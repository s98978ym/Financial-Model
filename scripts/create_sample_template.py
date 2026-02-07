#!/usr/bin/env python3
"""
Generate generic Excel templates for the Financial Model.

IMPORTANT: Templates must be BUSINESS-AGNOSTIC.
Sheet names, labels, and structure must use generic terms like
"セグメント1", "セグメント2" etc. -- NEVER use specific business names
(e.g. ミール, アカデミー, コンサル).  The LLM agents dynamically map
the user's actual business segments to these generic slots.

Creates:
  - templates/base.xlsx   (base-case financial model)
  - templates/worst.xlsx  (worst-case / conservative scenario)
"""

import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
ROOT_DIR = Path(__file__).resolve().parent.parent
TEMPLATES_DIR = ROOT_DIR / "templates"

FISCAL_YEARS = ["FY25", "FY26", "FY27", "FY28", "FY29"]
FY_COLS = list(range(2, 7))  # columns B-F  (index 2..6 in openpyxl)

# Number of generic revenue segments in the template
NUM_SEGMENTS = 3

YELLOW_FILL = PatternFill(
    patternType="solid",
    fgColor="FFFFF2CC",
)
HEADER_FILL = PatternFill(
    patternType="solid",
    fgColor="FF4472C4",
)
SUBHEADER_FILL = PatternFill(
    patternType="solid",
    fgColor="FFD9E2F3",
)
LIGHT_GRAY_FILL = PatternFill(
    patternType="solid",
    fgColor="FFF2F2F2",
)

TITLE_FONT = Font(name="Meiryo", size=14, bold=True, color="FFFFFFFF")
HEADER_FONT = Font(name="Meiryo", size=11, bold=True, color="FFFFFFFF")
SUBHEADER_FONT = Font(name="Meiryo", size=11, bold=True)
LABEL_FONT = Font(name="Meiryo", size=10)
NUMBER_FONT = Font(name="Meiryo", size=10)
FORMULA_FONT = Font(name="Meiryo", size=10, color="FF1F4E79")
PERCENT_FONT = Font(name="Meiryo", size=10, italic=True, color="FF1F4E79")

THIN_BORDER = Border(
    bottom=Side(style="thin", color="FFB4C6E7"),
)
BOTTOM_DOUBLE = Border(
    bottom=Side(style="double", color="FF4472C4"),
)

NUMBER_FMT = '#,##0'
CURRENCY_FMT = '#,##0'
PERCENT_FMT = '0.0%'


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _col(c):
    """Return Excel column letter (1-indexed)."""
    return get_column_letter(c)


def _set_col_widths(ws, widths: dict):
    for col_idx, w in widths.items():
        ws.column_dimensions[_col(col_idx)].width = w


def _write_title(ws, row, title, merge_end_col=6):
    """Write a merged title row with blue background."""
    ws.merge_cells(
        start_row=row, start_column=1,
        end_row=row, end_column=merge_end_col,
    )
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = TITLE_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 30


def _write_fy_headers(ws, row, start_col=2, label=""):
    """Write FY25-FY29 header row."""
    ws.cell(row=row, column=1, value=label).font = SUBHEADER_FONT
    for i, fy in enumerate(FISCAL_YEARS):
        c = ws.cell(row=row, column=start_col + i, value=fy)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")


def _write_label(ws, row, label, indent=0):
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = LABEL_FONT
    if indent:
        cell.alignment = Alignment(indent=indent)


def _write_input_row(ws, row, label, values, fmt=NUMBER_FMT, indent=0):
    """Write a row of hard-coded input values with yellow fill."""
    _write_label(ws, row, label, indent=indent)
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=FY_COLS[i], value=v)
        c.fill = YELLOW_FILL
        c.font = NUMBER_FONT
        c.number_format = fmt
        c.alignment = Alignment(horizontal="right")


def _write_formula_row(ws, row, label, formulas, fmt=NUMBER_FMT, indent=0,
                       is_percent=False, bold=False):
    """Write a row of Excel formulas."""
    _write_label(ws, row, label, indent=indent)
    for i, f in enumerate(formulas):
        c = ws.cell(row=row, column=FY_COLS[i])
        c.value = f
        c.font = FORMULA_FONT if not bold else Font(
            name="Meiryo", size=10, bold=True, color="FF1F4E79"
        )
        if is_percent:
            c.font = PERCENT_FONT
        c.number_format = fmt
        c.alignment = Alignment(horizontal="right")
    if bold:
        ws.cell(row=row, column=1).font = Font(
            name="Meiryo", size=10, bold=True
        )


def _apply_border_row(ws, row, border, max_col=6):
    for col in range(1, max_col + 1):
        ws.cell(row=row, column=col).border = border


def _seg_sheet_name(idx: int) -> str:
    """Generate a generic segment sheet name.

    Returns names like '収益モデル1', '収益モデル2', '収益モデル3'.
    These are intentionally business-agnostic.
    """
    return f"収益モデル{idx}"


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_pl_sheet(wb):
    """Sheet 1: PL設計 (P&L Design)

    Revenue line sums across generic segment sheets.
    """
    ws = wb.active
    ws.title = "PL設計"
    _set_col_widths(ws, {1: 22, 2: 16, 3: 16, 4: 16, 5: 16, 6: 16})

    # Title
    _write_title(ws, 1, "損益計算書（PL）")

    # Headers
    _write_fy_headers(ws, 2)
    ws.cell(row=2, column=1).value = "（単位：円）"
    ws.cell(row=2, column=1).font = Font(name="Meiryo", size=9, color="FFFFFFFF")
    ws.cell(row=2, column=1).fill = HEADER_FILL

    # Row 3: blank separator
    ws.row_dimensions[3].height = 6

    # Row 4: 売上高 = sum of all generic segment sheets row 10
    formulas = []
    for col in FY_COLS:
        cl = _col(col)
        refs = [f"'{_seg_sheet_name(i)}'!{cl}10" for i in range(1, NUM_SEGMENTS + 1)]
        formulas.append("=" + "+".join(refs))
    _write_formula_row(ws, 4, "売上高", formulas, bold=True)
    _apply_border_row(ws, 4, THIN_BORDER)

    # Row 5: 売上原価 (COGS) — simple % of revenue
    cogs_formulas = [f"={_col(c)}4*0.3" for c in FY_COLS]
    _write_formula_row(ws, 5, "売上原価", cogs_formulas)

    # Row 6: 粗利
    gp_formulas = [f"={_col(c)}4-{_col(c)}5" for c in FY_COLS]
    _write_formula_row(ws, 6, "粗利", gp_formulas, bold=True)
    _apply_border_row(ws, 6, BOTTOM_DOUBLE)

    # Row 7: 粗利率
    gp_rate = [f"=IF({_col(c)}4=0,0,{_col(c)}6/{_col(c)}4)" for c in FY_COLS]
    _write_formula_row(ws, 7, "粗利率", gp_rate, fmt=PERCENT_FMT, is_percent=True)

    # Row 8: blank
    ws.row_dimensions[8].height = 6

    # Row 9: 販管費合計
    sga_formulas = [f"='費用まとめ'!{_col(c)}12" for c in FY_COLS]
    _write_formula_row(ws, 9, "販管費合計", sga_formulas)

    # Row 10: 営業利益
    op_formulas = [f"={_col(c)}6-{_col(c)}9" for c in FY_COLS]
    _write_formula_row(ws, 10, "営業利益", op_formulas, bold=True)
    _apply_border_row(ws, 10, BOTTOM_DOUBLE)

    # Row 11: 営業利益率
    op_rate = [f"=IF({_col(c)}4=0,0,{_col(c)}10/{_col(c)}4)" for c in FY_COLS]
    _write_formula_row(ws, 11, "営業利益率", op_rate, fmt=PERCENT_FMT, is_percent=True)

    # Row 12: blank
    ws.row_dimensions[12].height = 6

    # Row 13: EBITDA (= operating profit + depreciation estimate)
    ebitda_formulas = [f"={_col(c)}10+500000" for c in FY_COLS]
    _write_formula_row(ws, 13, "EBITDA", ebitda_formulas, bold=True)

    # Row 14: EBITDA率
    ebitda_rate = [f"=IF({_col(c)}4=0,0,{_col(c)}13/{_col(c)}4)" for c in FY_COLS]
    _write_formula_row(ws, 14, "EBITDA率", ebitda_rate, fmt=PERCENT_FMT, is_percent=True)

    return ws


def build_segment_model(wb, seg_idx: int, scenario="base"):
    """Build a generic revenue segment sheet.

    Parameters
    ----------
    seg_idx : int
        1-based segment index (1, 2, 3, ...).
    scenario : str
        'base' or 'worst'.

    Each segment sheet has the SAME generic structure:
      Row 3: 顧客数/取引先数 (volume driver)
      Row 4: 単価 (price driver)
      Row 5: 頻度/回数 (frequency driver)
      Row 6: 成長率/解約率 (rate driver)
      Row 7: 月次売上 (formula)
      Row 8: 年間売上 (formula)
      Row 9: (blank)
      Row 10: 売上高 (reference for PL)

    Labels are intentionally generic so the LLM can map ANY business
    segment to them.
    """
    sheet_name = _seg_sheet_name(seg_idx)
    ws = wb.create_sheet(sheet_name)
    _set_col_widths(ws, {1: 24, 2: 16, 3: 16, 4: 16, 5: 16, 6: 16})

    _write_title(ws, 1, f"収益モデル{seg_idx}（セグメント{seg_idx}）")
    _write_fy_headers(ws, 2)

    # Generic sample data that varies by segment index and scenario
    if scenario == "base":
        volume = [v * (4 - seg_idx + 1) // 2 for v in [50, 100, 180, 280, 400]]
        price = [p + (seg_idx - 1) * 5000 for p in [5000, 5000, 5500, 5500, 6000]]
        frequency = [3, 3, 4, 4, 4]
        rate = [0.03, 0.03, 0.025, 0.025, 0.02]
    else:  # worst
        volume = [v * (4 - seg_idx + 1) // 3 for v in [30, 60, 100, 160, 240]]
        price = [p + (seg_idx - 1) * 3000 for p in [4000, 4000, 4500, 4500, 5000]]
        frequency = [2, 2, 3, 3, 3]
        rate = [0.05, 0.05, 0.04, 0.04, 0.035]

    # Row 3: 顧客数/取引数 (volume driver)
    _write_input_row(ws, 3, "顧客数/取引数", volume, fmt=NUMBER_FMT)

    # Row 4: 単価 (price per unit)
    _write_input_row(ws, 4, "単価（円）", price, fmt=CURRENCY_FMT)

    # Row 5: 頻度/回数 (frequency)
    _write_input_row(ws, 5, "頻度/回数（月間）", frequency, fmt='#,##0')

    # Row 6: 成長率/解約率 (rate)
    _write_input_row(ws, 6, "成長率/解約率", rate, fmt=PERCENT_FMT)

    # --- Formula rows ---
    # Row 7: 月次売上 = 顧客数 × 単価
    monthly_rev = [f"={_col(c)}3*{_col(c)}4" for c in FY_COLS]
    _write_formula_row(ws, 7, "月次売上", monthly_rev)

    # Row 8: 年間売上 = 月次売上 × 12
    annual_rev = [f"={_col(c)}7*12" for c in FY_COLS]
    _write_formula_row(ws, 8, "年間売上", annual_rev, bold=True)
    _apply_border_row(ws, 8, BOTTOM_DOUBLE)

    # Row 9: blank
    ws.row_dimensions[9].height = 6

    # Row 10: 売上高 (= Row 8, used as reference from PL)
    rev_ref = [f"={_col(c)}8" for c in FY_COLS]
    _write_formula_row(ws, 10, "売上高", rev_ref, bold=True)

    return ws


def build_cost_summary(wb):
    """Sheet: 費用まとめ (Cost Summary)"""
    ws = wb.create_sheet("費用まとめ")
    _set_col_widths(ws, {1: 24, 2: 16, 3: 16, 4: 16, 5: 16, 6: 16})

    _write_title(ws, 1, "費用まとめ（SGA）")
    _write_fy_headers(ws, 2)

    cost_list = "費用リスト"

    # Row 3: blank separator
    ws.row_dimensions[3].height = 6

    # Row 4: 人件費合計
    hr = [f"='{cost_list}'!{_col(c)}8" for c in FY_COLS]
    _write_formula_row(ws, 4, "人件費合計", hr)

    # Row 5: オフィス費
    office = [f"='{cost_list}'!{_col(c)}13" for c in FY_COLS]
    _write_formula_row(ws, 5, "オフィス費", office)

    # Row 6: マーケティング費
    mkt = [f"='{cost_list}'!{_col(c)}18" for c in FY_COLS]
    _write_formula_row(ws, 6, "マーケティング費", mkt)

    # Row 7: システム費
    sys_cost = [f"='{cost_list}'!{_col(c)}23" for c in FY_COLS]
    _write_formula_row(ws, 7, "システム費", sys_cost)

    # Row 8: その他
    other = [f"='{cost_list}'!{_col(c)}27" for c in FY_COLS]
    _write_formula_row(ws, 8, "その他", other)

    # Row 9: blank
    ws.row_dimensions[9].height = 6

    # Row 10: separator line
    _apply_border_row(ws, 9, THIN_BORDER)

    # Row 10: 販管費小計
    subtotal = [
        f"=SUM({_col(c)}4:{_col(c)}8)" for c in FY_COLS
    ]
    _write_formula_row(ws, 10, "販管費小計", subtotal, bold=True)

    # Row 11: 予備費 (contingency 5%)
    contingency = [f"={_col(c)}10*0.05" for c in FY_COLS]
    _write_formula_row(ws, 11, "予備費（5%）", contingency)

    # Row 12: 販管費合計  (this is what PL references)
    total = [f"={_col(c)}10+{_col(c)}11" for c in FY_COLS]
    _write_formula_row(ws, 12, "販管費合計", total, bold=True)
    _apply_border_row(ws, 12, BOTTOM_DOUBLE)

    return ws


def build_cost_list(wb, scenario="base"):
    """Sheet: 費用リスト (Cost List - detailed)"""
    ws = wb.create_sheet("費用リスト")
    _set_col_widths(ws, {1: 28, 2: 16, 3: 16, 4: 16, 5: 16, 6: 16})

    _write_title(ws, 1, "費用リスト（詳細）")
    _write_fy_headers(ws, 2)

    multiplier = 1.0 if scenario == "base" else 1.15  # worst: 15% higher costs

    # --- 人件費 section ---
    ws.cell(row=3, column=1, value="【人件費】").font = SUBHEADER_FONT
    for c in range(1, 7):
        ws.cell(row=3, column=c).fill = SUBHEADER_FILL

    # Row 4: 役員報酬
    exec_comp = [int(v * multiplier) for v in [6000000, 7200000, 8400000, 9600000, 12000000]]
    _write_input_row(ws, 4, "役員報酬", exec_comp, indent=1)

    # Row 5: エンジニア人件費
    eng = [int(v * multiplier) for v in [4800000, 9600000, 16800000, 24000000, 33600000]]
    _write_input_row(ws, 5, "エンジニア人件費", eng, indent=1)

    # Row 6: 営業人件費
    sales = [int(v * multiplier) for v in [3600000, 7200000, 10800000, 14400000, 21600000]]
    _write_input_row(ws, 6, "営業人件費", sales, indent=1)

    # Row 7: その他人件費
    other_hr = [int(v * multiplier) for v in [1200000, 2400000, 3600000, 4800000, 7200000]]
    _write_input_row(ws, 7, "その他人件費", other_hr, indent=1)

    # Row 8: 人件費小計
    hr_sub = [f"=SUM({_col(c)}4:{_col(c)}7)" for c in FY_COLS]
    _write_formula_row(ws, 8, "人件費小計", hr_sub, bold=True)
    _apply_border_row(ws, 8, THIN_BORDER)

    # --- オフィス section ---
    ws.cell(row=9, column=1, value="【オフィス】").font = SUBHEADER_FONT
    for c in range(1, 7):
        ws.cell(row=9, column=c).fill = SUBHEADER_FILL

    # Row 10: 家賃
    rent = [int(v * multiplier) for v in [1200000, 1200000, 2400000, 2400000, 3600000]]
    _write_input_row(ws, 10, "家賃", rent, indent=1)

    # Row 11: 光熱費
    util = [int(v * multiplier) for v in [120000, 144000, 240000, 288000, 360000]]
    _write_input_row(ws, 11, "光熱費", util, indent=1)

    # Row 12: 備品・消耗品
    supplies = [int(v * multiplier) for v in [180000, 240000, 360000, 480000, 600000]]
    _write_input_row(ws, 12, "備品・消耗品", supplies, indent=1)

    # Row 13: オフィス小計
    off_sub = [f"=SUM({_col(c)}10:{_col(c)}12)" for c in FY_COLS]
    _write_formula_row(ws, 13, "オフィス小計", off_sub, bold=True)
    _apply_border_row(ws, 13, THIN_BORDER)

    # --- マーケティング section ---
    ws.cell(row=14, column=1, value="【マーケティング】").font = SUBHEADER_FONT
    for c in range(1, 7):
        ws.cell(row=14, column=c).fill = SUBHEADER_FILL

    # Row 15: 広告費
    ads = [int(v * multiplier) for v in [600000, 1200000, 2400000, 3600000, 4800000]]
    _write_input_row(ws, 15, "広告費", ads, indent=1)

    # Row 16: PR費
    pr = [int(v * multiplier) for v in [240000, 480000, 720000, 960000, 1200000]]
    _write_input_row(ws, 16, "PR費", pr, indent=1)

    # Row 17: イベント費
    events = [int(v * multiplier) for v in [120000, 240000, 480000, 720000, 960000]]
    _write_input_row(ws, 17, "イベント費", events, indent=1)

    # Row 18: マーケティング小計
    mkt_sub = [f"=SUM({_col(c)}15:{_col(c)}17)" for c in FY_COLS]
    _write_formula_row(ws, 18, "マーケティング小計", mkt_sub, bold=True)
    _apply_border_row(ws, 18, THIN_BORDER)

    # --- システム section ---
    ws.cell(row=19, column=1, value="【システム】").font = SUBHEADER_FONT
    for c in range(1, 7):
        ws.cell(row=19, column=c).fill = SUBHEADER_FILL

    # Row 20: サーバー費
    server = [int(v * multiplier) for v in [360000, 600000, 960000, 1440000, 1800000]]
    _write_input_row(ws, 20, "サーバー費", server, indent=1)

    # Row 21: ツール費
    tools = [int(v * multiplier) for v in [240000, 360000, 480000, 720000, 960000]]
    _write_input_row(ws, 21, "ツール費", tools, indent=1)

    # Row 22: 外注開発費
    outsource = [int(v * multiplier) for v in [600000, 1200000, 1800000, 2400000, 3000000]]
    _write_input_row(ws, 22, "外注開発費", outsource, indent=1)

    # Row 23: システム小計
    sys_sub = [f"=SUM({_col(c)}20:{_col(c)}22)" for c in FY_COLS]
    _write_formula_row(ws, 23, "システム小計", sys_sub, bold=True)
    _apply_border_row(ws, 23, THIN_BORDER)

    # --- その他 section ---
    ws.cell(row=24, column=1, value="【その他】").font = SUBHEADER_FONT
    for c in range(1, 7):
        ws.cell(row=24, column=c).fill = SUBHEADER_FILL

    # Row 25: 交通費
    travel = [int(v * multiplier) for v in [180000, 360000, 540000, 720000, 960000]]
    _write_input_row(ws, 25, "交通費", travel, indent=1)

    # Row 26: 顧問・士業費
    advisory = [int(v * multiplier) for v in [600000, 600000, 900000, 1200000, 1500000]]
    _write_input_row(ws, 26, "顧問・士業費", advisory, indent=1)

    # Row 27: その他小計
    other_sub = [f"=SUM({_col(c)}25:{_col(c)}26)" for c in FY_COLS]
    _write_formula_row(ws, 27, "その他小計", other_sub, bold=True)
    _apply_border_row(ws, 27, BOTTOM_DOUBLE)

    return ws


def build_cost_tags(wb):
    """Sheet: 費用タグ (Cost Tags)"""
    ws = wb.create_sheet("費用タグ")
    _set_col_widths(ws, {1: 28, 2: 20, 3: 20})

    _write_title(ws, 1, "費用タグ（カテゴリ分類）", merge_end_col=3)

    # Header row
    for i, h in enumerate(["費用項目", "カテゴリ", "タグ"], start=1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")

    tags = [
        ("役員報酬",      "人件費", "固定費"),
        ("エンジニア人件費", "人件費", "固定費"),
        ("営業人件費",     "人件費", "固定費"),
        ("その他人件費",   "人件費", "固定費"),
        ("家賃",         "オフィス", "固定費"),
        ("光熱費",       "オフィス", "固定費"),
        ("備品・消耗品",   "オフィス", "変動費"),
        ("広告費",       "マーケティング", "変動費"),
        ("PR費",        "マーケティング", "変動費"),
        ("イベント費",    "マーケティング", "変動費"),
        ("サーバー費",    "システム", "変動費"),
        ("ツール費",     "システム", "固定費"),
        ("外注開発費",    "システム", "変動費"),
        ("交通費",       "その他", "変動費"),
        ("顧問・士業費",   "その他", "固定費"),
    ]

    for row_idx, (item, cat, tag) in enumerate(tags, start=3):
        ws.cell(row=row_idx, column=1, value=item).font = LABEL_FONT
        ws.cell(row=row_idx, column=2, value=cat).font = LABEL_FONT
        c = ws.cell(row=row_idx, column=3, value=tag)
        c.font = LABEL_FONT
        if tag == "固定費":
            c.fill = PatternFill(patternType="solid", fgColor="FFD6EAF8")
        else:
            c.fill = PatternFill(patternType="solid", fgColor="FFFCE4D6")

    return ws


def build_assumptions(wb, scenario="base"):
    """Sheet: （全Ver）前提条件 (All Versions - Assumptions)

    Assumptions are generic -- no references to any specific industry.
    """
    ws = wb.create_sheet("（全Ver）前提条件")
    _set_col_widths(ws, {1: 30, 2: 18, 3: 40})

    _write_title(ws, 1, "前提条件・マクロ仮定", merge_end_col=3)

    # Header
    for i, h in enumerate(["項目", "値", "備考"], start=1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")

    if scenario == "base":
        assumptions = [
            ("市場規模（TAM）",        "500億円",  "対象市場全体"),
            ("市場成長率（年率）",       "15%",     "業界レポート参照"),
            ("ターゲットシェア（5年後）", "0.5%",    "保守的見積もり"),
            ("顧客獲得コスト（CAC）",    "15,000円", "広告・営業含む"),
            ("平均契約期間",           "18ヶ月",   "過去実績ベース"),
            ("売上原価率",            "30%",      "変動費ベース"),
            ("人件費成長率",           "20%/年",   "採用計画連動"),
            ("為替前提",              "150円/$",  "経営計画ベース"),
            ("消費税率",              "10%",      "現行税制"),
            ("法人税実効税率",          "30%",      "中小企業前提"),
            ("割引率（WACC）",         "10%",      "DCF計算用"),
            ("ターミナル成長率",        "2%",       "永続成長前提"),
        ]
    else:
        assumptions = [
            ("市場規模（TAM）",        "400億円",  "縮小シナリオ"),
            ("市場成長率（年率）",       "8%",      "低成長前提"),
            ("ターゲットシェア（5年後）", "0.3%",    "保守的見積もり"),
            ("顧客獲得コスト（CAC）",    "22,000円", "広告効率悪化"),
            ("平均契約期間",           "12ヶ月",   "解約率上昇"),
            ("売上原価率",            "35%",      "原材料高騰"),
            ("人件費成長率",           "25%/年",   "人材確保コスト増"),
            ("為替前提",              "160円/$",  "円安進行"),
            ("消費税率",              "10%",      "現行税制"),
            ("法人税実効税率",          "30%",      "中小企業前提"),
            ("割引率（WACC）",         "12%",      "リスクプレミアム増"),
            ("ターミナル成長率",        "1%",       "保守的永続成長"),
        ]

    for row_idx, (item, value, note) in enumerate(assumptions, start=3):
        ws.cell(row=row_idx, column=1, value=item).font = LABEL_FONT
        vc = ws.cell(row=row_idx, column=2, value=value)
        vc.font = NUMBER_FONT
        vc.fill = YELLOW_FILL
        vc.alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=3, value=note).font = Font(
            name="Meiryo", size=9, color="FF808080"
        )

    return ws


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def create_workbook(scenario="base"):
    """Build a complete workbook for the given scenario."""
    wb = Workbook()

    # Sheet 1: PL設計
    build_pl_sheet(wb)

    # Sheets 2-4: Generic revenue segment models
    for seg_idx in range(1, NUM_SEGMENTS + 1):
        build_segment_model(wb, seg_idx, scenario)

    # Sheet 5: 費用まとめ
    build_cost_summary(wb)

    # Sheet 6: 費用リスト
    build_cost_list(wb, scenario)

    # Sheet 7: 費用タグ
    build_cost_tags(wb)

    # Sheet 8: （全Ver）前提条件
    build_assumptions(wb, scenario)

    return wb


def main():
    TEMPLATES_DIR.mkdir(parents=True, exist_ok=True)

    # --- Base case ---
    base_path = TEMPLATES_DIR / "base.xlsx"
    wb_base = create_workbook(scenario="base")
    wb_base.save(str(base_path))
    print(f"[OK] Created {base_path}")

    # --- Worst case ---
    worst_path = TEMPLATES_DIR / "worst.xlsx"
    wb_worst = create_workbook(scenario="worst")
    wb_worst.save(str(worst_path))
    print(f"[OK] Created {worst_path}")

    # Summary
    print()
    print("=== Template Summary ===")
    for path in [base_path, worst_path]:
        print(f"  {path.name}:")
        from openpyxl import load_workbook
        wb = load_workbook(str(path))
        for name in wb.sheetnames:
            ws = wb[name]
            print(f"    - {name:20s}  ({ws.max_row} rows x {ws.max_column} cols)")
        wb.close()


if __name__ == "__main__":
    main()
