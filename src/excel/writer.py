"""Excel writer that preserves formulas and only modifies input cells."""
import shutil
import logging
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
from copy import copy

import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

from ..config.models import (
    ExtractedParameter, CellTarget, PhaseAConfig, GenerationConfig
)

logger = logging.getLogger(__name__)


class PLWriter:
    """Writes extracted parameters to Excel template, preserving all formulas."""

    def __init__(self, template_path: str, output_path: str, config: PhaseAConfig):
        self.template_path = Path(template_path)
        self.output_path = Path(output_path)
        self.config = config
        self.wb = None
        self.change_log: List[Dict[str, Any]] = []
        self.skipped_log: List[Dict[str, Any]] = []

    def generate(self, parameters: List[ExtractedParameter]) -> str:
        """
        Generate PL Excel from template with extracted parameters.

        Steps:
        1. Copy template to output path
        2. Open copy with openpyxl (data_only=False to preserve formulas)
        3. For each parameter with mapped_targets:
           a. Check cell is input cell (correct color) and NOT a formula
           b. Write the value
           c. Log the change
        4. Apply optional color styling
        5. Set fullCalcOnLoad = True
        6. Save
        """
        # Step 1: Copy template
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(str(self.template_path), str(self.output_path))

        # Step 2: Open workbook
        self.wb = openpyxl.load_workbook(str(self.output_path))

        # Step 3: Write parameters
        for param in parameters:
            if not param.selected:
                continue

            value_to_write = param.adjusted_value if param.adjusted_value is not None else param.value

            for target in param.mapped_targets:
                self._write_cell(target, value_to_write, param)

        # Step 4: Apply optional styling
        self._apply_styling()

        # Step 5: Set fullCalcOnLoad
        self.wb.calculation = openpyxl.workbook.properties.CalcProperties(fullCalcOnLoad=True)

        # Step 6: Save
        self.wb.save(str(self.output_path))
        logger.info(f"Generated PL saved to {self.output_path}")

        return str(self.output_path)

    def _write_cell(
        self, target: CellTarget, value: Any, param: ExtractedParameter
    ) -> bool:
        """Write a value to a specific cell, with safety checks."""
        try:
            if target.sheet not in self.wb.sheetnames:
                logger.warning(f"Sheet '{target.sheet}' not found, skipping {target.cell}")
                self.skipped_log.append({
                    "sheet": target.sheet, "cell": target.cell,
                    "reason": "sheet_not_found", "param_key": param.key
                })
                return False

            ws = self.wb[target.sheet]
            cell = ws[target.cell]

            # CRITICAL: Never overwrite formula cells
            if cell.data_type == 'f' or (isinstance(cell.value, str) and str(cell.value).startswith('=')):
                logger.warning(
                    f"BLOCKED: Cell {target.sheet}!{target.cell} contains formula "
                    f"'{cell.value}', skipping write"
                )
                self.skipped_log.append({
                    "sheet": target.sheet, "cell": target.cell,
                    "reason": "formula_cell", "param_key": param.key,
                    "formula": str(cell.value)
                })
                return False

            # Check if cell has the expected input color
            if not self._is_input_cell(cell):
                logger.warning(
                    f"Cell {target.sheet}!{target.cell} does not have input color, "
                    f"writing anyway but flagging"
                )

            # Record the change
            old_value = cell.value
            cell.value = value

            self.change_log.append({
                "sheet": target.sheet,
                "cell": target.cell,
                "old_value": old_value,
                "new_value": value,
                "param_key": param.key,
                "confidence": param.confidence,
                "source": param.source,
            })

            logger.debug(
                f"Written {target.sheet}!{target.cell}: {old_value} -> {value} "
                f"(param={param.key}, confidence={param.confidence})"
            )
            return True

        except Exception as e:
            logger.error(f"Error writing to {target.sheet}!{target.cell}: {e}")
            self.skipped_log.append({
                "sheet": target.sheet, "cell": target.cell,
                "reason": "error", "error": str(e), "param_key": param.key
            })
            return False

    def _is_input_cell(self, cell) -> bool:
        """Check if a cell has the configured input color."""
        try:
            fill = cell.fill
            if fill and fill.fgColor and fill.fgColor.rgb:
                color = str(fill.fgColor.rgb)
                target = self.config.colors.input_color
                # Handle various color format differences
                return (
                    color == target or
                    color.lstrip("0") == target.lstrip("0") or
                    color[-6:] == target[-6:]
                )
        except Exception:
            pass
        return False

    def _apply_styling(self):
        """Apply optional color styling (formula cells, total cells)."""
        if not self.config.colors.apply_formula_color and not self.config.colors.apply_total_color:
            return

        formula_font = Font(color=self.config.colors.formula_color[2:])  # strip FF prefix
        total_fill = PatternFill(
            start_color=self.config.colors.total_color,
            end_color=self.config.colors.total_color,
            fill_type="solid"
        )

        for ws in self.wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue

                    # Apply formula color
                    if self.config.colors.apply_formula_color:
                        if cell.data_type == 'f' or (isinstance(cell.value, str) and str(cell.value).startswith('=')):
                            cell.font = copy(cell.font)
                            cell.font = Font(
                                name=cell.font.name,
                                size=cell.font.size,
                                bold=cell.font.bold,
                                italic=cell.font.italic,
                                color=self.config.colors.formula_color[2:]
                            )

                    # Apply total color
                    if self.config.colors.apply_total_color:
                        if self._is_total_cell(cell, ws):
                            cell.fill = total_fill

    def _is_total_cell(self, cell, ws) -> bool:
        """Detect if a cell is a total/subtotal cell."""
        # Check if formula contains SUM
        if isinstance(cell.value, str) and "SUM" in str(cell.value).upper():
            return True

        # Check if label to the left contains total keywords
        if cell.column > 1:
            label_cell = ws.cell(row=cell.row, column=cell.column - 1)
            if label_cell.value and isinstance(label_cell.value, str):
                label = label_cell.value.lower()
                if any(kw in label for kw in ["合計", "total", "計", "小計", "subtotal"]):
                    return True

        return False

    def get_change_summary(self) -> Dict[str, Any]:
        """Return summary of all changes made."""
        return {
            "total_changes": len(self.change_log),
            "total_skipped": len(self.skipped_log),
            "changes": self.change_log,
            "skipped": self.skipped_log,
        }
