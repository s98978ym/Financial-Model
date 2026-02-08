"""V2 Excel template builder – Investment-bank grade PL model.

Generates a 3-layer workbook:
  Layer A: シミュレーション分析  (Simulation / target comparison)
  Layer B: PL設計               (5-year PL with OPEX/CAPEX/FCF)
  Layer C: セグメント別モデル    (N segment revenue model sheets)

All input cells are highlighted yellow (FFFFF2CC) for scanner detection.
Cross-sheet references use direct cell references for robustness.
Named ranges are created for key PL KPIs so the simulation sheet
is decoupled from exact row positions.

IMPORTANT: This template is **business-agnostic**.  Sheet names use
generic labels; the LLM agents dynamically map real business data.
"""
from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Style constants (shared across all sheets)
# ---------------------------------------------------------------------------
YELLOW_FILL = PatternFill(patternType="solid", fgColor="FFFFF2CC")
HEADER_FILL = PatternFill(patternType="solid", fgColor="FF4472C4")
SUBHEADER_FILL = PatternFill(patternType="solid", fgColor="FFD9E2F3")
SECTION_FILL = PatternFill(patternType="solid", fgColor="FFF2F2F2")

TITLE_FONT = Font(name="Meiryo", size=14, bold=True, color="FFFFFFFF")
HEADER_FONT = Font(name="Meiryo", size=11, bold=True, color="FFFFFFFF")
SUBHEADER_FONT = Font(name="Meiryo", size=11, bold=True)
SECTION_FONT = Font(name="Meiryo", size=11, bold=True, color="FF4472C4")
LABEL_FONT = Font(name="Meiryo", size=10)
NUMBER_FONT = Font(name="Meiryo", size=10)
FORMULA_FONT = Font(name="Meiryo", size=10, color="FF1F4E79")
FORMULA_BOLD = Font(name="Meiryo", size=10, bold=True, color="FF1F4E79")
PERCENT_FONT = Font(name="Meiryo", size=10, italic=True, color="FF1F4E79")
BOLD_FONT = Font(name="Meiryo", size=10, bold=True)

THIN_BORDER = Border(bottom=Side(style="thin", color="FFB4C6E7"))
BOTTOM_DOUBLE = Border(bottom=Side(style="double", color="FF4472C4"))

NUMBER_FMT = '#,##0'
CURRENCY_FMT = '#,##0'
PERCENT_FMT = '0.0%'
MILLION_FMT = '#,##0,,"百万"'

# Default FY labels (overridden if user specifies start year)
DEFAULT_FY_LABELS = ["FY1", "FY2", "FY3", "FY4", "FY5"]
FY_COLS = [2, 3, 4, 5, 6]  # columns B-F

# Max revenue streams per segment sheet
MAX_STREAMS = 5

# Scenario parameter variables (generic, business-agnostic)
SCENARIO_VARS = [
    "売上倍率",
    "価格変動率",
    "数量変動率",
    "売上原価率調整",
    "人件費変動率",
    "マーケ費変動率",
    "その他SGA変動率",
    "CAPEX変動率",
]

# Default scenario multiplier values
DEFAULT_SCENARIO_VALUES = {
    "worst": [0.8, 0.9, 0.85, 1.1, 1.15, 0.7, 1.1, 1.2],
    "base":  [1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0],
    "best":  [1.2, 1.1, 1.15, 0.9, 0.95, 1.3, 0.9, 0.8],
}

# PL row positions (0-indexed offset from PL data start)
# These are the ROW NUMBERS in the PL設計 sheet
PL_ROWS = {
    "revenue":       5,
    "cogs_rate":     6,
    "cogs":          7,
    "gross_profit":  8,
    "gp_rate":       9,
    # OPEX section
    "opex_header":  11,
    "payroll":      12,
    "marketing":    13,
    "office":       14,
    "system":       15,
    "other_opex":   16,
    "opex_total":   17,
    # Depreciation
    "depr_header":  19,
    "depreciation": 20,
    # CAPEX
    "capex_header": 22,
    "capex":        23,
    # P&L
    "pl_header":    25,
    "op":           26,
    "op_rate":      27,
    # Cash flow
    "cf_header":    29,
    "fcf":          30,
    "cum_fcf":      31,
    # Funding
    "fund_header":  33,
    "funding_need": 34,
}

# Simulation sheet row positions
SIM_ROWS = {
    "scenario_select": 3,
    "target_header":   5,
    "target_fy_hdr":   6,
    "target_rev":      7,
    "target_op":       8,
    "target_sy_be":    9,
    "target_cum_be":  10,
    # Model vs Target
    "mvt_header":     12,
    "mvt_fy_hdr":     13,
    "model_rev":      14,
    "tgt_rev":        15,
    "gap_rev":        16,
    "pct_rev":        17,
    "model_op":       18,
    "tgt_op":         19,
    "gap_op":         20,
    "pct_op":         21,
    "model_sy_be":    22,
    "model_cum_be":   23,
    # Scenario parameters
    "sp_header":      25,
    "worst_header":   26,
    "worst_fy_hdr":   27,
    "worst_start":    28,  # rows 28-35
    "base_header":    37,
    "base_fy_hdr":    38,
    "base_start":     39,  # rows 39-46
    "best_header":    48,
    "best_fy_hdr":    49,
    "best_start":     50,  # rows 50-57
    # Simulation results
    "sr_header":      59,
    "sr_fy_hdr":      60,
    "sr_base_rev":    61,
    "sr_sim_rev":     62,
    "sr_base_gp":     63,
    "sr_sim_gp":      64,
    "sr_base_opex":   65,
    "sr_sim_opex":    66,
    "sr_base_depr":   67,
    "sr_sim_depr":    68,
    "sr_base_capex":  69,
    "sr_sim_capex":   70,
    "sr_base_op":     71,
    "sr_sim_op":      72,
    "sr_gp_rate":     73,
    "sr_op_rate":     74,
    # Funding simulator
    "fs_header":      76,
    "fs_fy_hdr":      77,
    "fs_gp":          78,
    "fs_need":        79,
    "fs_fcf":         80,
    "fs_cum_fcf":     81,
    "fs_buf_header":  83,
    "fs_buf_cols":    84,
    "fs_monthly_opex":85,
    "fs_buffer_amt":  86,
    "fs_cum_min":     87,
    "fs_min_year":    88,
    "fs_funding":     89,
}

# Segment sheet row positions per stream
SEG_STREAM_ROWS_PER = 7  # name, price, volume, freq, rate, revenue, blank
SEG_ROWS = {
    "header_fy": 3,
    "stream_start": 4,  # first stream starts here
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _col(c: int) -> str:
    return get_column_letter(c)


def _set_col_widths(ws, widths: Dict[int, float]) -> None:
    for col_idx, w in widths.items():
        ws.column_dimensions[_col(col_idx)].width = w


def _write_title(ws, row: int, title: str, merge_end_col: int = 6) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=merge_end_col)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = TITLE_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 30


def _write_fy_headers(ws, row: int, fy_labels: List[str], label: str = "") -> None:
    ws.cell(row=row, column=1, value=label).font = SUBHEADER_FONT
    for i, fy in enumerate(fy_labels):
        c = ws.cell(row=row, column=FY_COLS[i], value=fy)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")


def _write_section_header(ws, row: int, title: str) -> None:
    ws.cell(row=row, column=1, value=title).font = SECTION_FONT
    for c in range(1, 7):
        ws.cell(row=row, column=c).fill = SECTION_FILL


def _write_label(ws, row: int, label: str, indent: int = 0) -> None:
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = LABEL_FONT
    if indent:
        cell.alignment = Alignment(indent=indent)


def _write_input_row(
    ws, row: int, label: str, values: List, fmt: str = NUMBER_FMT, indent: int = 0,
) -> None:
    _write_label(ws, row, label, indent=indent)
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=FY_COLS[i], value=v)
        c.fill = YELLOW_FILL
        c.font = NUMBER_FONT
        c.number_format = fmt
        c.alignment = Alignment(horizontal="right")


def _write_formula_row(
    ws, row: int, label: str, formulas: List[str], fmt: str = NUMBER_FMT,
    indent: int = 0, is_percent: bool = False, bold: bool = False,
) -> None:
    _write_label(ws, row, label, indent=indent)
    for i, f in enumerate(formulas):
        c = ws.cell(row=row, column=FY_COLS[i])
        c.value = f
        if bold:
            c.font = FORMULA_BOLD
        elif is_percent:
            c.font = PERCENT_FONT
        else:
            c.font = FORMULA_FONT
        c.number_format = fmt
        c.alignment = Alignment(horizontal="right")
    if bold:
        ws.cell(row=row, column=1).font = BOLD_FONT


def _apply_border_row(ws, row: int, border, max_col: int = 6) -> None:
    for col in range(1, max_col + 1):
        ws.cell(row=row, column=col).border = border


def _separator_row(ws, row: int) -> None:
    ws.row_dimensions[row].height = 6


def _sanitize_sheet_name(name: str) -> str:
    """Sanitize a sheet name for Excel (max 31 chars, no special chars)."""
    name = re.sub(r'[\\/*?\[\]:]', '_', name)
    if len(name) > 28:
        name = name[:28]
    return name


def _seg_sheet_name(idx: int) -> str:
    """Generic segment sheet name."""
    return f"セグメント{idx}モデル"


# ---------------------------------------------------------------------------
# Named range registration
# ---------------------------------------------------------------------------
def _add_named_range(wb: Workbook, name: str, sheet: str, cell: str) -> None:
    """Add a workbook-level named range."""
    dn = DefinedName(name, attr_text=f"'{sheet}'!${cell}")
    wb.defined_names.add(dn)


def _register_pl_named_ranges(wb: Workbook, num_fys: int = 5) -> None:
    """Register named ranges for PL KPI rows."""
    sheet = "PL設計"
    ranges = {
        "PL_REVENUE":      PL_ROWS["revenue"],
        "PL_COGS_RATE":    PL_ROWS["cogs_rate"],
        "PL_COGS":         PL_ROWS["cogs"],
        "PL_GP":           PL_ROWS["gross_profit"],
        "PL_OPEX":         PL_ROWS["opex_total"],
        "PL_PAYROLL":      PL_ROWS["payroll"],
        "PL_MARKETING":    PL_ROWS["marketing"],
        "PL_OFFICE":       PL_ROWS["office"],
        "PL_SYSTEM":       PL_ROWS["system"],
        "PL_OTHER_OPEX":   PL_ROWS["other_opex"],
        "PL_DEPR":         PL_ROWS["depreciation"],
        "PL_CAPEX":        PL_ROWS["capex"],
        "PL_OP":           PL_ROWS["op"],
        "PL_FCF":          PL_ROWS["fcf"],
        "PL_CUM_FCF":      PL_ROWS["cum_fcf"],
    }
    for prefix, row in ranges.items():
        for i, col in enumerate(FY_COLS[:num_fys]):
            cl = _col(col)
            _add_named_range(wb, f"{prefix}_FY{i+1}", sheet, f"{cl}${row}")


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_pl_sheet(
    wb: Workbook,
    num_segments: int = 3,
    fy_labels: Optional[List[str]] = None,
) -> None:
    """Build PL設計 sheet (Layer B).

    Revenue = SUM of all segment total rows.
    Includes OPEX, CAPEX, depreciation, FCF, cumulative FCF, funding need.
    """
    fy = fy_labels or DEFAULT_FY_LABELS
    ws = wb.active
    ws.title = "PL設計"
    _set_col_widths(ws, {1: 24, 2: 16, 3: 16, 4: 16, 5: 16, 6: 16})

    # Title
    _write_title(ws, 1, "PL設計（5年計画）")

    # Row 2: unit + FY headers
    _write_fy_headers(ws, 2, fy, label="（単位：円）")
    ws.cell(row=2, column=1).font = Font(name="Meiryo", size=9, color="FFFFFFFF")
    ws.cell(row=2, column=1).fill = HEADER_FILL

    # Row 3: separator
    _separator_row(ws, 3)

    # ---------- 【売上】 ----------
    _write_section_header(ws, 4, "【売上】")
    r = PL_ROWS

    # Row 5: 売上高 = SUM of segment totals (last row of each segment sheet)
    seg_total_row = SEG_ROWS["stream_start"] + MAX_STREAMS * SEG_STREAM_ROWS_PER + 1
    rev_formulas = []
    for col in FY_COLS:
        cl = _col(col)
        refs = [f"'{_seg_sheet_name(i)}'!{cl}{seg_total_row}" for i in range(1, num_segments + 1)]
        rev_formulas.append("=" + "+".join(refs))
    _write_formula_row(ws, r["revenue"], "売上高", rev_formulas, bold=True)
    _apply_border_row(ws, r["revenue"], THIN_BORDER)

    # Row 6: 売上原価率 (INPUT)
    _write_input_row(ws, r["cogs_rate"], "売上原価率", [0.3] * 5, fmt=PERCENT_FMT)

    # Row 7: 売上原価 = 売上高 × 原価率
    cogs_f = [f"={_col(c)}{r['revenue']}*{_col(c)}{r['cogs_rate']}" for c in FY_COLS]
    _write_formula_row(ws, r["cogs"], "売上原価", cogs_f)

    # Row 8: 粗利
    gp_f = [f"={_col(c)}{r['revenue']}-{_col(c)}{r['cogs']}" for c in FY_COLS]
    _write_formula_row(ws, r["gross_profit"], "粗利", gp_f, bold=True)
    _apply_border_row(ws, r["gross_profit"], BOTTOM_DOUBLE)

    # Row 9: 粗利率
    gp_rate_f = [f"=IF({_col(c)}{r['revenue']}=0,0,{_col(c)}{r['gross_profit']}/{_col(c)}{r['revenue']})" for c in FY_COLS]
    _write_formula_row(ws, r["gp_rate"], "粗利率", gp_rate_f, fmt=PERCENT_FMT, is_percent=True)

    # Row 10: separator
    _separator_row(ws, 10)

    # ---------- 【OPEX】 ----------
    _write_section_header(ws, r["opex_header"], "【OPEX（事業運営費）】")

    _write_input_row(ws, r["payroll"], "人件費", [0] * 5, indent=1)
    _write_input_row(ws, r["marketing"], "マーケティング費", [0] * 5, indent=1)
    _write_input_row(ws, r["office"], "オフィス・一般管理費", [0] * 5, indent=1)
    _write_input_row(ws, r["system"], "システム・開発費", [0] * 5, indent=1)
    _write_input_row(ws, r["other_opex"], "その他OPEX", [0] * 5, indent=1)

    # Row 17: OPEX合計
    opex_f = [f"=SUM({_col(c)}{r['payroll']}:{_col(c)}{r['other_opex']})" for c in FY_COLS]
    _write_formula_row(ws, r["opex_total"], "OPEX合計", opex_f, bold=True)
    _apply_border_row(ws, r["opex_total"], BOTTOM_DOUBLE)

    # Row 18: separator
    _separator_row(ws, 18)

    # ---------- 【減価償却費】 ----------
    _write_section_header(ws, r["depr_header"], "【減価償却費】")
    _write_input_row(ws, r["depreciation"], "減価償却費", [0] * 5, indent=1)

    # Row 21: separator
    _separator_row(ws, 21)

    # ---------- 【CAPEX】 ----------
    _write_section_header(ws, r["capex_header"], "【CAPEX（設備投資）】")
    _write_input_row(ws, r["capex"], "CAPEX", [0] * 5, indent=1)

    # Row 24: separator
    _separator_row(ws, 24)

    # ---------- 【損益】 ----------
    _write_section_header(ws, r["pl_header"], "【損益】")

    # Row 26: 営業利益 = 粗利 - OPEX合計 - 減価償却
    op_f = [f"={_col(c)}{r['gross_profit']}-{_col(c)}{r['opex_total']}-{_col(c)}{r['depreciation']}" for c in FY_COLS]
    _write_formula_row(ws, r["op"], "営業利益", op_f, bold=True)
    _apply_border_row(ws, r["op"], BOTTOM_DOUBLE)

    # Row 27: 営業利益率
    op_rate_f = [f"=IF({_col(c)}{r['revenue']}=0,0,{_col(c)}{r['op']}/{_col(c)}{r['revenue']})" for c in FY_COLS]
    _write_formula_row(ws, r["op_rate"], "営業利益率", op_rate_f, fmt=PERCENT_FMT, is_percent=True)

    # Row 28: separator
    _separator_row(ws, 28)

    # ---------- 【キャッシュフロー】 ----------
    _write_section_header(ws, r["cf_header"], "【キャッシュフロー】")

    # Row 30: FCF = 営業利益 + 減価償却 - CAPEX
    fcf_f = [f"={_col(c)}{r['op']}+{_col(c)}{r['depreciation']}-{_col(c)}{r['capex']}" for c in FY_COLS]
    _write_formula_row(ws, r["fcf"], "FCF（簡易営業CF）", fcf_f, bold=True)

    # Row 31: 累積FCF
    cum_fcf_f = []
    for i, c in enumerate(FY_COLS):
        cl = _col(c)
        if i == 0:
            cum_fcf_f.append(f"={cl}{r['fcf']}")
        else:
            prev = _col(FY_COLS[i - 1])
            cum_fcf_f.append(f"={prev}{r['cum_fcf']}+{cl}{r['fcf']}")
    _write_formula_row(ws, r["cum_fcf"], "累積FCF", cum_fcf_f, bold=True)
    _apply_border_row(ws, r["cum_fcf"], BOTTOM_DOUBLE)

    # Row 32: separator
    _separator_row(ws, 32)

    # ---------- 【資金需要】 ----------
    _write_section_header(ws, r["fund_header"], "【資金需要】")

    # Row 34: 必要資本金 = |MIN(累積FCF)| (if negative) + buffer
    cum_range = f"{_col(FY_COLS[0])}{r['cum_fcf']}:{_col(FY_COLS[-1])}{r['cum_fcf']}"
    fund_f = f"=IF(MIN({cum_range})<0, -MIN({cum_range}), 0)"
    ws.cell(row=r["funding_need"], column=1, value="必要資本金（最低限）").font = BOLD_FONT
    c = ws.cell(row=r["funding_need"], column=2, value=fund_f)
    c.font = FORMULA_BOLD
    c.number_format = NUMBER_FMT
    ws.merge_cells(start_row=r["funding_need"], start_column=2,
                   end_row=r["funding_need"], end_column=6)


def build_segment_sheet(
    wb: Workbook,
    seg_idx: int,
    fy_labels: Optional[List[str]] = None,
    model_type: str = "subscription",
) -> str:
    """Build a generic segment revenue model sheet (Layer C).

    Each sheet supports up to MAX_STREAMS revenue streams with:
      - stream name, unit price, volume, frequency, rate (all INPUT)
      - stream revenue (formula)
    Plus a summary section with total revenue, variable cost rate, gross profit.

    Returns the sheet name created.
    """
    fy = fy_labels or DEFAULT_FY_LABELS
    sheet_name = _seg_sheet_name(seg_idx)
    ws = wb.create_sheet(sheet_name)
    _set_col_widths(ws, {1: 26, 2: 16, 3: 16, 4: 16, 5: 16, 6: 16})

    # Title
    _write_title(ws, 1, f"セグメント{seg_idx} 収益モデル")

    # Row 2: model type info
    ws.cell(row=2, column=1, value=f"モデルタイプ: {model_type}").font = LABEL_FONT

    # Row 3: FY headers
    _write_fy_headers(ws, SEG_ROWS["header_fy"], fy)

    # Build streams
    for stream in range(1, MAX_STREAMS + 1):
        base_row = SEG_ROWS["stream_start"] + (stream - 1) * SEG_STREAM_ROWS_PER
        _build_stream_block(ws, base_row, stream, fy)

    # Summary section
    summary_start = SEG_ROWS["stream_start"] + MAX_STREAMS * SEG_STREAM_ROWS_PER
    _build_segment_summary(ws, summary_start, fy)

    return sheet_name


def _build_stream_block(ws, base_row: int, stream_num: int, fy: List[str]) -> None:
    """Build one revenue stream block within a segment sheet."""
    r = base_row

    # Stream header
    _write_section_header(ws, r, f"【ストリーム{stream_num}】")

    # Row +1: ストリーム名 (text input, single cell)
    ws.cell(row=r + 1, column=1, value="ストリーム名").font = LABEL_FONT
    name_cell = ws.cell(row=r + 1, column=2, value=f"ストリーム{stream_num}")
    name_cell.fill = YELLOW_FILL
    name_cell.font = NUMBER_FONT

    # Row +2: 単価 (INPUT × 5)
    _write_input_row(ws, r + 2, "単価（円）", [0] * 5, indent=1)

    # Row +3: 数量 (INPUT × 5)
    _write_input_row(ws, r + 3, "数量（顧客/取引/件）", [0] * 5, indent=1)

    # Row +4: 頻度 (INPUT × 5)
    _write_input_row(ws, r + 4, "頻度（月間）", [1] * 5, indent=1)

    # Row +5: 継続率/解約率 (INPUT × 5)
    _write_input_row(ws, r + 5, "継続率/解約率", [0] * 5, fmt=PERCENT_FMT, indent=1)

    # Row +6: ストリーム売上 = 単価 × 数量 × 頻度 × 12
    rev_f = [f"={_col(c)}{r+2}*{_col(c)}{r+3}*{_col(c)}{r+4}*12" for c in FY_COLS]
    _write_formula_row(ws, r + 6, f"ストリーム{stream_num}売上", rev_f, bold=True)
    _apply_border_row(ws, r + 6, THIN_BORDER)


def _build_segment_summary(ws, start_row: int, fy: List[str]) -> None:
    """Build the segment summary section (total revenue, variable cost, gross profit)."""
    r = start_row

    _write_section_header(ws, r, "【セグメント集計】")

    # Row +1: セグメント売上合計 = SUM of all stream revenues
    rev_rows = [
        SEG_ROWS["stream_start"] + (s - 1) * SEG_STREAM_ROWS_PER + 6
        for s in range(1, MAX_STREAMS + 1)
    ]
    total_f = []
    for c in FY_COLS:
        cl = _col(c)
        refs = "+".join(f"{cl}{rr}" for rr in rev_rows)
        total_f.append(f"={refs}")
    _write_formula_row(ws, r + 1, "セグメント売上合計", total_f, bold=True)
    _apply_border_row(ws, r + 1, BOTTOM_DOUBLE)

    # Row +2: 変動費率 (INPUT)
    _write_input_row(ws, r + 2, "変動費率", [0.0] * 5, fmt=PERCENT_FMT)

    # Row +3: 変動費 = 売上 × 変動費率
    vc_f = [f"={_col(c)}{r+1}*{_col(c)}{r+2}" for c in FY_COLS]
    _write_formula_row(ws, r + 3, "変動費", vc_f)

    # Row +4: セグメント粗利 = 売上 - 変動費
    gp_f = [f"={_col(c)}{r+1}-{_col(c)}{r+3}" for c in FY_COLS]
    _write_formula_row(ws, r + 4, "セグメント粗利", gp_f, bold=True)
    _apply_border_row(ws, r + 4, BOTTOM_DOUBLE)


def build_simulation_sheet(
    wb: Workbook,
    fy_labels: Optional[List[str]] = None,
) -> None:
    """Build シミュレーション分析 sheet (Layer A).

    Contains:
      - Scenario dropdown (Worst/Base/Best)
      - Target PL section (editable targets + gap analysis)
      - Scenario parameter tables (8 variables × 3 scenarios × 5 years)
      - Simulation results (base vs scenario-adjusted)
      - Funding simulator with buffer analysis
    """
    fy = fy_labels or DEFAULT_FY_LABELS
    ws = wb.create_sheet("シミュレーション分析")
    # Move to first position
    wb.move_sheet(ws, offset=-(len(wb.sheetnames) - 1))

    _set_col_widths(ws, {1: 28, 2: 16, 3: 16, 4: 16, 5: 16, 6: 16})

    R = SIM_ROWS
    pl = "PL設計"

    # Title
    _write_title(ws, 1, "シミュレーション分析")
    _separator_row(ws, 2)

    # ===== Scenario selector =====
    ws.cell(row=R["scenario_select"], column=1, value="シナリオ選択").font = BOLD_FONT
    sc = ws.cell(row=R["scenario_select"], column=4, value="Base")
    sc.fill = YELLOW_FILL
    sc.font = Font(name="Meiryo", size=12, bold=True, color="FF4472C4")
    sc.alignment = Alignment(horizontal="center")
    # Data validation dropdown
    dv = DataValidation(type="list", formula1='"Worst,Base,Best"', allow_blank=False)
    dv.error = "Worst, Base, Best から選んでください"
    dv.errorTitle = "シナリオ選択エラー"
    ws.add_data_validation(dv)
    dv.add(ws.cell(row=R["scenario_select"], column=4))

    _separator_row(ws, 4)

    # ===== ターゲットPL =====
    _write_section_header(ws, R["target_header"], "【ターゲットPL】")
    _write_fy_headers(ws, R["target_fy_hdr"], fy)
    _write_input_row(ws, R["target_rev"], "売上ターゲット", [0] * 5)
    _write_input_row(ws, R["target_op"], "営業利益ターゲット", [0] * 5)

    ws.cell(row=R["target_sy_be"], column=1, value="単年黒字目標年度").font = LABEL_FONT
    sc = ws.cell(row=R["target_sy_be"], column=2, value="")
    sc.fill = YELLOW_FILL
    sc.font = NUMBER_FONT

    ws.cell(row=R["target_cum_be"], column=1, value="累積黒字目標年度").font = LABEL_FONT
    sc = ws.cell(row=R["target_cum_be"], column=2, value="")
    sc.fill = YELLOW_FILL
    sc.font = NUMBER_FONT

    _separator_row(ws, 11)

    # ===== モデル値 vs ターゲット =====
    _write_section_header(ws, R["mvt_header"], "【モデル値 vs ターゲット】")
    _write_fy_headers(ws, R["mvt_fy_hdr"], fy)

    # Model revenue from PL
    rev_f = [f"='{pl}'!{_col(c)}{PL_ROWS['revenue']}" for c in FY_COLS]
    _write_formula_row(ws, R["model_rev"], "モデル売上", rev_f, bold=True)

    # Target revenue (from input above)
    tgt_rev_f = [f"={_col(c)}{R['target_rev']}" for c in FY_COLS]
    _write_formula_row(ws, R["tgt_rev"], "ターゲット売上", tgt_rev_f)

    # Gap
    gap_rev_f = [f"={_col(c)}{R['model_rev']}-{_col(c)}{R['tgt_rev']}" for c in FY_COLS]
    _write_formula_row(ws, R["gap_rev"], "Gap（売上）", gap_rev_f)

    # Achievement rate
    pct_rev_f = [f"=IF({_col(c)}{R['tgt_rev']}=0,0,{_col(c)}{R['model_rev']}/{_col(c)}{R['tgt_rev']})" for c in FY_COLS]
    _write_formula_row(ws, R["pct_rev"], "達成率（売上）", pct_rev_f, fmt=PERCENT_FMT, is_percent=True)

    # Model OP from PL
    op_f = [f"='{pl}'!{_col(c)}{PL_ROWS['op']}" for c in FY_COLS]
    _write_formula_row(ws, R["model_op"], "モデル営業利益", op_f, bold=True)

    tgt_op_f = [f"={_col(c)}{R['target_op']}" for c in FY_COLS]
    _write_formula_row(ws, R["tgt_op"], "ターゲット営業利益", tgt_op_f)

    gap_op_f = [f"={_col(c)}{R['model_op']}-{_col(c)}{R['tgt_op']}" for c in FY_COLS]
    _write_formula_row(ws, R["gap_op"], "Gap（営利）", gap_op_f)

    pct_op_f = [f"=IF({_col(c)}{R['tgt_op']}=0,0,{_col(c)}{R['model_op']}/{_col(c)}{R['tgt_op']})" for c in FY_COLS]
    _write_formula_row(ws, R["pct_op"], "達成率（営利）", pct_op_f, fmt=PERCENT_FMT, is_percent=True)

    # Breakeven timing (model)
    # Single year breakeven: first FY where OP > 0
    op_cells = ",".join(f"'{pl}'!{_col(c)}{PL_ROWS['op']}" for c in FY_COLS)
    be_labels = ",".join(f'"{fl}"' for fl in fy)
    ws.cell(row=R["model_sy_be"], column=1, value="単年黒字タイミング（モデル）").font = LABEL_FONT
    # Use nested IFs to find first positive year
    _write_breakeven_formula(ws, R["model_sy_be"], fy, PL_ROWS["op"], pl, "single")

    ws.cell(row=R["model_cum_be"], column=1, value="累積黒字タイミング（モデル）").font = LABEL_FONT
    _write_breakeven_formula(ws, R["model_cum_be"], fy, PL_ROWS["cum_fcf"], pl, "cumulative")

    _separator_row(ws, 24)

    # ===== シナリオ別パラメータ設定 =====
    _write_section_header(ws, R["sp_header"], "【シナリオ別パラメータ設定（年度別）】")

    for scenario, label, header_row, start_row in [
        ("worst", "■ Worst（悲観）", R["worst_header"], R["worst_start"]),
        ("base", "■ Base（基本）", R["base_header"], R["base_start"]),
        ("best", "■ Best（楽観）", R["best_header"], R["best_start"]),
    ]:
        ws.cell(row=header_row, column=1, value=label).font = SUBHEADER_FONT
        for c in range(1, 7):
            ws.cell(row=header_row, column=c).fill = SUBHEADER_FILL
        _write_fy_headers(ws, header_row + 1, fy, label="変数")

        defaults = DEFAULT_SCENARIO_VALUES[scenario]
        for vi, var_name in enumerate(SCENARIO_VARS):
            row = start_row + vi
            val = defaults[vi]
            _write_input_row(ws, row, var_name, [val] * 5, fmt='0.00', indent=1)

        # Separator after last var
        _separator_row(ws, start_row + len(SCENARIO_VARS))

    # ===== シミュレーション結果 =====
    _write_section_header(ws, R["sr_header"], "【シミュレーション結果（PL vs シナリオ適用後）】")
    _write_fy_headers(ws, R["sr_fy_hdr"], fy)

    # Helper: scenario multiplier lookup
    # $D$3 = selected scenario; uses IF to pick from Worst/Base/Best tables
    def _scenario_cell(var_idx: int, col: int) -> str:
        """Return IF formula that picks the right scenario multiplier."""
        cl = _col(col)
        w_row = R["worst_start"] + var_idx
        b_row = R["base_start"] + var_idx
        t_row = R["best_start"] + var_idx
        return (
            f'IF($D${R["scenario_select"]}="Worst",{cl}{w_row},'
            f'IF($D${R["scenario_select"]}="Best",{cl}{t_row},{cl}{b_row}))'
        )

    # idx 0 = 売上倍率
    _sr_row(ws, R["sr_base_rev"], "Base売上", PL_ROWS["revenue"], pl, None, None)
    _sr_row(ws, R["sr_sim_rev"], "シミュ後売上", PL_ROWS["revenue"], pl, 0, R)  # var 0 = 売上倍率

    _sr_row(ws, R["sr_base_gp"], "Base粗利", PL_ROWS["gross_profit"], pl, None, None)
    # GP adjusted = Base GP * (1 - cogs_rate_adj + base_cogs_rate) / (1 - base_cogs_rate)
    # Simplified: GP × revenue_multiplier × (1 - cogs_adj) where cogs_adj is relative
    _sr_gp_row(ws, R["sr_sim_gp"], "シミュ後粗利", R)

    _sr_row(ws, R["sr_base_opex"], "Base OPEX", PL_ROWS["opex_total"], pl, None, None)
    _sr_opex_row(ws, R["sr_sim_opex"], "シミュ後OPEX", R)

    _sr_row(ws, R["sr_base_depr"], "Base償却費", PL_ROWS["depreciation"], pl, None, None)
    _sr_row(ws, R["sr_sim_depr"], "シミュ後償却費", PL_ROWS["depreciation"], pl, None, None)  # depr unchanged

    _sr_row(ws, R["sr_base_capex"], "Base CAPEX", PL_ROWS["capex"], pl, None, None)
    _sr_row(ws, R["sr_sim_capex"], "シミュ後CAPEX", PL_ROWS["capex"], pl, 7, R)  # var 7 = CAPEX変動率

    _sr_row(ws, R["sr_base_op"], "Base営業利益", PL_ROWS["op"], pl, None, None)
    # Simulated OP = sim GP - sim OPEX - sim depr
    sim_op_f = [f"={_col(c)}{R['sr_sim_gp']}-{_col(c)}{R['sr_sim_opex']}-{_col(c)}{R['sr_sim_depr']}" for c in FY_COLS]
    _write_formula_row(ws, R["sr_sim_op"], "シミュ後営業利益", sim_op_f, bold=True)
    _apply_border_row(ws, R["sr_sim_op"], BOTTOM_DOUBLE)

    # Rates
    gp_rate_f = [f"=IF({_col(c)}{R['sr_sim_rev']}=0,0,{_col(c)}{R['sr_sim_gp']}/{_col(c)}{R['sr_sim_rev']})" for c in FY_COLS]
    _write_formula_row(ws, R["sr_gp_rate"], "粗利率（シミュ後）", gp_rate_f, fmt=PERCENT_FMT, is_percent=True)

    op_rate_f = [f"=IF({_col(c)}{R['sr_sim_rev']}=0,0,{_col(c)}{R['sr_sim_op']}/{_col(c)}{R['sr_sim_rev']})" for c in FY_COLS]
    _write_formula_row(ws, R["sr_op_rate"], "営業利益率（シミュ後）", op_rate_f, fmt=PERCENT_FMT, is_percent=True)

    _separator_row(ws, 75)

    # ===== 必要資本金シミュレーター =====
    _write_section_header(ws, R["fs_header"], "【必要資本金シミュレーター】")
    _write_fy_headers(ws, R["fs_fy_hdr"], fy)

    # GP from simulation result
    fs_gp_f = [f"={_col(c)}{R['sr_sim_gp']}" for c in FY_COLS]
    _write_formula_row(ws, R["fs_gp"], "粗利(B)", fs_gp_f)

    # Need = OPEX + CAPEX (simulated)
    fs_need_f = [f"={_col(c)}{R['sr_sim_opex']}+{_col(c)}{R['sr_sim_capex']}" for c in FY_COLS]
    _write_formula_row(ws, R["fs_need"], "必要資金(A)=OPEX+CAPEX", fs_need_f)

    # FCF = B - A
    fs_fcf_f = [f"={_col(c)}{R['fs_gp']}-{_col(c)}{R['fs_need']}" for c in FY_COLS]
    _write_formula_row(ws, R["fs_fcf"], "FCF = B-A", fs_fcf_f, bold=True)

    # Cumulative FCF
    fs_cum_f = []
    for i, c in enumerate(FY_COLS):
        cl = _col(c)
        if i == 0:
            fs_cum_f.append(f"={cl}{R['fs_fcf']}")
        else:
            prev = _col(FY_COLS[i - 1])
            fs_cum_f.append(f"={prev}{R['fs_cum_fcf']}+{cl}{R['fs_fcf']}")
    _write_formula_row(ws, R["fs_cum_fcf"], "累積FCF", fs_cum_f, bold=True)
    _apply_border_row(ws, R["fs_cum_fcf"], BOTTOM_DOUBLE)

    _separator_row(ws, 82)

    # Buffer analysis
    _write_section_header(ws, R["fs_buf_header"], "【バッファ月数別 必要資本金】")

    # Column headers: A=label, B=3ヶ月, C=6ヶ月, D=9ヶ月, E=12ヶ月
    for i, (months, label) in enumerate([(3, "3ヶ月"), (6, "6ヶ月"), (9, "9ヶ月"), (12, "12ヶ月")]):
        c = ws.cell(row=R["fs_buf_cols"], column=2 + i, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")

    # Monthly OPEX = OPEX total / 12 (use max year OPEX as proxy)
    opex_range = f"{_col(FY_COLS[0])}{R['sr_sim_opex']}:{_col(FY_COLS[-1])}{R['sr_sim_opex']}"
    ws.cell(row=R["fs_monthly_opex"], column=1, value="月次OPEX（最大年）").font = LABEL_FONT
    for i in range(4):
        c = ws.cell(row=R["fs_monthly_opex"], column=2 + i,
                    value=f"=MAX({opex_range})/12")
        c.font = FORMULA_FONT
        c.number_format = NUMBER_FMT

    # Buffer amount = monthly OPEX × months
    ws.cell(row=R["fs_buffer_amt"], column=1, value="バッファ金額").font = LABEL_FONT
    for i, months in enumerate([3, 6, 9, 12]):
        c = ws.cell(row=R["fs_buffer_amt"], column=2 + i,
                    value=f"={_col(2+i)}{R['fs_monthly_opex']}*{months}")
        c.font = FORMULA_FONT
        c.number_format = NUMBER_FMT

    # Cumulative FCF floor (minimum)
    cum_range = f"{_col(FY_COLS[0])}{R['fs_cum_fcf']}:{_col(FY_COLS[-1])}{R['fs_cum_fcf']}"
    ws.cell(row=R["fs_cum_min"], column=1, value="累積FCF底（最小値）").font = LABEL_FONT
    for i in range(4):
        c = ws.cell(row=R["fs_cum_min"], column=2 + i,
                    value=f"=MIN({cum_range})")
        c.font = FORMULA_FONT
        c.number_format = NUMBER_FMT

    # Floor year
    ws.cell(row=R["fs_min_year"], column=1, value="底発生年度").font = LABEL_FONT
    fy_array = "{" + ",".join(f'"{fl}"' for fl in fy) + "}"
    for i in range(4):
        c = ws.cell(row=R["fs_min_year"], column=2 + i,
                    value=f'=IFERROR(INDEX({fy_array},MATCH(MIN({cum_range}),{_col(FY_COLS[0])}{R["fs_cum_fcf"]}:{_col(FY_COLS[-1])}{R["fs_cum_fcf"]},0)),"-")')
        c.font = FORMULA_FONT

    # Required funding = |floor| + buffer (if floor < 0)
    ws.cell(row=R["fs_funding"], column=1, value="必要資本金").font = BOLD_FONT
    for i in range(4):
        col = 2 + i
        c = ws.cell(row=R["fs_funding"], column=col,
                    value=f"=IF({_col(col)}{R['fs_cum_min']}<0,-{_col(col)}{R['fs_cum_min']}+{_col(col)}{R['fs_buffer_amt']},0)")
        c.font = FORMULA_BOLD
        c.number_format = NUMBER_FMT
    _apply_border_row(ws, R["fs_funding"], BOTTOM_DOUBLE, max_col=5)


# ---------------------------------------------------------------------------
# Simulation sheet helper formulas
# ---------------------------------------------------------------------------

def _write_breakeven_formula(ws, row: int, fy: List[str], pl_row: int, pl_sheet: str, kind: str) -> None:
    """Write breakeven year formula (nested IF checking each FY)."""
    # Build nested IF from last to first
    formula = '"-"'
    for i in range(len(fy) - 1, -1, -1):
        cl = _col(FY_COLS[i])
        ref = f"'{pl_sheet}'!{cl}{pl_row}"
        formula = f'IF({ref}>0,"{fy[i]}",{formula})'
    c = ws.cell(row=row, column=2, value=f"={formula}")
    c.font = FORMULA_FONT
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)


def _sr_row(ws, row: int, label: str, pl_row: int, pl_sheet: str,
            var_idx: Optional[int], R: Optional[dict]) -> None:
    """Write a simulation result row.

    If var_idx is None, just reference PL (base value).
    If var_idx is given, multiply by scenario parameter.
    """
    formulas = []
    for c in FY_COLS:
        cl = _col(c)
        base_ref = f"'{pl_sheet}'!{cl}{pl_row}"
        if var_idx is not None and R is not None:
            # Multiply by scenario multiplier
            w_row = R["worst_start"] + var_idx
            b_row = R["base_start"] + var_idx
            t_row = R["best_start"] + var_idx
            mult = (
                f'IF($D${R["scenario_select"]}="Worst",{cl}{w_row},'
                f'IF($D${R["scenario_select"]}="Best",{cl}{t_row},{cl}{b_row}))'
            )
            formulas.append(f"={base_ref}*{mult}")
        else:
            formulas.append(f"={base_ref}")
    bold = var_idx is not None
    _write_formula_row(ws, row, label, formulas, bold=bold)


def _sr_gp_row(ws, row: int, label: str, R: dict) -> None:
    """Simulated GP = simulated revenue × (1 - adjusted COGS rate).

    COGS rate adjustment: base_cogs_rate * scenario_cogs_adj (var idx 3)
    """
    formulas = []
    for c in FY_COLS:
        cl = _col(c)
        sim_rev = f"{cl}{R['sr_sim_rev']}"
        base_cogs_rate = f"'PL設計'!{cl}{PL_ROWS['cogs_rate']}"
        # var 3 = 売上原価率調整
        w_row = R["worst_start"] + 3
        b_row = R["base_start"] + 3
        t_row = R["best_start"] + 3
        cogs_adj = (
            f'IF($D${R["scenario_select"]}="Worst",{cl}{w_row},'
            f'IF($D${R["scenario_select"]}="Best",{cl}{t_row},{cl}{b_row}))'
        )
        formulas.append(f"={sim_rev}*(1-{base_cogs_rate}*{cogs_adj})")
    _write_formula_row(ws, row, label, formulas, bold=True)


def _sr_opex_row(ws, row: int, label: str, R: dict) -> None:
    """Simulated OPEX: each component × its scenario multiplier.

    payroll × var4, marketing × var5, (office+system+other) × var6
    """
    formulas = []
    pl = "PL設計"
    for c in FY_COLS:
        cl = _col(c)

        def _mult(var_idx: int) -> str:
            w = R["worst_start"] + var_idx
            b = R["base_start"] + var_idx
            t = R["best_start"] + var_idx
            return (
                f'IF($D${R["scenario_select"]}="Worst",{cl}{w},'
                f'IF($D${R["scenario_select"]}="Best",{cl}{t},{cl}{b}))'
            )

        payroll = f"'{pl}'!{cl}{PL_ROWS['payroll']}*{_mult(4)}"
        marketing = f"'{pl}'!{cl}{PL_ROWS['marketing']}*{_mult(5)}"
        rest = (
            f"('{pl}'!{cl}{PL_ROWS['office']}"
            f"+'{pl}'!{cl}{PL_ROWS['system']}"
            f"+'{pl}'!{cl}{PL_ROWS['other_opex']})*{_mult(6)}"
        )
        formulas.append(f"={payroll}+{marketing}+{rest}")
    _write_formula_row(ws, row, label, formulas, bold=True)


# ---------------------------------------------------------------------------
# Public API: create full v2 workbook
# ---------------------------------------------------------------------------

def create_v2_workbook(
    num_segments: int = 3,
    fy_labels: Optional[List[str]] = None,
    segment_model_types: Optional[List[str]] = None,
) -> Workbook:
    """Create a complete v2 workbook with 3-layer structure.

    Parameters
    ----------
    num_segments : int
        Number of business segments (1-10).
    fy_labels : list[str] | None
        Custom FY labels (e.g. ["FY26", "FY27", ...]).
    segment_model_types : list[str] | None
        Model type per segment (e.g. ["subscription", "transaction"]).

    Returns
    -------
    Workbook
        openpyxl Workbook ready to save.
    """
    fy = fy_labels or DEFAULT_FY_LABELS
    model_types = segment_model_types or ["subscription"] * num_segments

    wb = Workbook()

    # Layer B: PL設計 (built first as active sheet)
    build_pl_sheet(wb, num_segments=num_segments, fy_labels=fy)

    # Layer C: Segment model sheets
    for i in range(1, num_segments + 1):
        mt = model_types[i - 1] if i - 1 < len(model_types) else "subscription"
        build_segment_sheet(wb, i, fy_labels=fy, model_type=mt)

    # Layer A: Simulation (moved to first position)
    build_simulation_sheet(wb, fy_labels=fy)

    # Register named ranges for PL KPIs
    _register_pl_named_ranges(wb, num_fys=len(fy))

    # Set calculation mode
    from openpyxl.workbook.properties import CalcProperties
    wb.calculation = CalcProperties(fullCalcOnLoad=True)

    return wb


def save_v2_template(
    output_path: str,
    num_segments: int = 3,
    fy_labels: Optional[List[str]] = None,
    segment_model_types: Optional[List[str]] = None,
) -> str:
    """Create and save a v2 template to disk.

    Returns the output path.
    """
    wb = create_v2_workbook(
        num_segments=num_segments,
        fy_labels=fy_labels,
        segment_model_types=segment_model_types,
    )
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info("V2 template saved to %s", output_path)
    return output_path
