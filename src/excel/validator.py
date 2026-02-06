"""Post-generation Excel validation - ensures template integrity."""
import logging
from pathlib import Path
from typing import List, Dict, Any, Set, Tuple

import openpyxl

from ..config.models import ValidationResult, CellTarget

logger = logging.getLogger(__name__)

# Excel error values to check for
EXCEL_ERRORS = {"#REF!", "#DIV/0!", "#NAME?", "#NULL!", "#N/A", "#VALUE!", "#NUM!"}


class PLValidator:
    """Validates generated Excel against template to ensure formula integrity."""

    def __init__(self, template_path: str, generated_path: str, input_color: str = "FFFFF2CC"):
        self.template_path = Path(template_path)
        self.generated_path = Path(generated_path)
        self.input_color = input_color

    def validate(self) -> ValidationResult:
        """
        Run all validation checks.

        Checks:
        1. Formulas are preserved (non-input cells unchanged)
        2. No new Excel error values introduced
        3. fullCalcOnLoad is True
        4. Only input-colored constant cells were modified
        """
        template_wb = openpyxl.load_workbook(str(self.template_path))
        generated_wb = openpyxl.load_workbook(str(self.generated_path))

        errors = []
        warnings = []
        changed_cells = []
        formula_preserved = True
        no_new_errors = True

        # Check fullCalcOnLoad
        full_calc = False
        if generated_wb.calculation is not None:
            full_calc = getattr(generated_wb.calculation, 'fullCalcOnLoad', False)

        if not full_calc:
            errors.append("fullCalcOnLoad is not set to True")

        # Check each sheet
        for sheet_name in template_wb.sheetnames:
            if sheet_name not in generated_wb.sheetnames:
                errors.append(f"Sheet '{sheet_name}' missing from generated file")
                continue

            t_ws = template_wb[sheet_name]
            g_ws = generated_wb[sheet_name]

            # Collect template errors for comparison
            template_errors = self._collect_errors(t_ws)
            generated_errors = self._collect_errors(g_ws)
            new_errors = generated_errors - template_errors

            if new_errors:
                no_new_errors = False
                for cell_ref, error_val in new_errors:
                    errors.append(
                        f"New error in {sheet_name}!{cell_ref}: {error_val}"
                    )

            # Check cell-by-cell
            for row in t_ws.iter_rows():
                for t_cell in row:
                    cell_ref = t_cell.coordinate
                    g_cell = g_ws[cell_ref]

                    t_val = t_cell.value
                    g_val = g_cell.value

                    if t_val == g_val:
                        continue

                    # Value changed - check if this is an allowed change
                    is_input = self._is_input_color(t_cell)
                    is_formula = (
                        t_cell.data_type == 'f' or
                        (isinstance(t_val, str) and str(t_val).startswith('='))
                    )

                    if is_formula:
                        # Formula was modified - THIS IS AN ERROR
                        formula_preserved = False
                        errors.append(
                            f"FORMULA CHANGED in {sheet_name}!{cell_ref}: "
                            f"'{t_val}' -> '{g_val}'"
                        )
                    elif is_input:
                        # Input cell changed - this is expected
                        changed_cells.append(CellTarget(sheet=sheet_name, cell=cell_ref))
                    else:
                        # Non-input, non-formula cell changed - warning
                        warnings.append(
                            f"Non-input cell changed in {sheet_name}!{cell_ref}: "
                            f"'{t_val}' -> '{g_val}'"
                        )

        template_wb.close()
        generated_wb.close()

        passed = formula_preserved and no_new_errors and full_calc and len(errors) == 0

        return ValidationResult(
            passed=passed,
            formula_preserved=formula_preserved,
            no_new_errors=no_new_errors,
            full_calc_on_load=full_calc,
            changed_cells=changed_cells,
            errors_found=errors,
            warnings=warnings,
        )

    def _is_input_color(self, cell) -> bool:
        """Check if cell has input color."""
        try:
            fill = cell.fill
            if fill and fill.fgColor and fill.fgColor.rgb:
                color = str(fill.fgColor.rgb)
                target = self.input_color
                return (
                    color == target or
                    color.lstrip("0") == target.lstrip("0") or
                    color[-6:] == target[-6:]
                )
        except Exception:
            pass
        return False

    def _collect_errors(self, ws) -> Set[Tuple[str, str]]:
        """Collect all Excel error values in a worksheet."""
        errors = set()
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value in EXCEL_ERRORS:
                    errors.add((cell.coordinate, cell.value))
        return errors


def generate_needs_review_csv(
    parameters: list,
    output_path: str,
) -> str:
    """Generate needs_review.csv for parameters that need human review."""
    import csv

    needs_review = [
        p for p in parameters
        if p.confidence < 0.7 or p.source == "inferred" or p.source == "template_default"
    ]

    with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow([
            "Key", "Label", "Value", "Confidence", "Source",
            "Evidence", "Mapped Cells", "Status"
        ])
        for p in needs_review:
            status = "NEEDS REVIEW" if p.confidence < 0.5 else "LOW CONFIDENCE"
            if p.source == "template_default":
                status = "TEMPLATE DEFAULT"
            writer.writerow([
                p.key, p.label, p.value, f"{p.confidence:.2f}", p.source,
                p.evidence.quote[:100] if p.evidence else "",
                "; ".join(f"{t.sheet}!{t.cell}" for t in p.mapped_targets),
                status
            ])

    return output_path
