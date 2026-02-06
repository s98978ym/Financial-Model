"""Simulation engine for PL model - Monte Carlo and sensitivity analysis."""
import logging
import json
import random
import math
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
from copy import deepcopy
from dataclasses import dataclass, field

logger = logging.getLogger(__name__)


@dataclass
class SimulationResult:
    """Result of a single simulation run."""
    iteration: int
    parameter_values: Dict[str, float]
    kpi_results: Dict[str, float]


@dataclass
class SimulationSummary:
    """Summary statistics from simulation runs."""
    kpi_name: str
    mean: float
    median: float
    std_dev: float
    p10: float
    p50: float
    p90: float
    min_val: float
    max_val: float
    distribution_type: str = "assumed_normal"
    notes: str = ""


@dataclass
class FullSimulationReport:
    """Complete simulation report."""
    method: str  # "xlwings" or "fallback"
    iterations: int
    summaries: List[SimulationSummary]
    parameter_distributions: Dict[str, Dict[str, float]]
    raw_results: List[SimulationResult] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)


class SimulationEngine:
    """Monte Carlo simulation for PL models."""

    def __init__(
        self,
        iterations: int = 1000,
        kpi_targets: Optional[List[str]] = None,
    ):
        self.iterations = iterations
        self.kpi_targets = kpi_targets or [
            "revenue", "gross_profit", "operating_profit", "ebitda"
        ]

    def run(
        self,
        base_parameters: List[Any],  # List[ExtractedParameter]
        parameter_ranges: Optional[Dict[str, Tuple[float, float]]] = None,
        template_path: Optional[str] = None,
    ) -> FullSimulationReport:
        """
        Run simulation. Try xlwings first, fall back to analytical.
        """
        # Try xlwings-based simulation
        if template_path:
            try:
                return self._run_xlwings(base_parameters, parameter_ranges, template_path)
            except Exception as e:
                logger.warning(f"xlwings simulation failed: {e}, falling back to analytical")

        # Fallback: analytical/parameter sweep simulation
        return self._run_fallback(base_parameters, parameter_ranges)

    def _run_xlwings(
        self,
        parameters: List[Any],
        ranges: Optional[Dict[str, Tuple[float, float]]],
        template_path: str,
    ) -> FullSimulationReport:
        """Run simulation using xlwings for Excel recalculation."""
        try:
            import xlwings as xw
        except ImportError:
            raise ImportError("xlwings not available")

        param_ranges = self._build_ranges(parameters, ranges)
        results = []

        app = xw.App(visible=False)
        try:
            wb = app.books.open(template_path)

            for i in range(self.iterations):
                # Sample parameters
                sampled = self._sample_parameters(param_ranges)

                # Write sampled values to Excel
                for param in parameters:
                    if param.key in sampled and param.selected:
                        for target in param.mapped_targets:
                            try:
                                ws = wb.sheets[target.sheet]
                                ws.range(target.cell).value = sampled[param.key]
                            except Exception:
                                pass

                # Recalculate
                wb.app.calculate()

                # Read KPIs (simplified - would need KPI cell mapping)
                kpi_values = self._read_kpis(wb)

                results.append(SimulationResult(
                    iteration=i,
                    parameter_values=sampled,
                    kpi_results=kpi_values,
                ))

            wb.close()
        finally:
            app.quit()

        return self._build_report(results, param_ranges, method="xlwings")

    def _run_fallback(
        self,
        parameters: List[Any],
        ranges: Optional[Dict[str, Tuple[float, float]]],
    ) -> FullSimulationReport:
        """
        Fallback simulation without Excel.
        Uses simplified analytical relationships between inputs and KPIs.
        """
        param_ranges = self._build_ranges(parameters, ranges)
        results = []

        for i in range(self.iterations):
            sampled = self._sample_parameters(param_ranges)

            # Simplified KPI estimation based on sampled parameters
            kpi_values = self._estimate_kpis(sampled, parameters)

            results.append(SimulationResult(
                iteration=i,
                parameter_values=sampled,
                kpi_results=kpi_values,
            ))

        report = self._build_report(results, param_ranges, method="fallback")
        report.warnings.append(
            "This simulation uses analytical approximation. "
            "For accurate results with full formula recalculation, "
            "use xlwings with Excel installed."
        )
        return report

    def _build_ranges(
        self,
        parameters: List[Any],
        custom_ranges: Optional[Dict[str, Tuple[float, float]]],
    ) -> Dict[str, Dict[str, float]]:
        """Build parameter ranges for simulation."""
        ranges = {}
        custom = custom_ranges or {}

        for param in parameters:
            if not param.selected or not isinstance(param.value, (int, float)):
                continue

            val = float(param.adjusted_value if param.adjusted_value is not None else param.value)

            if param.key in custom:
                low, high = custom[param.key]
                ranges[param.key] = {"base": val, "low": low, "high": high}
            else:
                # Default: +/- 20% based on confidence
                uncertainty = 0.2 * (1.0 + (1.0 - param.confidence))
                ranges[param.key] = {
                    "base": val,
                    "low": val * (1 - uncertainty),
                    "high": val * (1 + uncertainty),
                }

        return ranges

    def _sample_parameters(
        self, ranges: Dict[str, Dict[str, float]]
    ) -> Dict[str, float]:
        """Sample parameter values from their ranges using triangular distribution."""
        sampled = {}
        for key, r in ranges.items():
            # Triangular distribution: mode at base, bounded by low/high
            sampled[key] = random.triangular(r["low"], r["high"], r["base"])
        return sampled

    def _estimate_kpis(
        self, sampled: Dict[str, float], parameters: List[Any]
    ) -> Dict[str, float]:
        """
        Estimate KPI values from sampled parameters.
        This is a simplified estimation for fallback mode.
        """
        kpis = {}

        # Try to compute revenue from revenue-related params
        revenue_keys = [k for k in sampled if any(
            w in k.lower() for w in ["revenue", "売上", "arpu", "price", "単価"]
        )]
        volume_keys = [k for k in sampled if any(
            w in k.lower() for w in ["customer", "顧客", "user", "数", "count", "volume"]
        )]
        cost_keys = [k for k in sampled if any(
            w in k.lower() for w in ["cost", "費", "cogs", "原価"]
        )]

        # Simple revenue estimation
        if revenue_keys and volume_keys:
            rev = sampled[revenue_keys[0]] * sampled[volume_keys[0]]
        elif revenue_keys:
            rev = sampled[revenue_keys[0]]
        else:
            rev = sum(v for k, v in sampled.items() if "revenue" in k.lower() or "売上" in k.lower()) or 0

        kpis["revenue"] = rev

        # Cost estimation
        total_cost = sum(sampled.get(k, 0) for k in cost_keys)
        kpis["total_cost"] = total_cost

        # Gross profit
        cogs = total_cost * 0.4  # Simplified
        kpis["gross_profit"] = rev - cogs

        # Operating profit
        opex = total_cost * 0.6
        kpis["operating_profit"] = rev - cogs - opex

        # EBITDA (simplified)
        kpis["ebitda"] = kpis["operating_profit"] * 1.1

        return kpis

    def _read_kpis(self, wb) -> Dict[str, float]:
        """Read KPI values from Excel workbook (xlwings)."""
        kpis = {}
        # This would need proper KPI cell mapping from modelmap
        # Simplified version reads from known locations
        try:
            if "PL設計" in [s.name for s in wb.sheets]:
                ws = wb.sheets["PL設計"]
                # These would be configured based on template analysis
                kpis["revenue"] = ws.range("B10").value or 0
                kpis["gross_profit"] = ws.range("B20").value or 0
                kpis["operating_profit"] = ws.range("B30").value or 0
                kpis["ebitda"] = ws.range("B35").value or 0
        except Exception as e:
            logger.warning(f"Error reading KPIs: {e}")
        return kpis

    def _build_report(
        self,
        results: List[SimulationResult],
        param_ranges: Dict[str, Dict[str, float]],
        method: str,
    ) -> FullSimulationReport:
        """Build simulation summary report from raw results."""
        summaries = []

        # Collect KPI values across iterations
        kpi_collections: Dict[str, List[float]] = {}
        for r in results:
            for kpi_name, kpi_val in r.kpi_results.items():
                if kpi_name not in kpi_collections:
                    kpi_collections[kpi_name] = []
                kpi_collections[kpi_name].append(kpi_val)

        for kpi_name, values in kpi_collections.items():
            if not values:
                continue
            values_sorted = sorted(values)
            n = len(values_sorted)

            summaries.append(SimulationSummary(
                kpi_name=kpi_name,
                mean=sum(values) / n,
                median=values_sorted[n // 2],
                std_dev=self._std_dev(values),
                p10=values_sorted[int(n * 0.1)],
                p50=values_sorted[int(n * 0.5)],
                p90=values_sorted[int(n * 0.9)],
                min_val=values_sorted[0],
                max_val=values_sorted[-1],
                distribution_type="triangular_sampled",
                notes=f"Based on {n} iterations ({method})",
            ))

        return FullSimulationReport(
            method=method,
            iterations=len(results),
            summaries=summaries,
            parameter_distributions={
                k: v for k, v in param_ranges.items()
            },
            raw_results=results if len(results) <= 100 else results[:100],
            warnings=[],
        )

    @staticmethod
    def _std_dev(values: List[float]) -> float:
        """Calculate standard deviation."""
        n = len(values)
        if n < 2:
            return 0.0
        mean = sum(values) / n
        variance = sum((x - mean) ** 2 for x in values) / (n - 1)
        return math.sqrt(variance)


def export_simulation_summary(
    report: FullSimulationReport, output_path: str
) -> str:
    """Export simulation summary to Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Simulation Summary"

    # Header styling
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_font = Font(color="FFFFFFFF", bold=True)

    # Title
    ws["A1"] = "Monte Carlo Simulation Summary"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = f"Method: {report.method} | Iterations: {report.iterations}"

    # KPI Summary table
    headers = ["KPI", "Mean", "Median", "Std Dev", "P10", "P50", "P90", "Min", "Max"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, summary in enumerate(report.summaries, 5):
        ws.cell(row=row_idx, column=1, value=summary.kpi_name)
        ws.cell(row=row_idx, column=2, value=round(summary.mean, 2))
        ws.cell(row=row_idx, column=3, value=round(summary.median, 2))
        ws.cell(row=row_idx, column=4, value=round(summary.std_dev, 2))
        ws.cell(row=row_idx, column=5, value=round(summary.p10, 2))
        ws.cell(row=row_idx, column=6, value=round(summary.p50, 2))
        ws.cell(row=row_idx, column=7, value=round(summary.p90, 2))
        ws.cell(row=row_idx, column=8, value=round(summary.min_val, 2))
        ws.cell(row=row_idx, column=9, value=round(summary.max_val, 2))

    # Parameter distributions
    param_start = len(report.summaries) + 7
    ws.cell(row=param_start, column=1, value="Parameter Distributions").font = Font(size=12, bold=True)

    param_headers = ["Parameter", "Base", "Low", "High"]
    for col, header in enumerate(param_headers, 1):
        cell = ws.cell(row=param_start + 1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, (param_name, dist) in enumerate(report.parameter_distributions.items(), param_start + 2):
        ws.cell(row=row_idx, column=1, value=param_name)
        ws.cell(row=row_idx, column=2, value=round(dist.get("base", 0), 2))
        ws.cell(row=row_idx, column=3, value=round(dist.get("low", 0), 2))
        ws.cell(row=row_idx, column=4, value=round(dist.get("high", 0), 2))

    # Warnings
    if report.warnings:
        warn_start = param_start + len(report.parameter_distributions) + 4
        ws.cell(row=warn_start, column=1, value="Warnings").font = Font(size=12, bold=True, color="FFFF0000")
        for i, warning in enumerate(report.warnings):
            ws.cell(row=warn_start + 1 + i, column=1, value=warning)

    # Auto-fit columns
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 30)

    wb.save(output_path)
    return output_path
