"""Workbook export helpers for FAM PDCA artifacts."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .source_registry import analysis_source_refs


YEAR_HEADERS = ["FY1", "FY2", "FY3", "FY4", "FY5"]

ACADEMY_LEVELS = [
    {
        "code": "c",
        "label": "C級課程",
        "share_default": [1.00, 0.65, 0.45, 0.35, 0.25],
        "price_multiplier": 1.0,
        "completion": 0.90,
        "certification": 1.00,
        "progression_to_next": 0.60,
    },
    {
        "code": "b",
        "label": "B級課程",
        "share_default": [0.00, 0.20, 0.20, 0.18, 0.15],
        "price_multiplier": 100000 / 70000,
        "completion": 0.90,
        "certification": 0.50,
        "progression_to_next": 0.30,
    },
    {
        "code": "a",
        "label": "A級課程",
        "share_default": [0.00, 0.10, 0.18, 0.22, 0.25],
        "price_multiplier": 300000 / 70000,
        "completion": 0.80,
        "certification": 0.50,
        "progression_to_next": 0.90,
    },
    {
        "code": "s",
        "label": "S級課程",
        "share_default": [0.00, 0.05, 0.17, 0.25, 0.35],
        "price_multiplier": 0.0,
        "completion": 0.90,
        "certification": 0.75,
        "progression_to_next": 0.00,
    },
]

CONSULT_SKUS = [
    {"sku": "P1", "service_name": "S総合（年契）", "unit": "契約/年", "base_price": 15000000, "retention_multiplier": 1.00, "standard_hours": 480, "share_default": [0.18, 0.20, 0.22, 0.24, 0.25]},
    {"sku": "P2", "service_name": "A総合（年契）", "unit": "契約/年", "base_price": 10000000, "retention_multiplier": 0.83, "standard_hours": 360, "share_default": [0.20, 0.22, 0.22, 0.22, 0.23]},
    {"sku": "P3", "service_name": "Sチーム年契", "unit": "契約/年", "base_price": 5000000, "retention_multiplier": 0.83, "standard_hours": 240, "share_default": [0.10, 0.10, 0.10, 0.10, 0.10]},
    {"sku": "P4", "service_name": "Aチーム年契", "unit": "契約/年", "base_price": 1500000, "retention_multiplier": 0.83, "standard_hours": 120, "share_default": [0.10, 0.12, 0.13, 0.14, 0.15]},
    {"sku": "P5", "service_name": "セミナー（単発）", "unit": "回", "base_price": 300000, "retention_multiplier": 0.00, "standard_hours": 10, "share_default": [0.10, 0.08, 0.07, 0.06, 0.05]},
    {"sku": "P6", "service_name": "ライトセミナー", "unit": "回", "base_price": 60000, "retention_multiplier": 0.00, "standard_hours": 6, "share_default": [0.06, 0.05, 0.04, 0.03, 0.03]},
    {"sku": "P8", "service_name": "S個別（年契）", "unit": "人/年", "base_price": 500000, "retention_multiplier": 0.83, "standard_hours": 30, "share_default": [0.07, 0.07, 0.08, 0.08, 0.08]},
    {"sku": "P9", "service_name": "A個別（年契）", "unit": "人/年", "base_price": 500000, "retention_multiplier": 0.83, "standard_hours": 30, "share_default": [0.07, 0.06, 0.06, 0.05, 0.05]},
    {"sku": "P10", "service_name": "S個別（単発）", "unit": "回", "base_price": 50000, "retention_multiplier": 0.00, "standard_hours": 3, "share_default": [0.05, 0.04, 0.03, 0.03, 0.03]},
    {"sku": "P11", "service_name": "A個別（単発）", "unit": "回", "base_price": 50000, "retention_multiplier": 0.00, "standard_hours": 3, "share_default": [0.05, 0.04, 0.03, 0.03, 0.02]},
    {"sku": "P12", "service_name": "OJT", "unit": "回", "base_price": 0, "retention_multiplier": 0.00, "standard_hours": 3, "share_default": [0.02, 0.02, 0.02, 0.02, 0.01]},
]

ASSUMPTION_ROWS = {
    "revenue_target": 2,
    "gross_profit_target": 3,
    "opex_target": 4,
    "academy_public_price": 6,
    "academy_students": 7,
    "academy_certified": 8,
    "academy_revenue": 9,
    "academy_effective_price": 10,
    "meal_price_per_item": 12,
    "meal_items_per_meal": 13,
    "meal_meals_per_year": 14,
    "meal_retention_rate": 15,
    "meal_share_non_academy": 16,
    "meal_unit_count": 17,
    "meal_revenue": 18,
    "consult_unit_price": 20,
    "consult_retention": 21,
    "consult_standard_hours": 22,
    "consult_revenue": 23,
    "consult_project_count": 24,
    "gross_margin_ratio": 26,
    "personnel_ratio": 28,
    "marketing_ratio": 29,
    "development_ratio": 30,
    "other_ratio": 31,
    "academy_c_share": 33,
    "academy_b_share": 34,
    "academy_a_share": 35,
    "academy_s_share": 36,
    "academy_c_to_b": 38,
    "academy_b_to_a": 39,
    "academy_a_to_s": 40,
    "academy_c_price_multiplier": 42,
    "academy_b_price_multiplier": 43,
    "academy_a_price_multiplier": 44,
    "academy_s_price_multiplier": 45,
    "academy_c_completion": 47,
    "academy_c_certification": 48,
    "academy_b_completion": 49,
    "academy_b_certification": 50,
    "academy_a_completion": 51,
    "academy_a_certification": 52,
    "academy_s_completion": 53,
    "academy_s_certification": 54,
    "blended_hourly_rate": 56,
    "consult_p1_share": 58,
    "consult_p2_share": 59,
    "consult_p3_share": 60,
    "consult_p4_share": 61,
    "consult_p5_share": 62,
    "consult_p6_share": 63,
    "consult_p8_share": 64,
    "consult_p9_share": 65,
    "consult_p10_share": 66,
    "consult_p11_share": 67,
    "consult_p12_share": 68,
    "development_investment": 70,
    "development_amortization_years": 71,
    "development_amortization_expense": 72,
    "development_unamortized_balance": 73,
}

PL_ROWS = {
    "revenue_total": 2,
    "academy_revenue": 3,
    "consult_revenue": 4,
    "meal_revenue": 5,
    "cogs": 6,
    "gross_profit": 7,
    "gross_margin_ratio": 8,
    "personnel_cost": 9,
    "marketing_cost": 10,
    "development_cost": 11,
    "other_opex": 12,
    "opex_total": 13,
    "operating_profit": 14,
    "operating_margin": 15,
}

COST_SUMMARY_ROWS = {
    "personnel_cost": 3,
    "marketing_cost": 4,
    "development_cost": 5,
    "other_opex": 6,
    "opex_total": 7,
}

COST_LIST_ROWS = [
    ("営業人件費", "人件費", 0.35, "personnel_cost"),
    ("運営人件費", "人件費", 0.30, "personnel_cost"),
    ("管理人件費", "人件費", 0.20, "personnel_cost"),
    ("採用・教育人件費", "人件費", 0.15, "personnel_cost"),
    ("メディア費", "マーケ費", 0.50, "marketing_cost"),
    ("イベント・販促費", "マーケ費", 0.30, "marketing_cost"),
    ("パートナーインセンティブ", "マーケ費", 0.20, "marketing_cost"),
    ("外部開発償却費", "開発費（償却）", 0.60, "development_cost"),
    ("内部開発償却費", "開発費（償却）", 0.40, "development_cost"),
    ("その他固定費", "その他OPEX", 1.00, "other_opex"),
]

COST_PLAN_ROWS = {
    "summary_title": 1,
    "summary_header": 2,
    "amortization_title": 9,
    "amortization_start": 10,
    "detail_title": 16,
    "detail_header": 17,
    "detail_start": 18,
}

ACADEMY_ROW_LAYOUT = {
    "c": {"label": 2, "price": 3, "students": 4, "certified": 5, "revenue": 6},
    "b": {"label": 7, "price": 8, "students": 9, "certified": 10, "revenue": 11},
    "a": {"label": 12, "price": 13, "students": 14, "certified": 15, "revenue": 16},
    "s": {"label": 17, "price": 18, "students": 19, "certified": 20, "revenue": 21},
    "total_students": 23,
    "total_revenue": 24,
}

CONSULT_DETAIL_START_ROW = 3
CONSULT_TOTAL_REVENUE_ROW = 15
CONSULT_TOTAL_DELIVERY_ROW = 16
CONSULT_TOTAL_GROSS_PROFIT_ROW = 17

INPUT_FILL = PatternFill(fill_type="solid", fgColor="DDEBF7")
FORMULA_FILL = PatternFill(fill_type="solid", fgColor="E2F0D9")
SUBTOTAL_FILL = PatternFill(fill_type="solid", fgColor="D9E2F3")
TOTAL_FILL = PatternFill(fill_type="solid", fgColor="A6A6A6")
UPDATED_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
BOLD_FONT = Font(bold=True)
TOTAL_FONT = Font(bold=True, color="FFFFFF")
NUMBER_FORMAT = "#,##0"
COUNT_DECIMAL_FORMAT = "#,##0.0"
PERCENT_FORMAT = "0.0%"
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
WRAP_LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)


def export_candidate_workbook(
    *,
    output_path: Path,
    candidate_id: str,
    candidate_payload: dict[str, Any],
    diagnosis: dict[str, Any],
    baseline_total: float,
    run_root: Path,
    iteration_summaries: list[dict[str, Any]] | None = None,
) -> None:
    """Export one candidate workbook for human review."""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    review_sheet = workbook.active
    review_sheet.title = "PDCAチェックシート"

    assumptions = _build_workbook_assumptions(candidate_payload)
    _write_pdca_check_sheet(
        review_sheet,
        candidate_id=candidate_id,
        diagnosis=diagnosis,
        baseline_total=baseline_total,
        candidate_payload=candidate_payload,
        run_root=run_root,
        iteration_summaries=iteration_summaries or [],
    )
    _write_qa_sheet(
        workbook.create_sheet("想定Q&A"),
        candidate_id=candidate_id,
        candidate_payload=candidate_payload,
        diagnosis=diagnosis,
        assumptions=assumptions,
        iteration_summaries=iteration_summaries or [],
        qa_mode="pdca",
    )
    _write_qa_sheet(
        workbook.create_sheet("収益計画Q&A"),
        candidate_id=candidate_id,
        candidate_payload=candidate_payload,
        diagnosis=diagnosis,
        assumptions=assumptions,
        iteration_summaries=iteration_summaries or [],
        qa_mode="revenue_plan",
    )
    _write_pl_sheet(workbook.create_sheet("PL設計"))
    _write_meal_sheet(workbook.create_sheet("ミールモデル"))
    _write_academy_sheet(workbook.create_sheet("アカデミーモデル"))
    _write_consulting_sheet(workbook.create_sheet("コンサルモデル"))
    _write_cost_plan_sheet(workbook.create_sheet("費用計画"), assumptions)
    _write_plan_assumptions_sheet(workbook.create_sheet("（全Ver）前提条件"), assumptions)
    workbook.save(output_path)


def _write_pdca_check_sheet(
    sheet,
    *,
    candidate_id: str,
    diagnosis: dict[str, Any],
    baseline_total: float,
    candidate_payload: dict[str, Any],
    run_root: Path,
    iteration_summaries: list[dict[str, Any]],
) -> None:
    hypothesis = diagnosis.get("hypothesis", {})
    verdict = diagnosis.get("verdict", {})
    score = diagnosis.get("score", {})
    logic = diagnosis.get("logic", {})
    evidence = diagnosis.get("evidence", {})

    row = 1
    row = _write_section_title(sheet, row, "今回の結論")
    conclusion_rows = [
        ("候補ID", candidate_id),
        ("判定", verdict.get("status", "")),
        ("判定理由", verdict.get("reason", "")),
        ("総合スコア", score.get("total", 0.0)),
        ("baseline比", score.get("delta_vs_baseline", round(score.get("total", 0.0) - baseline_total, 4))),
        ("baselineスコア", baseline_total),
    ]
    row = _write_labeled_values(sheet, row, conclusion_rows, numeric_labels={"総合スコア", "baseline比", "baselineスコア"})

    row += 1
    row = _write_section_title(sheet, row, "仮説")
    hypothesis_rows = [
        ("今回の仮説", hypothesis.get("title", "") or "-"),
        ("仮説の詳細", hypothesis.get("detail", "") or "-"),
        ("ONにした要素", ", ".join(logic.get("toggles_on", [])) or "-"),
        ("OFFにした要素", ", ".join(logic.get("toggles_off", [])) or "-"),
        ("ロジック", "\n".join(logic.get("steps", [])) or "-"),
    ]
    row = _write_labeled_values(sheet, row, hypothesis_rows)

    row += 1
    row = _write_section_title(sheet, row, "PDCA全体推移")
    row = _write_iteration_trend_table(
        sheet,
        row,
        iteration_summaries
        or [
            {
                "iteration": 1,
                "candidate_id": candidate_id,
                "hypothesis": hypothesis.get("title", ""),
                "result": verdict.get("reason", ""),
                "total_score": score.get("total", 0.0),
                "delta_vs_baseline": score.get("delta_vs_baseline", round(score.get("total", 0.0) - baseline_total, 4)),
                "next_action": (diagnosis.get("next_actions") or [""])[0],
            }
        ],
    )

    row += 1
    row = _write_section_title(sheet, row, "評価スコア")
    row = _write_score_table(sheet, row, score, baseline_total)

    row += 1
    row = _write_section_title(sheet, row, "根拠と前提")
    row = _write_assumption_table(sheet, row, candidate_payload.get("assumptions", []))

    row += 1
    row = _write_section_title(sheet, row, "関連ファイル")
    artifact_rows = [
        ("summary", str(run_root / "summary.md")),
        ("scores", str(run_root / "scores.json")),
        ("diagnosis", str(run_root / "diagnosis.json")),
        ("candidate_json", str(run_root / "candidates" / f"{candidate_id}.json")),
        ("reference", str(run_root / "reference.json")),
        ("baseline", str(run_root / "baseline.json")),
        ("外部根拠タイプ", ", ".join(evidence.get("source_types", [])) or "-"),
    ]
    row = _write_labeled_values(sheet, row, artifact_rows)

    row += 1
    row = _write_section_title(sheet, row, "次の改善施策")
    next_actions = diagnosis.get("next_actions", []) or ["次の施策は未設定です。"]
    row = _write_bullet_list(sheet, row, next_actions)

    sheet.freeze_panes = "B2"
    _set_column_widths(
        sheet,
        {
            "A": 24,
            "B": 24,
            "C": 24,
            "D": 22,
            "E": 24,
            "F": 24,
            "G": 12,
            "H": 14,
            "I": 10,
            "J": 16,
            "K": 20,
            "L": 12,
            "M": 30,
        },
    )
    _set_row_heights(sheet, {1: 24})
    _align_range(sheet, start_row=1, end_row=sheet.max_row, start_col=1, end_col=1, alignment=LEFT_ALIGN)
    _align_range(sheet, start_row=1, end_row=sheet.max_row, start_col=2, end_col=2, alignment=LEFT_ALIGN)
    _align_range(sheet, start_row=1, end_row=sheet.max_row, start_col=3, end_col=6, alignment=WRAP_LEFT_ALIGN)
    _align_range(sheet, start_row=1, end_row=sheet.max_row, start_col=7, end_col=10, alignment=RIGHT_ALIGN)
    _align_range(sheet, start_row=1, end_row=sheet.max_row, start_col=11, end_col=13, alignment=WRAP_LEFT_ALIGN)


def _write_pl_sheet(sheet) -> None:
    _write_header_row(sheet)
    labels = [
        ("revenue_total", "売上"),
        ("academy_revenue", "アカデミー"),
        ("consult_revenue", "コンサル"),
        ("meal_revenue", "ミール"),
        ("cogs", "売上原価"),
        ("gross_profit", "粗利"),
        ("gross_margin_ratio", "粗利率"),
        ("personnel_cost", "人件費"),
        ("marketing_cost", "マーケ費"),
        ("development_cost", "開発費（償却）"),
        ("other_opex", "その他OPEX"),
        ("opex_total", "事業運営費（OPEX）"),
        ("operating_profit", "営業利益"),
        ("operating_margin", "営業利益率"),
    ]
    for key, label in labels:
        sheet.cell(row=PL_ROWS[key], column=1, value=label)

    for column_index, assumption_col in enumerate(_year_columns(), start=2):
        model_col = excel_col(column_index)
        gross_margin_ref = _sheet_ref("（全Ver）前提条件", f"{assumption_col}{ASSUMPTION_ROWS['gross_margin_ratio']}")
        personnel_ref = _sheet_ref("費用計画", f"{model_col}{COST_SUMMARY_ROWS['personnel_cost']}")
        marketing_ref = _sheet_ref("費用計画", f"{model_col}{COST_SUMMARY_ROWS['marketing_cost']}")
        development_ref = _sheet_ref("費用計画", f"{model_col}{COST_SUMMARY_ROWS['development_cost']}")
        other_opex_ref = _sheet_ref("費用計画", f"{model_col}{COST_SUMMARY_ROWS['other_opex']}")
        academy_total_row = ACADEMY_ROW_LAYOUT["total_revenue"]
        consult_total_col = excel_col(column_index)
        sheet[f"{model_col}{PL_ROWS['academy_revenue']}"] = f"={_sheet_ref('アカデミーモデル', f'{model_col}{academy_total_row}')}"
        sheet[f"{model_col}{PL_ROWS['consult_revenue']}"] = f"={_sheet_ref('コンサルモデル', f'{consult_total_col}{CONSULT_TOTAL_REVENUE_ROW}')}"
        sheet[f"{model_col}{PL_ROWS['meal_revenue']}"] = f"={_sheet_ref('ミールモデル', f'{model_col}7')}"
        sheet[f"{model_col}{PL_ROWS['revenue_total']}"] = (
            f"=SUM({model_col}{PL_ROWS['academy_revenue']}:{model_col}{PL_ROWS['meal_revenue']})"
        )
        sheet[f"{model_col}{PL_ROWS['cogs']}"] = (
            f"={model_col}{PL_ROWS['revenue_total']}*(1-{gross_margin_ref})"
        )
        sheet[f"{model_col}{PL_ROWS['gross_profit']}"] = (
            f"={model_col}{PL_ROWS['revenue_total']}-{model_col}{PL_ROWS['cogs']}"
        )
        sheet[f"{model_col}{PL_ROWS['gross_margin_ratio']}"] = (
            f"=IF({model_col}{PL_ROWS['revenue_total']}<>0,{model_col}{PL_ROWS['gross_profit']}/{model_col}{PL_ROWS['revenue_total']},0)"
        )
        sheet[f"{model_col}{PL_ROWS['personnel_cost']}"] = f"={personnel_ref}"
        sheet[f"{model_col}{PL_ROWS['marketing_cost']}"] = f"={marketing_ref}"
        sheet[f"{model_col}{PL_ROWS['development_cost']}"] = f"={development_ref}"
        sheet[f"{model_col}{PL_ROWS['other_opex']}"] = f"={other_opex_ref}"
        sheet[f"{model_col}{PL_ROWS['opex_total']}"] = (
            f"=SUM({model_col}{PL_ROWS['personnel_cost']}:{model_col}{PL_ROWS['other_opex']})"
        )
        sheet[f"{model_col}{PL_ROWS['operating_profit']}"] = (
            f"={model_col}{PL_ROWS['gross_profit']}-{model_col}{PL_ROWS['opex_total']}"
        )
        sheet[f"{model_col}{PL_ROWS['operating_margin']}"] = (
            f"=IF({model_col}{PL_ROWS['revenue_total']}<>0,{model_col}{PL_ROWS['operating_profit']}/{model_col}{PL_ROWS['revenue_total']},0)"
        )
    _style_subtotal_rows(sheet, [PL_ROWS["academy_revenue"], PL_ROWS["consult_revenue"], PL_ROWS["meal_revenue"]])
    _style_formula_rows(sheet, [PL_ROWS["gross_margin_ratio"], PL_ROWS["operating_margin"]])
    _style_total_rows(sheet, [PL_ROWS["revenue_total"], PL_ROWS["gross_profit"], PL_ROWS["opex_total"], PL_ROWS["operating_profit"]])
    _set_number_format_rows(
        sheet,
        [
            PL_ROWS["revenue_total"],
            PL_ROWS["academy_revenue"],
            PL_ROWS["consult_revenue"],
            PL_ROWS["meal_revenue"],
            PL_ROWS["cogs"],
            PL_ROWS["gross_profit"],
            PL_ROWS["personnel_cost"],
            PL_ROWS["marketing_cost"],
            PL_ROWS["development_cost"],
            PL_ROWS["other_opex"],
            PL_ROWS["opex_total"],
            PL_ROWS["operating_profit"],
        ],
        number_format=NUMBER_FORMAT,
    )
    _set_number_format_rows(sheet, [PL_ROWS["gross_margin_ratio"], PL_ROWS["operating_margin"]], number_format=PERCENT_FORMAT)
    _apply_standard_layout(sheet, freeze_panes="B2", label_width=18, year_width=12)
    _set_row_heights(
        sheet,
        {
            PL_ROWS["revenue_total"]: 24,
            PL_ROWS["gross_profit"]: 24,
            PL_ROWS["opex_total"]: 24,
            PL_ROWS["operating_profit"]: 26,
        },
    )


def _write_meal_sheet(sheet) -> None:
    _write_header_row(sheet)
    labels = [
        "価格/アイテム",
        "アイテム/食事",
        "食事数/年",
        "継続率",
        "推定ユニット数",
        "売上",
    ]
    for row_index, label in enumerate(labels, start=2):
        sheet.cell(row=row_index, column=1, value=label)

    mapping = {
        2: "meal_price_per_item",
        3: "meal_items_per_meal",
        4: "meal_meals_per_year",
        5: "meal_retention_rate",
        6: "meal_unit_count",
    }
    for column_index, assumption_col in enumerate(_year_columns(), start=2):
        model_col = excel_col(column_index)
        for row_index, assumption_key in mapping.items():
            sheet[f"{model_col}{row_index}"] = f"={_sheet_ref('（全Ver）前提条件', f'{assumption_col}{ASSUMPTION_ROWS[assumption_key]}')}"
        sheet[f"{model_col}7"] = f"={model_col}2*{model_col}3*{model_col}4*{model_col}6"
    _style_formula_rows(sheet, [2, 3, 4, 5, 6])
    _style_subtotal_rows(sheet, [7])
    _set_number_format_rows(sheet, [2, 3, 4, 6, 7], number_format=NUMBER_FORMAT)
    _set_number_format_rows(sheet, [5], number_format=PERCENT_FORMAT)
    _apply_standard_layout(sheet, freeze_panes="B2", label_width=20, year_width=12)


def _write_academy_sheet(sheet) -> None:
    _write_header_row(sheet)
    for level in ACADEMY_LEVELS:
        rows = ACADEMY_ROW_LAYOUT[level["code"]]
        sheet.cell(row=rows["label"], column=1, value=level["label"])
        sheet.cell(row=rows["price"], column=1, value="単価")
        sheet.cell(row=rows["students"], column=1, value="受講人数")
        sheet.cell(row=rows["certified"], column=1, value="認証人数(期末)")
        sheet.cell(row=rows["revenue"], column=1, value="売上")

    sheet.cell(row=ACADEMY_ROW_LAYOUT["total_students"], column=1, value="アカデミー合計人数")
    sheet.cell(row=ACADEMY_ROW_LAYOUT["total_revenue"], column=1, value="アカデミー合計売上")

    for column_index, assumption_col in enumerate(_year_columns(), start=2):
        model_col = excel_col(column_index)
        total_students_ref = _sheet_ref("（全Ver）前提条件", f"{assumption_col}{ASSUMPTION_ROWS['academy_students']}")
        effective_price_ref = _sheet_ref("（全Ver）前提条件", f"{assumption_col}{ASSUMPTION_ROWS['academy_effective_price']}")
        raw_students = {
            "c": _academy_raw_student_expr("c", assumption_col, None),
            "b": _academy_raw_student_expr("b", assumption_col, excel_col(column_index - 1) if column_index > 2 else None),
            "a": _academy_raw_student_expr("a", assumption_col, excel_col(column_index - 1) if column_index > 2 else None),
            "s": _academy_raw_student_expr("s", assumption_col, excel_col(column_index - 1) if column_index > 2 else None),
        }
        raw_total = "+".join(f"({expr})" for expr in raw_students.values())

        for level in ACADEMY_LEVELS:
            code = level["code"]
            rows = ACADEMY_ROW_LAYOUT[code]
            students_cell = f"{model_col}{rows['students']}"
            price_cell = f"{model_col}{rows['price']}"
            certified_cell = f"{model_col}{rows['certified']}"
            revenue_cell = f"{model_col}{rows['revenue']}"

            sheet[students_cell] = (
                f"=IF(({raw_total})<>0,{total_students_ref}*({raw_students[code]})/({raw_total}),0)"
            )
            completion_ref = _sheet_ref("（全Ver）前提条件", f"{assumption_col}{ASSUMPTION_ROWS[f'academy_{code}_completion']}")
            certification_ref = _sheet_ref("（全Ver）前提条件", f"{assumption_col}{ASSUMPTION_ROWS[f'academy_{code}_certification']}")
            multiplier_ref = _sheet_ref("（全Ver）前提条件", f"{assumption_col}{ASSUMPTION_ROWS[f'academy_{code}_price_multiplier']}")

            weighted_terms: list[str] = []
            for other in ACADEMY_LEVELS:
                other_code = other["code"]
                other_multiplier_ref = _sheet_ref(
                    "（全Ver）前提条件",
                    f"{assumption_col}{ASSUMPTION_ROWS[f'academy_{other_code}_price_multiplier']}",
                )
                other_students_cell = f"{model_col}{ACADEMY_ROW_LAYOUT[other_code]['students']}"
                weighted_terms.append(f"{other_multiplier_ref}*{other_students_cell}")
            weighted_multiplier = "+".join(weighted_terms)
            sheet[price_cell] = (
                f"=IF(({weighted_multiplier})<>0,{effective_price_ref}*{multiplier_ref}*{total_students_ref}/({weighted_multiplier}),0)"
            )
            sheet[certified_cell] = f"={students_cell}*{completion_ref}*{certification_ref}"
            sheet[revenue_cell] = f"={price_cell}*{students_cell}"

        sheet[f"{model_col}{ACADEMY_ROW_LAYOUT['total_students']}"] = (
            f"=SUM({model_col}{ACADEMY_ROW_LAYOUT['c']['students']},{model_col}{ACADEMY_ROW_LAYOUT['b']['students']},"
            f"{model_col}{ACADEMY_ROW_LAYOUT['a']['students']},{model_col}{ACADEMY_ROW_LAYOUT['s']['students']})"
        )
        sheet[f"{model_col}{ACADEMY_ROW_LAYOUT['total_revenue']}"] = (
            f"=SUM({model_col}{ACADEMY_ROW_LAYOUT['c']['revenue']},{model_col}{ACADEMY_ROW_LAYOUT['b']['revenue']},"
            f"{model_col}{ACADEMY_ROW_LAYOUT['a']['revenue']},{model_col}{ACADEMY_ROW_LAYOUT['s']['revenue']})"
        )
    _style_formula_rows(sheet, [3, 4, 5, 6, 8, 9, 10, 11, 13, 14, 15, 16, 18, 19, 20, 21])
    _style_subtotal_rows(sheet, [ACADEMY_ROW_LAYOUT["total_students"]])
    _style_total_rows(sheet, [ACADEMY_ROW_LAYOUT["total_revenue"]])
    _set_number_format_rows(sheet, [3, 4, 5, 6, 8, 9, 10, 11, 13, 14, 15, 16, 18, 19, 20, 21, ACADEMY_ROW_LAYOUT["total_students"], ACADEMY_ROW_LAYOUT["total_revenue"]], number_format=NUMBER_FORMAT)
    _apply_standard_layout(sheet, freeze_panes="B2", label_width=18, year_width=12)
    _set_row_heights(
        sheet,
        {
            2: 24,
            7: 24,
            12: 24,
            17: 24,
            ACADEMY_ROW_LAYOUT["total_students"]: 24,
            ACADEMY_ROW_LAYOUT["total_revenue"]: 26,
        },
    )


def _write_consulting_sheet(sheet) -> None:
    year_count_columns = [excel_col(column_index) for column_index in range(8, 13)]
    year_revenue_columns = [excel_col(column_index) for column_index in range(13, 18)]
    year_cost_columns = [excel_col(column_index) for column_index in range(18, 23)]

    for column_index, header in enumerate(["SKU", "サービス名", "単位", "単価（円）", "継続率", "デリバリー原価単価", "標準時間"], start=1):
        sheet.cell(row=2, column=column_index, value=header)
        sheet.cell(row=2, column=column_index).font = Font(bold=True)

    for idx, year in enumerate(YEAR_HEADERS):
        sheet.cell(row=1, column=8 + idx, value=f"{year} 件数")
        sheet.cell(row=1, column=13 + idx, value=f"{year} 売上")
        sheet.cell(row=1, column=18 + idx, value=f"{year} 原価")
        sheet.cell(row=1, column=8 + idx).font = BOLD_FONT
        sheet.cell(row=1, column=13 + idx).font = BOLD_FONT
        sheet.cell(row=1, column=18 + idx).font = BOLD_FONT

    unit_price_scale_ref = _sheet_ref("（全Ver）前提条件", f"B{ASSUMPTION_ROWS['consult_unit_price']}")
    retention_ref = _sheet_ref("（全Ver）前提条件", f"B{ASSUMPTION_ROWS['consult_retention']}")
    blended_hourly_rate_ref = _sheet_ref("（全Ver）前提条件", f"B{ASSUMPTION_ROWS['blended_hourly_rate']}")

    for offset, sku in enumerate(CONSULT_SKUS):
        row_index = CONSULT_DETAIL_START_ROW + offset
        sheet.cell(row=row_index, column=1, value=sku["sku"])
        sheet.cell(row=row_index, column=2, value=sku["service_name"])
        sheet.cell(row=row_index, column=3, value=sku["unit"])
        price_multiplier = sku["base_price"] / CONSULT_SKUS[0]["base_price"] if CONSULT_SKUS[0]["base_price"] else 0
        sheet.cell(row=row_index, column=4, value=f"={unit_price_scale_ref}*{price_multiplier}")
        sheet.cell(row=row_index, column=5, value=f"=MIN(1,{retention_ref}*{sku['retention_multiplier']})")
        sheet.cell(row=row_index, column=7, value=sku["standard_hours"])
        sheet.cell(row=row_index, column=6, value=f"=G{row_index}*{blended_hourly_rate_ref}")

        for year_idx, year_col in enumerate(_year_columns()):
            share_row = ASSUMPTION_ROWS[f"consult_{sku['sku'].lower()}_share"]
            share_ref = _sheet_ref(
                "（全Ver）前提条件",
                f"{year_col}{share_row}",
            )
            target_revenue_ref = _sheet_ref("（全Ver）前提条件", f"{year_col}{ASSUMPTION_ROWS['consult_revenue']}")
            count_col = year_count_columns[year_idx]
            revenue_col = year_revenue_columns[year_idx]
            cost_col = year_cost_columns[year_idx]
            sheet[f"{count_col}{row_index}"] = f"=IF($D{row_index}<>0,({target_revenue_ref}*{share_ref})/$D{row_index},0)"
            sheet[f"{revenue_col}{row_index}"] = f"={count_col}{row_index}*$D{row_index}"
            sheet[f"{cost_col}{row_index}"] = f"={count_col}{row_index}*$F{row_index}"

    sheet.cell(row=CONSULT_TOTAL_REVENUE_ROW, column=1, value="コンサル売上合計")
    sheet.cell(row=CONSULT_TOTAL_DELIVERY_ROW, column=1, value="デリバリー原価合計")
    sheet.cell(row=CONSULT_TOTAL_GROSS_PROFIT_ROW, column=1, value="コンサル粗利")
    for column_index, model_col in enumerate(_year_columns(), start=13):
        summary_col = excel_col(column_index - 11)
        cost_col = excel_col(column_index + 5)
        revenue_col = excel_col(column_index)
        sheet[f"{summary_col}{CONSULT_TOTAL_REVENUE_ROW}"] = f"=SUM({revenue_col}{CONSULT_DETAIL_START_ROW}:{revenue_col}{CONSULT_DETAIL_START_ROW + len(CONSULT_SKUS) - 1})"
        sheet[f"{summary_col}{CONSULT_TOTAL_DELIVERY_ROW}"] = f"=SUM({cost_col}{CONSULT_DETAIL_START_ROW}:{cost_col}{CONSULT_DETAIL_START_ROW + len(CONSULT_SKUS) - 1})"
        sheet[f"{summary_col}{CONSULT_TOTAL_GROSS_PROFIT_ROW}"] = f"={summary_col}{CONSULT_TOTAL_REVENUE_ROW}-{summary_col}{CONSULT_TOTAL_DELIVERY_ROW}"
    for row_index in range(CONSULT_DETAIL_START_ROW, CONSULT_DETAIL_START_ROW + len(CONSULT_SKUS)):
        _apply_row_style(sheet, row_index, 4, 6, fill=FORMULA_FILL)
        _apply_row_style(sheet, row_index, 8, 22, fill=FORMULA_FILL)
    _style_subtotal_rows(sheet, [CONSULT_TOTAL_REVENUE_ROW, CONSULT_TOTAL_DELIVERY_ROW], start_col=1, end_col=17)
    _style_total_rows(sheet, [CONSULT_TOTAL_GROSS_PROFIT_ROW], start_col=1, end_col=17)
    for row_index in range(CONSULT_DETAIL_START_ROW, CONSULT_DETAIL_START_ROW + len(CONSULT_SKUS)):
        _set_number_format_rows(sheet, [row_index], number_format=NUMBER_FORMAT, start_col=4, end_col=4)
        _set_number_format_rows(sheet, [row_index], number_format=PERCENT_FORMAT, start_col=5, end_col=5)
        _set_number_format_rows(sheet, [row_index], number_format=NUMBER_FORMAT, start_col=6, end_col=7)
        _set_number_format_rows(sheet, [row_index], number_format=COUNT_DECIMAL_FORMAT, start_col=8, end_col=12)
        _set_number_format_rows(sheet, [row_index], number_format=NUMBER_FORMAT, start_col=13, end_col=22)
    _set_number_format_rows(sheet, [CONSULT_TOTAL_REVENUE_ROW, CONSULT_TOTAL_DELIVERY_ROW, CONSULT_TOTAL_GROSS_PROFIT_ROW], number_format=NUMBER_FORMAT)
    sheet.freeze_panes = "A3"
    _set_column_widths(
        sheet,
        {
            "A": 10,
            "B": 28,
            "C": 12,
            "D": 14,
            "E": 12,
            "F": 16,
            "G": 12,
            "H": 12,
            "I": 12,
            "J": 12,
            "K": 12,
            "L": 12,
            "M": 14,
            "N": 14,
            "O": 14,
            "P": 14,
            "Q": 14,
            "R": 14,
            "S": 14,
            "T": 14,
            "U": 14,
            "V": 14,
        },
    )
    _set_row_heights(sheet, {1: 22, 2: 22, CONSULT_TOTAL_REVENUE_ROW: 24, CONSULT_TOTAL_DELIVERY_ROW: 24, CONSULT_TOTAL_GROSS_PROFIT_ROW: 26})
    for row_index in range(CONSULT_DETAIL_START_ROW, CONSULT_DETAIL_START_ROW + len(CONSULT_SKUS)):
        sheet.row_dimensions[row_index].height = 32
    _align_range(sheet, start_row=1, end_row=sheet.max_row, start_col=1, end_col=1, alignment=LEFT_ALIGN)
    _align_range(sheet, start_row=1, end_row=sheet.max_row, start_col=2, end_col=2, alignment=WRAP_LEFT_ALIGN)
    _align_range(sheet, start_row=1, end_row=sheet.max_row, start_col=3, end_col=22, alignment=RIGHT_ALIGN)


def _write_cost_plan_sheet(sheet, assumptions: dict[str, list[float]]) -> None:
    sheet.cell(row=COST_PLAN_ROWS["summary_title"], column=1, value="費用サマリー")
    _apply_row_style(sheet, COST_PLAN_ROWS["summary_title"], 1, 6, fill=SUBTOTAL_FILL, font=BOLD_FONT)
    sheet.cell(row=COST_PLAN_ROWS["summary_header"], column=1, value="項目")
    sheet.cell(row=COST_PLAN_ROWS["summary_header"], column=1).font = BOLD_FONT
    for column_index, header in enumerate(YEAR_HEADERS, start=2):
        sheet.cell(row=COST_PLAN_ROWS["summary_header"], column=column_index, value=header)
        sheet.cell(row=COST_PLAN_ROWS["summary_header"], column=column_index).font = BOLD_FONT
    labels = {
        COST_SUMMARY_ROWS["personnel_cost"]: "人件費",
        COST_SUMMARY_ROWS["marketing_cost"]: "マーケ費",
        COST_SUMMARY_ROWS["development_cost"]: "開発費（PL計上・償却）",
        COST_SUMMARY_ROWS["other_opex"]: "その他OPEX",
        COST_SUMMARY_ROWS["opex_total"]: "OPEX合計",
    }
    for row_index, label in labels.items():
        sheet.cell(row=row_index, column=1, value=label)

    ratio_map = {
        COST_SUMMARY_ROWS["personnel_cost"]: "personnel_ratio",
        COST_SUMMARY_ROWS["marketing_cost"]: "marketing_ratio",
        COST_SUMMARY_ROWS["development_cost"]: "development_ratio",
        COST_SUMMARY_ROWS["other_opex"]: "other_ratio",
    }
    for column_index, assumption_col in enumerate(_year_columns(), start=2):
        model_col = excel_col(column_index)
        opex_ref = _sheet_ref("（全Ver）前提条件", f"{assumption_col}{ASSUMPTION_ROWS['opex_target']}")
        personnel_ratio_ref = _sheet_ref("（全Ver）前提条件", f"{assumption_col}{ASSUMPTION_ROWS['personnel_ratio']}")
        marketing_ratio_ref = _sheet_ref("（全Ver）前提条件", f"{assumption_col}{ASSUMPTION_ROWS['marketing_ratio']}")
        other_ratio_ref = _sheet_ref("（全Ver）前提条件", f"{assumption_col}{ASSUMPTION_ROWS['other_ratio']}")
        development_amortization_ref = _sheet_ref(
            "（全Ver）前提条件",
            f"{assumption_col}{ASSUMPTION_ROWS['development_amortization_expense']}",
        )
        for row_index, ratio_key in ratio_map.items():
            ratio_ref = {
                "personnel_ratio": personnel_ratio_ref,
                "marketing_ratio": marketing_ratio_ref,
                "development_ratio": development_amortization_ref,
                "other_ratio": other_ratio_ref,
            }[ratio_key]
            if ratio_key == "development_ratio":
                sheet[f"{model_col}{row_index}"] = f"={ratio_ref}"
            else:
                sheet[f"{model_col}{row_index}"] = f"={opex_ref}*{ratio_ref}"
        sheet[f"{model_col}{COST_SUMMARY_ROWS['opex_total']}"] = (
            f"=SUM({model_col}{COST_SUMMARY_ROWS['personnel_cost']}:{model_col}{COST_SUMMARY_ROWS['other_opex']})"
        )
    _style_subtotal_rows(
        sheet,
        [
            COST_SUMMARY_ROWS["personnel_cost"],
            COST_SUMMARY_ROWS["marketing_cost"],
            COST_SUMMARY_ROWS["development_cost"],
            COST_SUMMARY_ROWS["other_opex"],
        ],
    )
    _style_total_rows(sheet, [COST_SUMMARY_ROWS["opex_total"]])
    _set_number_format_rows(
        sheet,
        [
            COST_SUMMARY_ROWS["personnel_cost"],
            COST_SUMMARY_ROWS["marketing_cost"],
            COST_SUMMARY_ROWS["development_cost"],
            COST_SUMMARY_ROWS["other_opex"],
            COST_SUMMARY_ROWS["opex_total"],
        ],
        number_format=NUMBER_FORMAT,
    )
    sheet.cell(row=COST_PLAN_ROWS["amortization_title"], column=1, value="開発償却ブロック")
    _apply_row_style(sheet, COST_PLAN_ROWS["amortization_title"], 1, 7, fill=SUBTOTAL_FILL, font=BOLD_FONT)
    amortization_labels = {
        COST_PLAN_ROWS["amortization_start"]: "開発投資（キャッシュ）",
        COST_PLAN_ROWS["amortization_start"] + 1: "償却方法",
        COST_PLAN_ROWS["amortization_start"] + 2: "償却期間（年）",
        COST_PLAN_ROWS["amortization_start"] + 3: "当期償却額（PL計上）",
        COST_PLAN_ROWS["amortization_start"] + 4: "当期投資の期末未償却残高",
    }
    for row_index, label in amortization_labels.items():
        sheet.cell(row=row_index, column=1, value=label)
    for column_index, assumption_col in enumerate(_year_columns(), start=2):
        model_col = excel_col(column_index)
        development_investment_ref = _sheet_ref(
            "（全Ver）前提条件",
            f"{assumption_col}{ASSUMPTION_ROWS['development_investment']}",
        )
        development_expense_ref = _sheet_ref(
            "（全Ver）前提条件",
            f"{assumption_col}{ASSUMPTION_ROWS['development_amortization_expense']}",
        )
        development_unamortized_ref = _sheet_ref(
            "（全Ver）前提条件",
            f"{assumption_col}{ASSUMPTION_ROWS['development_unamortized_balance']}",
        )
        sheet[f"{model_col}{COST_PLAN_ROWS['amortization_start'] + 1}"] = "定額法"
        sheet[f"{model_col}{COST_PLAN_ROWS['amortization_start']}"] = f"={development_investment_ref}"
        sheet[f"{model_col}{COST_PLAN_ROWS['amortization_start'] + 2}"] = assumptions["development_amortization_years"][column_index - 2]
        sheet[f"{model_col}{COST_PLAN_ROWS['amortization_start'] + 3}"] = f"={development_expense_ref}"
        sheet[f"{model_col}{COST_PLAN_ROWS['amortization_start'] + 4}"] = f"={development_unamortized_ref}"
    _style_formula_rows(
        sheet,
        [
            COST_PLAN_ROWS["amortization_start"],
            COST_PLAN_ROWS["amortization_start"] + 3,
            COST_PLAN_ROWS["amortization_start"] + 4,
        ],
    )
    _style_input_rows(sheet, [COST_PLAN_ROWS["amortization_start"] + 2])
    _set_number_format_rows(
        sheet,
        [COST_PLAN_ROWS["amortization_start"], COST_PLAN_ROWS["amortization_start"] + 3, COST_PLAN_ROWS["amortization_start"] + 4],
        number_format=NUMBER_FORMAT,
    )
    _set_number_format_rows(sheet, [COST_PLAN_ROWS["amortization_start"] + 2], number_format=NUMBER_FORMAT)
    sheet.cell(row=COST_PLAN_ROWS["detail_title"], column=1, value="費用明細")
    _apply_row_style(sheet, COST_PLAN_ROWS["detail_title"], 1, 7, fill=SUBTOTAL_FILL, font=BOLD_FONT)
    detail_header_row = COST_PLAN_ROWS["detail_header"]
    sheet.cell(row=detail_header_row, column=1, value="アイテム")
    for column_index, header in enumerate(YEAR_HEADERS, start=2):
        sheet.cell(row=detail_header_row, column=column_index, value=header)
    sheet.cell(row=detail_header_row, column=7, value="カテゴリ")
    sheet["A10"].font = Font(bold=True)
    sheet["G10"].font = Font(bold=True)
    for column_index in range(2, 7):
        sheet.cell(row=detail_header_row, column=column_index).font = Font(bold=True)

    for row_index, (item_name, category_name, share, cost_key) in enumerate(COST_LIST_ROWS, start=COST_PLAN_ROWS["detail_start"]):
        sheet.cell(row=row_index, column=1, value=item_name)
        sheet.cell(row=row_index, column=7, value=category_name)
        cost_row = COST_SUMMARY_ROWS[cost_key]
        for column_index in range(2, 7):
            model_col = excel_col(column_index)
            sheet[f"{model_col}{row_index}"] = f"={_sheet_ref('費用計画', f'{model_col}{cost_row}')}*{share}"
    for row_index in range(COST_PLAN_ROWS["detail_start"], COST_PLAN_ROWS["detail_start"] + len(COST_LIST_ROWS)):
        _apply_row_style(sheet, row_index, 2, 6, fill=FORMULA_FILL)
        _set_number_format_rows(sheet, [row_index], number_format=NUMBER_FORMAT, start_col=2, end_col=6)
    _apply_standard_layout(sheet, freeze_panes="B2", label_width=22, year_width=12)
    _set_column_widths(sheet, {"A": 24, "B": 12, "C": 12, "D": 12, "E": 12, "F": 12, "G": 14})
    _set_row_heights(
        sheet,
        {
            COST_SUMMARY_ROWS["opex_total"]: 24,
            COST_PLAN_ROWS["summary_title"]: 24,
            COST_PLAN_ROWS["amortization_title"]: 24,
            COST_PLAN_ROWS["detail_title"]: 24,
        },
    )
    _align_range(sheet, start_row=1, end_row=sheet.max_row, start_col=1, end_col=1, alignment=LEFT_ALIGN)
    _align_range(sheet, start_row=1, end_row=sheet.max_row, start_col=2, end_col=6, alignment=RIGHT_ALIGN)


def _write_qa_sheet(
    sheet,
    *,
    candidate_id: str,
    candidate_payload: dict[str, Any],
    diagnosis: dict[str, Any],
    assumptions: dict[str, list[float]],
    iteration_summaries: list[dict[str, Any]],
    qa_mode: str,
) -> None:
    headers = [
        "区分",
        "カテゴリ",
        "想定質問",
        "回答",
        "頻度",
        "精度",
        "初回追加Iteration",
        "今回更新Iteration",
        "状態",
        "採用状況",
        "根拠区分",
        "根拠の数値・内容",
        "根拠ソース",
    ]
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column_index, value=header)
        cell.font = BOLD_FONT
        cell.fill = SUBTOTAL_FILL
        cell.alignment = CENTER_ALIGN if column_index >= 5 else LEFT_ALIGN

    for row_index, item in enumerate(
        _build_workbook_qa_items(
            candidate_id=candidate_id,
            candidate_payload=candidate_payload,
            diagnosis=diagnosis,
            assumptions=assumptions,
            iteration_summaries=iteration_summaries,
            qa_mode=qa_mode,
        ),
        start=2,
    ):
        sheet.cell(row=row_index, column=1, value=item["scope"])
        sheet.cell(row=row_index, column=2, value=item["category"])
        sheet.cell(row=row_index, column=3, value=item["question"])
        sheet.cell(row=row_index, column=4, value=item["answer"])
        sheet.cell(row=row_index, column=5, value=item["frequency"])
        sheet.cell(row=row_index, column=6, value=item["accuracy"])
        sheet.cell(row=row_index, column=7, value=item["first_added_iteration"])
        sheet.cell(row=row_index, column=8, value=item["last_updated_iteration"])
        sheet.cell(row=row_index, column=9, value=item["status"])
        sheet.cell(row=row_index, column=10, value=item["adoption"])
        sheet.cell(row=row_index, column=11, value=item["support_categories"])
        sheet.cell(row=row_index, column=12, value=item["support_details"])
        sheet.cell(row=row_index, column=13, value=item["support_sources"])
        if item["status"] == "新規":
            sheet.cell(row=row_index, column=9).fill = FORMULA_FILL
        elif item["status"] == "更新":
            sheet.cell(row=row_index, column=9).fill = UPDATED_FILL
        if item["adoption"] == "今回採用":
            sheet.cell(row=row_index, column=10).fill = INPUT_FILL
        elif item["adoption"] == "比較のみ":
            sheet.cell(row=row_index, column=10).fill = SUBTOTAL_FILL

    sheet.freeze_panes = "A2"
    _set_column_widths(
        sheet,
        {
            "A": 12,
            "B": 12,
            "C": 30,
            "D": 52,
            "E": 8,
            "F": 8,
            "G": 14,
            "H": 14,
            "I": 10,
            "J": 12,
            "K": 16,
            "L": 42,
            "M": 34,
        },
    )
    _align_range(sheet, start_row=1, end_row=sheet.max_row, start_col=1, end_col=2, alignment=CENTER_ALIGN)
    _align_range(sheet, start_row=1, end_row=sheet.max_row, start_col=3, end_col=4, alignment=WRAP_LEFT_ALIGN)
    _align_range(sheet, start_row=1, end_row=sheet.max_row, start_col=5, end_col=6, alignment=CENTER_ALIGN)
    _align_range(sheet, start_row=1, end_row=sheet.max_row, start_col=7, end_col=8, alignment=RIGHT_ALIGN)
    _align_range(sheet, start_row=1, end_row=sheet.max_row, start_col=9, end_col=11, alignment=CENTER_ALIGN)
    _align_range(sheet, start_row=1, end_row=sheet.max_row, start_col=12, end_col=13, alignment=WRAP_LEFT_ALIGN)


def _write_plan_assumptions_sheet(sheet, assumptions: dict[str, list[float]]) -> None:
    sheet.cell(row=1, column=1, value="項目")
    for column_index, header in enumerate(YEAR_HEADERS, start=2):
        sheet.cell(row=1, column=column_index, value=header)
    sheet.cell(row=1, column=7, value="説明")
    for cell in ("A1", "G1", "B1", "C1", "D1", "E1", "F1"):
        sheet[cell].font = Font(bold=True)

    row_specs = [
        ("revenue_target", "売上目標", "candidate の PL 売上系列"),
        ("gross_profit_target", "粗利目標", "candidate の PL 粗利系列"),
        ("opex_target", "OPEX目標", "candidate の PL OPEX 系列"),
        ("academy_public_price", "アカデミー公開単価", "抽出または補完した公開単価"),
        ("academy_students", "アカデミー受講人数", "抽出した受講人数系列"),
        ("academy_certified", "アカデミー認証人数", "抽出した認証人数系列"),
        ("academy_revenue", "アカデミー売上目標", "抽出または trend 補完した売上系列"),
        ("academy_effective_price", "アカデミー実効単価", "売上目標 ÷ 受講人数"),
        ("meal_price_per_item", "ミール価格/アイテム", "industry 補完または抽出値"),
        ("meal_items_per_meal", "ミールアイテム/食事", "industry 補完または抽出値"),
        ("meal_meals_per_year", "ミール食事数/年", "industry 補完または抽出値"),
        ("meal_retention_rate", "ミール継続率", "industry 補完または抽出値"),
        ("meal_share_non_academy", "ミール残余売上比率", "非アカデミー残余売上に対するミール比率の透明な仮定"),
        ("meal_unit_count", "ミール推定ユニット数", "残余売上比率から逆算した推定ユニット数"),
        ("meal_revenue", "ミール売上", "ユニットエコノミクスから計算した売上"),
        ("consult_unit_price", "コンサルSKU単価", "benchmark または抽出値"),
        ("consult_retention", "コンサル継続率", "benchmark または抽出値"),
        ("consult_standard_hours", "コンサル標準工数", "benchmark または抽出値"),
        ("consult_revenue", "コンサル売上", "総売上の残余から計算した売上"),
        ("consult_project_count", "コンサル推定案件数", "コンサル売上 ÷ SKU単価"),
        ("gross_margin_ratio", "粗利率", "粗利目標 ÷ 売上目標"),
        ("personnel_ratio", "人件費比率", "OPEX のうち人件費の比率"),
        ("marketing_ratio", "マーケ費比率", "OPEX のうちマーケ費の比率"),
        ("development_ratio", "開発費比率", "OPEX のうち開発費の比率"),
        ("other_ratio", "その他費用比率", "OPEX のうちその他費用の比率"),
        ("development_investment", "開発投資額（キャッシュ）", "PL計上する開発償却費と償却年数から逆算した投資額"),
        ("development_amortization_years", "開発償却期間（年）", "開発投資の初期値。定額法で5年償却"),
        ("development_amortization_expense", "PL計上開発費（償却）", "OPEX比率から計算した当期の償却費"),
        ("development_unamortized_balance", "開発投資の期末未償却残高", "当期投資額から当期償却額を差し引いた残高"),
        ("academy_c_share", "C級新規構成比", "アカデミー総受講人数のうち C級に配分する比率"),
        ("academy_b_share", "B級新規構成比", "アカデミー総受講人数のうち B級に配分する比率"),
        ("academy_a_share", "A級新規構成比", "アカデミー総受講人数のうち A級に配分する比率"),
        ("academy_s_share", "S級新規構成比", "アカデミー総受講人数のうち S級に配分する比率"),
        ("academy_c_to_b", "C→B進級率", "前年度のC級認証者がB級へ進む比率"),
        ("academy_b_to_a", "B→A進級率", "前年度のB級認証者がA級へ進む比率"),
        ("academy_a_to_s", "A→S進級率", "前年度のA級認証者がS級へ進む比率"),
        ("academy_c_price_multiplier", "C級価格倍率", "アカデミー実効単価に対する C級倍率"),
        ("academy_b_price_multiplier", "B級価格倍率", "アカデミー実効単価に対する B級倍率"),
        ("academy_a_price_multiplier", "A級価格倍率", "アカデミー実効単価に対する A級倍率"),
        ("academy_s_price_multiplier", "S級価格倍率", "アカデミー実効単価に対する S級倍率"),
        ("academy_c_completion", "C級修了率", "C級の修了率仮定"),
        ("academy_c_certification", "C級認証率", "C級の認証率仮定"),
        ("academy_b_completion", "B級修了率", "B級の修了率仮定"),
        ("academy_b_certification", "B級認証率", "B級の認証率仮定"),
        ("academy_a_completion", "A級修了率", "A級の修了率仮定"),
        ("academy_a_certification", "A級認証率", "A級の認証率仮定"),
        ("academy_s_completion", "S級修了率", "S級の修了率仮定"),
        ("academy_s_certification", "S級認証率", "S級の認証率仮定"),
        ("blended_hourly_rate", "ブレンド時給", "コンサルのデリバリー原価に使う透明な時給仮定"),
        ("consult_p1_share", "P1売上構成比", "コンサル売上のうち P1 に配分する比率"),
        ("consult_p2_share", "P2売上構成比", "コンサル売上のうち P2 に配分する比率"),
        ("consult_p3_share", "P3売上構成比", "コンサル売上のうち P3 に配分する比率"),
        ("consult_p4_share", "P4売上構成比", "コンサル売上のうち P4 に配分する比率"),
        ("consult_p5_share", "P5売上構成比", "コンサル売上のうち P5 に配分する比率"),
        ("consult_p6_share", "P6売上構成比", "コンサル売上のうち P6 に配分する比率"),
        ("consult_p8_share", "P8売上構成比", "コンサル売上のうち P8 に配分する比率"),
        ("consult_p9_share", "P9売上構成比", "コンサル売上のうち P9 に配分する比率"),
        ("consult_p10_share", "P10売上構成比", "コンサル売上のうち P10 に配分する比率"),
        ("consult_p11_share", "P11売上構成比", "コンサル売上のうち P11 に配分する比率"),
        ("consult_p12_share", "P12売上構成比", "コンサル売上のうち P12 に配分する比率"),
    ]

    for key, label, description in row_specs:
        row_index = ASSUMPTION_ROWS[key]
        sheet.cell(row=row_index, column=1, value=label)
        sheet.cell(row=row_index, column=7, value=description)

    for key, series in assumptions.items():
        row_index = ASSUMPTION_ROWS[key]
        for column_index, value in enumerate(series, start=2):
            sheet.cell(row=row_index, column=column_index, value=value)

    for column_index, year_col in enumerate(_year_columns(), start=2):
        col = excel_col(column_index)
        sheet[f"{col}{ASSUMPTION_ROWS['academy_effective_price']}"] = (
            f"=IF({col}{ASSUMPTION_ROWS['academy_students']}<>0,{col}{ASSUMPTION_ROWS['academy_revenue']}/{col}{ASSUMPTION_ROWS['academy_students']},0)"
        )
        sheet[f"{col}{ASSUMPTION_ROWS['meal_unit_count']}"] = (
            f"=MAX(0,(({col}{ASSUMPTION_ROWS['revenue_target']}-{col}{ASSUMPTION_ROWS['academy_revenue']})*"
            f"{col}{ASSUMPTION_ROWS['meal_share_non_academy']})/"
            f"({col}{ASSUMPTION_ROWS['meal_price_per_item']}*{col}{ASSUMPTION_ROWS['meal_items_per_meal']}*{col}{ASSUMPTION_ROWS['meal_meals_per_year']}))"
        )
        sheet[f"{col}{ASSUMPTION_ROWS['meal_revenue']}"] = (
            f"={col}{ASSUMPTION_ROWS['meal_price_per_item']}*{col}{ASSUMPTION_ROWS['meal_items_per_meal']}*"
            f"{col}{ASSUMPTION_ROWS['meal_meals_per_year']}*{col}{ASSUMPTION_ROWS['meal_unit_count']}"
        )
        sheet[f"{col}{ASSUMPTION_ROWS['consult_revenue']}"] = (
            f"=MAX(0,{col}{ASSUMPTION_ROWS['revenue_target']}-{col}{ASSUMPTION_ROWS['academy_revenue']}-{col}{ASSUMPTION_ROWS['meal_revenue']})"
        )
        sheet[f"{col}{ASSUMPTION_ROWS['consult_project_count']}"] = (
            f"=IF({col}{ASSUMPTION_ROWS['consult_unit_price']}<>0,{col}{ASSUMPTION_ROWS['consult_revenue']}/{col}{ASSUMPTION_ROWS['consult_unit_price']},0)"
        )
        sheet[f"{col}{ASSUMPTION_ROWS['gross_margin_ratio']}"] = (
            f"=IF({col}{ASSUMPTION_ROWS['revenue_target']}<>0,{col}{ASSUMPTION_ROWS['gross_profit_target']}/{col}{ASSUMPTION_ROWS['revenue_target']},0)"
        )
        sheet[f"{col}{ASSUMPTION_ROWS['development_amortization_expense']}"] = (
            f"={col}{ASSUMPTION_ROWS['opex_target']}*{col}{ASSUMPTION_ROWS['development_ratio']}"
        )
        sheet[f"{col}{ASSUMPTION_ROWS['development_investment']}"] = (
            f"={col}{ASSUMPTION_ROWS['development_amortization_expense']}*{col}{ASSUMPTION_ROWS['development_amortization_years']}"
        )
        sheet[f"{col}{ASSUMPTION_ROWS['development_unamortized_balance']}"] = (
            f"=MAX(0,{col}{ASSUMPTION_ROWS['development_investment']}-{col}{ASSUMPTION_ROWS['development_amortization_expense']})"
        )
    derived_rows = [
        ASSUMPTION_ROWS["academy_effective_price"],
        ASSUMPTION_ROWS["meal_unit_count"],
        ASSUMPTION_ROWS["meal_revenue"],
        ASSUMPTION_ROWS["consult_revenue"],
        ASSUMPTION_ROWS["consult_project_count"],
        ASSUMPTION_ROWS["gross_margin_ratio"],
        ASSUMPTION_ROWS["development_investment"],
        ASSUMPTION_ROWS["development_amortization_expense"],
        ASSUMPTION_ROWS["development_unamortized_balance"],
    ]
    input_rows = [
        row_index
        for row_key, row_index in ASSUMPTION_ROWS.items()
        if row_key
        not in {
            "academy_effective_price",
            "meal_unit_count",
            "meal_revenue",
            "consult_revenue",
            "consult_project_count",
            "gross_margin_ratio",
            "development_investment",
            "development_amortization_expense",
            "development_unamortized_balance",
        }
    ]
    _style_input_rows(sheet, input_rows)
    _style_formula_rows(sheet, derived_rows)
    percentage_rows = [
        ASSUMPTION_ROWS["meal_retention_rate"],
        ASSUMPTION_ROWS["consult_retention"],
        ASSUMPTION_ROWS["gross_margin_ratio"],
        ASSUMPTION_ROWS["personnel_ratio"],
        ASSUMPTION_ROWS["marketing_ratio"],
        ASSUMPTION_ROWS["development_ratio"],
        ASSUMPTION_ROWS["other_ratio"],
        ASSUMPTION_ROWS["academy_c_share"],
        ASSUMPTION_ROWS["academy_b_share"],
        ASSUMPTION_ROWS["academy_a_share"],
        ASSUMPTION_ROWS["academy_s_share"],
        ASSUMPTION_ROWS["academy_c_to_b"],
        ASSUMPTION_ROWS["academy_b_to_a"],
        ASSUMPTION_ROWS["academy_a_to_s"],
        ASSUMPTION_ROWS["academy_c_completion"],
        ASSUMPTION_ROWS["academy_c_certification"],
        ASSUMPTION_ROWS["academy_b_completion"],
        ASSUMPTION_ROWS["academy_b_certification"],
        ASSUMPTION_ROWS["academy_a_completion"],
        ASSUMPTION_ROWS["academy_a_certification"],
        ASSUMPTION_ROWS["academy_s_completion"],
        ASSUMPTION_ROWS["academy_s_certification"],
    ]
    decimal_count_rows = [ASSUMPTION_ROWS["consult_project_count"]]
    integer_input_rows = [ASSUMPTION_ROWS["development_amortization_years"]]
    for row_index in input_rows + derived_rows:
        fmt = PERCENT_FORMAT if row_index in percentage_rows else NUMBER_FORMAT
        if row_index in decimal_count_rows:
            fmt = COUNT_DECIMAL_FORMAT
        if row_index in integer_input_rows:
            fmt = NUMBER_FORMAT
        _set_number_format_rows(sheet, [row_index], number_format=fmt)
    sheet.freeze_panes = "B2"
    _set_column_widths(
        sheet,
        {
            "A": 20,
            "B": 12,
            "C": 12,
            "D": 12,
            "E": 12,
            "F": 12,
            "G": 34,
        },
    )
    _set_row_heights(sheet, {1: 22})
    _align_range(sheet, start_row=1, end_row=sheet.max_row, start_col=1, end_col=1, alignment=LEFT_ALIGN)
    _align_range(sheet, start_row=1, end_row=sheet.max_row, start_col=2, end_col=6, alignment=RIGHT_ALIGN)
    _align_range(sheet, start_row=1, end_row=sheet.max_row, start_col=7, end_col=7, alignment=WRAP_LEFT_ALIGN)


def _academy_raw_student_expr(code: str, assumption_col: str, previous_model_col: str | None) -> str:
    total_students_ref = _sheet_ref("（全Ver）前提条件", f"{assumption_col}{ASSUMPTION_ROWS['academy_students']}")
    share_ref = _sheet_ref("（全Ver）前提条件", f"{assumption_col}{ASSUMPTION_ROWS[f'academy_{code}_share']}")
    base_expr = f"{total_students_ref}*{share_ref}"
    if previous_model_col is None:
        return base_expr

    if code == "b":
        previous_ref = f"{previous_model_col}{ACADEMY_ROW_LAYOUT['c']['certified']}"
        progression_ref = _sheet_ref("（全Ver）前提条件", f"{assumption_col}{ASSUMPTION_ROWS['academy_c_to_b']}")
        return f"{base_expr}+{previous_ref}*{progression_ref}"
    if code == "a":
        previous_ref = f"{previous_model_col}{ACADEMY_ROW_LAYOUT['b']['certified']}"
        progression_ref = _sheet_ref("（全Ver）前提条件", f"{assumption_col}{ASSUMPTION_ROWS['academy_b_to_a']}")
        return f"{base_expr}+{previous_ref}*{progression_ref}"
    if code == "s":
        previous_ref = f"{previous_model_col}{ACADEMY_ROW_LAYOUT['a']['certified']}"
        progression_ref = _sheet_ref("（全Ver）前提条件", f"{assumption_col}{ASSUMPTION_ROWS['academy_a_to_s']}")
        return f"{base_expr}+{previous_ref}*{progression_ref}"
    return base_expr


def _write_assumptions_sheet(sheet, assumptions: list[dict[str, Any]]) -> None:
    headers = ["source_type", "review_status", "evidence_count", "evidence_ids"]
    for column_index, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=column_index, value=header)
        sheet.cell(row=1, column=column_index).font = Font(bold=True)

    for row_index, assumption in enumerate(assumptions, start=2):
        evidence_refs = assumption.get("evidence_refs", [])
        evidence_ids = ", ".join(
            ref.get("source_id") or ref.get("title") or ""
            for ref in evidence_refs
            if ref.get("source_id") or ref.get("title")
        )
        values = [
            assumption.get("source_type", ""),
            assumption.get("review_status", ""),
            len(evidence_refs),
            evidence_ids,
        ]
        for column_index, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=column_index, value=value)


def _write_artifacts_sheet(sheet, *, run_root: Path, candidate_id: str) -> None:
    rows = [
        ["run_root", str(run_root)],
        ["summary", str(run_root / "summary.md")],
        ["scores", str(run_root / "scores.json")],
        ["diagnosis", str(run_root / "diagnosis.json")],
        ["candidate_json", str(run_root / "candidates" / f"{candidate_id}.json")],
        ["reference", str(run_root / "reference.json")],
        ["baseline", str(run_root / "baseline.json")],
    ]
    _write_key_value_rows(sheet, rows)


def _write_section_title(sheet, row_index: int, title: str) -> int:
    sheet.cell(row=row_index, column=1, value=title)
    _apply_row_style(sheet, row_index, 1, 4, fill=SUBTOTAL_FILL, font=BOLD_FONT)
    return row_index + 1


def _write_labeled_values(
    sheet,
    row_index: int,
    rows: list[tuple[str, Any]],
    *,
    numeric_labels: set[str] | None = None,
) -> int:
    numeric_labels = numeric_labels or set()
    for label, value in rows:
        sheet.cell(row=row_index, column=1, value=label)
        target_col = 2 if label in numeric_labels else 4
        sheet.cell(row=row_index, column=target_col, value=value)
        row_index += 1
    return row_index


def _build_workbook_qa_items(
    *,
    candidate_id: str,
    candidate_payload: dict[str, Any],
    diagnosis: dict[str, Any],
    assumptions: dict[str, list[float]],
    iteration_summaries: list[dict[str, Any]],
    qa_mode: str,
) -> list[dict[str, Any]]:
    current_iteration = _current_iteration(iteration_summaries, candidate_id)
    current_summary = next(
        (item for item in iteration_summaries if item.get("candidate_id") == candidate_id),
        {},
    )
    score_layers = diagnosis.get("score", {}).get("layers", {})
    next_actions = diagnosis.get("next_actions") or ["次の改善施策は未設定です。"]
    source_types = ", ".join(diagnosis.get("evidence", {}).get("source_types", [])) or "pdf"
    pdf_facts = diagnosis.get("evidence", {}).get("pdf_facts", []) or []
    external_sources = diagnosis.get("evidence", {}).get("external_sources", []) or []
    external_label = ", ".join(source.get("title", "") for source in external_sources if source.get("title")) or source_types
    segments = list(candidate_payload.get("model_sheets", {}).keys())
    revenue_start = assumptions["revenue_target"][0]
    revenue_end = assumptions["revenue_target"][-1]
    gross_margin_start = (
        assumptions["gross_profit_target"][0] / revenue_start
        if revenue_start
        else 0.0
    )
    gross_margin_end = (
        assumptions["gross_profit_target"][-1] / revenue_end
        if revenue_end
        else 0.0
    )
    meal_retention_start = assumptions["meal_retention_rate"][0]
    consult_retention_start = assumptions["consult_retention"][0]
    personnel_ratio_start = assumptions["personnel_ratio"][0]
    marketing_ratio_start = assumptions["marketing_ratio"][0]
    development_ratio_start = assumptions["development_ratio"][0]
    development_years = assumptions["development_amortization_years"][0]
    pl_delta = score_layers.get("pl", {}).get("delta", 0.0)
    total_delta = diagnosis.get("score", {}).get("delta_vs_baseline", 0.0)

    if qa_mode == "revenue_plan":
        qa_specs = [
            {
                "scope": "一般",
                "category": "収益",
                "question": "売上は価格・数量・継続率のどこで伸ばすのか？",
                "answer": (
                    f"現状の成長仮説は価格改定より数量増が中心で、売上は FY1 {_fmt_int(revenue_start)} から FY5 {_fmt_int(revenue_end)} へ伸ばす計画です。"
                    f" 継続率はミール {_fmt_pct(meal_retention_start)}、コンサル {_fmt_pct(consult_retention_start)} を起点に置いており、まずは数量と継続の維持が蓋然性の核です。"
                ),
                "metrics_to_check": "売上目標 / meal 食数 / academy 受講人数 / consult 数量 / 継続率",
                "evidence_to_check": "PL設計 / 各モデルシート / （全Ver）前提条件",
                "frequency": "高",
                "accuracy": "中",
                "tags": ["sales", "staged"],
            },
            {
                "scope": "一般",
                "category": "収益",
                "question": "その成長は市場成長ではなく、どの顧客獲得仮説で達成するのか？",
                "answer": (
                    "市場の追い風を前提にするのではなく、検証期間の後に営業効率の高いチャネルへ投資を寄せて顧客獲得を伸ばす仮説です。"
                    " ただし市場シェアの取り切り方はまだ粗く、チャネル別獲得件数の裏付けを追加で詰める必要があります。"
                ),
                "metrics_to_check": "sales overlay 差分 / partner 差分 / 売上目標",
                "evidence_to_check": f"PDCA全体推移 / 外部根拠 / {external_label}",
                "frequency": "高",
                "accuracy": "中",
                "tags": ["sales", "partner", "staged"],
            },
            {
                "scope": "一般",
                "category": "収益",
                "question": "売上計画はどの先行指標で裏打ちされているか？",
                "answer": (
                    "現時点で直接追えている先行指標は、アカデミー受講人数、ミールの食数、コンサル売上構成比です。"
                    " 一方でパイプライン量、受注残、営業担当の立ち上がり速度までは未実装なので、投資家説明ではここが弱点になります。"
                ),
                "metrics_to_check": "受講人数 / meal 食数 / consult 構成比 / pipeline 未整備箇所",
                "evidence_to_check": "モデルシート / 次の改善施策",
                "frequency": "高",
                "accuracy": "低",
                "tags": ["sales"],
            },
            {
                "scope": "一般",
                "category": "収益",
                "question": "営業体制と採用計画で本当に取り切れるか？",
                "answer": (
                    "現時点では営業効率を重ねたときに PL が改善することまでは確認できていますが、担当者数・生産性・立ち上がり曲線の明細までは持てていません。"
                    " したがって、営業容量の説明はまだ中身を足す必要があります。"
                ),
                "metrics_to_check": "sales 候補の pl Δ / コンサル数量 / 人件費",
                "evidence_to_check": "PDCA全体推移 / コンサルモデル / 費用計画",
                "frequency": "高",
                "accuracy": "低",
                "tags": ["sales"],
            },
            {
                "scope": "一般",
                "category": "収益",
                "question": "営業・マーケ・パートナーのどのチャネルが主役で、効率はどう違うのか？",
                "answer": (
                    "今回の比較では営業効率が主役、パートナーが準主役、ブランドは補助レバーという順番です。"
                    " 営業を主軸としたのは PL 改善幅が最も大きかったためで、パートナーは補完的に効く位置づけです。"
                ),
                "metrics_to_check": "sales / partner / branding の total Δ / pl Δ",
                "evidence_to_check": "PDCA全体推移 / diagnosis verdict",
                "frequency": "高",
                "accuracy": "中",
                "tags": ["sales", "partner", "branding"],
            },
            {
                "scope": "一般",
                "category": "収益",
                "question": "営業投資を強める前提条件は何か？",
                "answer": (
                    "前提条件は、前半で unit economics の筋が見え、後半で sales overlay を載せても PL が悪化しないことです。"
                    " つまりアクセルを踏む条件は、売上の質と損益の両方で確認する設計になっています。"
                ),
                "metrics_to_check": "staged_acceleration Δ / pl Δ / 営業利益率",
                "evidence_to_check": "PDCA全体推移 / 仮説の詳細",
                "frequency": "高",
                "accuracy": "高",
                "tags": ["staged", "sales"],
            },
            {
                "scope": "一般",
                "category": "成長性",
                "question": "継続率やリピートが計画に与える影響はどの程度か？",
                "answer": (
                    f"継続率はミール {_fmt_pct(meal_retention_start)}、コンサル {_fmt_pct(consult_retention_start)} を起点にしているため、売上の下振れ感応度は小さくありません。"
                    " とくに数量前提が強い計画では、継続率が崩れると想定より早く PL が悪化します。"
                ),
                "metrics_to_check": "meal retention / consult retention / 売上感応度",
                "evidence_to_check": "ミールモデル / コンサルモデル / 前提条件",
                "frequency": "高",
                "accuracy": "中",
                "tags": [],
            },
            {
                "scope": "一般",
                "category": "成長性",
                "question": "市場成長だけでなく、自社が獲得できる根拠は何か？",
                "answer": (
                    "自社が獲得できる根拠として使えているのは、事業構造の分解とチャネル別 overlay の比較までです。"
                    " 顧客セグメント別の商談化率や受注率までは弱いため、現状は『完全に言い切る』より『仮説として説明する』段階です。"
                ),
                "metrics_to_check": "candidate 間比較 / structure / model_sheets / pl",
                "evidence_to_check": f"PDCA全体推移 / {external_label}",
                "frequency": "中",
                "accuracy": "低",
                "tags": ["sales", "partner", "branding", "staged"],
            },
            {
                "scope": "一般",
                "category": "収益性",
                "question": "粗利率は伸びるほど改善する構造か？",
                "answer": (
                    f"現状の粗利率は FY1 {_fmt_pct(gross_margin_start)} から FY5 {_fmt_pct(gross_margin_end)} のレンジで見ています。"
                    " ただし、構造的な改善を言い切るには原価項目の分解がまだ足りず、事業別原価の精緻化が必要です。"
                ),
                "metrics_to_check": "粗利 / 粗利率 / 売上原価",
                "evidence_to_check": "PL設計 / ミールモデル / コンサルモデル",
                "frequency": "高",
                "accuracy": "中",
                "tags": [],
            },
            {
                "scope": "一般",
                "category": "収益性",
                "question": "黒字化の条件は何か？",
                "answer": (
                    "黒字化の条件は売上成長そのものより、営業効率を高めながら OPEX の先行を抑えることです。"
                    f" 今の候補では PL の差分が {pl_delta:+.4f} 改善しており、黒字化に向かう方向性は示せています。"
                ),
                "metrics_to_check": "営業利益 / 営業利益率 / pl Δ / OPEX",
                "evidence_to_check": "PL設計 / PDCA全体推移",
                "frequency": "高",
                "accuracy": "中",
                "tags": ["sales", "staged"],
            },
            {
                "scope": "一般",
                "category": "コスト",
                "question": "費用は売上成長に対してどうコントロールしているか？",
                "answer": (
                    f"費用は人件費 { _fmt_pct(personnel_ratio_start) }、マーケ費 { _fmt_pct(marketing_ratio_start) }、開発費 { _fmt_pct(development_ratio_start) } の比率から管理しています。"
                    " 検証段階で固定費を先行させすぎないことが、計画の基本ルールです。"
                ),
                "metrics_to_check": "人件費 / マーケ費 / 開発費（償却） / OPEX",
                "evidence_to_check": "費用計画 / （全Ver）前提条件",
                "frequency": "高",
                "accuracy": "高",
                "tags": ["staged"],
            },
            {
                "scope": "一般",
                "category": "リスク",
                "question": "下振れした場合、どの前提が先に崩れるか？",
                "answer": (
                    "いちばん先に崩れやすいのは sales overlay の実現、次に consulting 数量ロジック、最後に継続率です。"
                    " つまり下振れはまず獲得効率の未達として現れ、その影響が PL に波及すると見るのが自然です。"
                ),
                "metrics_to_check": "worsened points / pl Δ / コンサル数量 / retention",
                "evidence_to_check": current_summary.get("worsened_points", "-"),
                "frequency": "高",
                "accuracy": "中",
                "tags": ["sales", "staged"],
            },
            {
                "scope": "一般",
                "category": "リスク",
                "question": "追加資金が必要になる条件は何か？",
                "answer": (
                    "追加資金が必要になるのは、営業投資を増やしても売上がついてこず、OPEX だけが先行するケースです。"
                    " この計画では、そのリスクを避けるために検証後アクセルの構造を採っています。"
                ),
                "metrics_to_check": "OPEX / 営業利益 / sales 候補の差分",
                "evidence_to_check": "費用計画 / PDCA全体推移 / 仮説の詳細",
                "frequency": "高",
                "accuracy": "中",
                "tags": ["staged", "sales"],
            },
            {
                "scope": "一般",
                "category": "リスク",
                "question": "感応度が高い前提は何か？",
                "answer": (
                    "感応度が高いのは sales efficiency、partner 寄与、継続率、consulting 数量ロジックです。"
                    " 総合スコアよりも、どのレバーが PL を動かしたかを見て管理すべき計画です。"
                ),
                "metrics_to_check": "候補別 pl Δ / explainability Δ / retention",
                "evidence_to_check": "PDCA全体推移 / モデルシート",
                "frequency": "高",
                "accuracy": "中",
                "tags": ["sales", "partner", "branding", "staged"],
            },
            {
                "scope": "一般",
                "category": "資金",
                "question": "追加投資のアクセルを踏む解放条件は何か？",
                "answer": (
                    "解放条件は、検証期間を経たうえで sales overlay を載せても PL が悪化しないことです。"
                    " したがって投資判断は『成長したいから増やす』ではなく、『再現できると確認できたから増やす』に置いています。"
                ),
                "metrics_to_check": "staged_acceleration Δ / candidate-revenue-staged-sales の total Δ",
                "evidence_to_check": "PDCA全体推移 / verdict",
                "frequency": "高",
                "accuracy": "高",
                "tags": ["staged"],
            },
            {
                "scope": "一般",
                "category": "資金",
                "question": "投資回収の筋はどこで確認するのか？",
                "answer": (
                    "投資回収は、営業利益率と OPEX の伸びのバランスで確認します。"
                    " ただし CAC 回収期間や rep payback のような詳細指標は未実装で、ここは追加で補う必要があります。"
                ),
                "metrics_to_check": "営業利益率 / OPEX / マーケ費 / sales 候補差分",
                "evidence_to_check": "PL設計 / 費用計画 / PDCA全体推移",
                "frequency": "高",
                "accuracy": "低",
                "tags": ["sales", "staged"],
            },
            {
                "scope": "一般",
                "category": "オペレーション",
                "question": "毎月の先行管理KPIとして何を見るべきか？",
                "answer": (
                    "最低限、受講人数、meal の数量、consulting 数量、継続率、営業効率の5つを見るべきです。"
                    " 逆にこの先行指標が揃わない限り、売上計画の蓋然性を強く言い切るのは難しいです。"
                ),
                "metrics_to_check": "受講人数 / meal 数量 / consult 数量 / retention / sales 効率",
                "evidence_to_check": "各モデルシート / PDCA全体推移",
                "frequency": "高",
                "accuracy": "中",
                "tags": ["sales"],
            },
            {
                "scope": "企画書固有",
                "category": "収益",
                "question": "なぜ3年間は検証期間なのか？",
                "answer": (
                    "この計画では、unit economics が確認できる前に営業投資を積み増すと失敗コストが大きいからです。"
                    " 検証期間を置くことで、成長投資の前に『何が本当に効くか』を確かめる意図があります。"
                ),
                "metrics_to_check": "staged_acceleration Δ / acceleration period 単独との差分",
                "evidence_to_check": "PDCA全体推移 / 仮説の詳細",
                "frequency": "高",
                "accuracy": "高",
                "tags": ["staged"],
            },
            {
                "scope": "企画書固有",
                "category": "収益",
                "question": "なぜ sales efficiency を一番強く見ているのか？",
                "answer": (
                    f"比較した候補の中で、sales を重ねた案が総合 {total_delta:+.4f}、PL {pl_delta:+.4f} と最も改善幅が大きかったからです。"
                    " つまり、今の計画では営業効率が最も直接的に収益計画の蓋然性を押し上げています。"
                ),
                "metrics_to_check": "candidate-revenue-staged-sales の total Δ / pl Δ",
                "evidence_to_check": "PDCA全体推移 / diagnosis verdict",
                "frequency": "高",
                "accuracy": "高",
                "tags": ["sales"],
            },
            {
                "scope": "企画書固有",
                "category": "収益",
                "question": "partner や branding はなぜ主軸にしなかったのか？",
                "answer": (
                    "比較上は一定の改善がありましたが、営業効率を重ねた案ほど PL を押し上げませんでした。"
                    " そのため、いまは主軸ではなく補助レバーとして扱うのが妥当です。"
                ),
                "metrics_to_check": "partner / branding 候補との差分",
                "evidence_to_check": "PDCA全体推移 / 外部根拠",
                "frequency": "中",
                "accuracy": "高",
                "tags": ["partner", "branding"],
            },
            {
                "scope": "企画書固有",
                "category": "オペレーション",
                "question": "ミール・アカデミー・コンサルの役割分担は何か？",
                "answer": (
                    "ミールは継続的な unit economics、アカデミーは人材育成と送客、コンサルは高単価収益の柱という役割です。"
                    " 3つを分けて見ることで、どこが成長と利益を作っているかを説明しやすくしています。"
                ),
                "metrics_to_check": "事業別売上 / 事業別粗利",
                "evidence_to_check": "各モデルシート / PL設計",
                "frequency": "高",
                "accuracy": "高",
                "tags": [],
            },
            {
                "scope": "企画書固有",
                "category": "オペレーション",
                "question": "アカデミーを C/B/A/S に分けている意味は何か？",
                "answer": (
                    "単価、人数、進級、認証の構造を分けないと、どの階層が売上と粗利を作っているか見えません。"
                    " 階層分解は、成長前提を '人数増' だけにしないための管理軸です。"
                ),
                "metrics_to_check": "各級売上 / 構成比 / 進級率",
                "evidence_to_check": "アカデミーモデル / 前提条件",
                "frequency": "中",
                "accuracy": "中",
                "tags": [],
            },
            {
                "scope": "企画書固有",
                "category": "オペレーション",
                "question": "コンサル数量ロジックはどこまで信頼できるか？",
                "answer": (
                    "現時点では P1〜P12 の売上構成比から件数を逆算しており、厳密な受注データ起点ではありません。"
                    " したがって、投資家向けには『方向は示せるが、精緻化が必要』と答えるのが正直です。"
                ),
                "metrics_to_check": "P1〜P12 売上構成比 / 推定件数",
                "evidence_to_check": "コンサルモデル / 次の改善施策",
                "frequency": "高",
                "accuracy": "低",
                "tags": [],
            },
            {
                "scope": "企画書固有",
                "category": "オペレーション",
                "question": "ミールの unit economics はどこまで実証済みか？",
                "answer": (
                    "ミールは価格・食数・継続率まで整理していますが、配送やサポートなど下流コストの実証はまだ薄いです。"
                    " そのため収益面はある程度説明できても、原価と再現性の詰めは追加で必要です。"
                ),
                "metrics_to_check": "price per item / meals per year / retention / 原価率",
                "evidence_to_check": "ミールモデル / PL設計",
                "frequency": "中",
                "accuracy": "中",
                "tags": [],
            },
            {
                "scope": "企画書固有",
                "category": "資金",
                "question": "なぜ開発費を償却で見ているのか？",
                "answer": (
                    f"開発投資は {development_years:.0f} 年の定額償却で見せ、cash と P/L を分けて説明するためです。"
                    " 収益計画と資金計画の論点を混ぜないためにも、この見せ方が妥当です。"
                ),
                "metrics_to_check": "開発投資額 / 当期償却額 / 未償却残高",
                "evidence_to_check": "費用計画の開発償却ブロック",
                "frequency": "中",
                "accuracy": "高",
                "tags": [],
            },
            {
                "scope": "企画書固有",
                "category": "リスク",
                "question": "この収益計画で、いま最も説明が弱い論点は何か？",
                "answer": (
                    f"現時点で一番弱いのは {next_actions[0]} です。"
                    " つまり、収益計画の方向性よりも、数量ロジックと PL への接続の説明が課題です。"
                ),
                "metrics_to_check": "次の改善施策 / worsened points",
                "evidence_to_check": next_actions[0],
                "frequency": "高",
                "accuracy": "高",
                "tags": ["sales"],
            },
            {
                "scope": "企画書固有",
                "category": "リスク",
                "question": "この計画を役員・投資家に出す前に、追加で確認すべきファクトは何か？",
                "answer": (
                    "追加で確認すべきなのは、営業容量、チャネル別獲得効率、コンサル数量の実測データです。"
                    " ここが埋まると、収益計画の蓋然性はかなり強く説明できます。"
                ),
                "metrics_to_check": "営業容量 / CAC proxy / コンサル件数",
                "evidence_to_check": "次の改善施策 / 外部根拠 / コンサルモデル",
                "frequency": "高",
                "accuracy": "高",
                "tags": ["sales", "partner"],
            },
        ]
    else:
        qa_specs = [
            {
                "scope": "一般",
                "category": "収益",
                "question": "売上は誰に何をどう売って伸ばす計画か？",
            "answer": (
                f"売上の主軸は {' / '.join(segments) or 'ミール・アカデミー・コンサル'} の3系統です。"
                f" FY1 {_fmt_int(assumptions['revenue_target'][0])} から FY5 {_fmt_int(assumptions['revenue_target'][-1])} まで伸ばす前提で、"
                "検証後のアクセルと営業効率の改善で成長を作ります。"
            ),
            "metrics_to_check": f"売上目標 FY1={_fmt_int(assumptions['revenue_target'][0])}, FY5={_fmt_int(assumptions['revenue_target'][-1])}",
            "evidence_to_check": f"PL設計の売上行 / PDF方針 / {source_types}",
            "frequency": "高",
            "accuracy": "高",
            "tags": ["sales"],
        },
        {
            "scope": "一般",
            "category": "収益",
            "question": "売上成長を最も左右するレバーは何か？",
            "answer": (
                f"現時点では「{diagnosis.get('hypothesis', {}).get('title', '-') }」が最有力です。"
                " 前半の検証で筋を確認し、後半で営業投資を加速する段階設計が成長ドライバーです。"
            ),
            "metrics_to_check": "pl Δ, model_sheets Δ, 総合スコア差分",
            "evidence_to_check": "PDCA全体推移 / diagnosis の verdict",
            "frequency": "高",
            "accuracy": "高",
            "tags": ["sales", "staged"],
        },
        {
            "scope": "一般",
            "category": "収益",
            "question": "売上予測は営業施策とどう結びついているか？",
            "answer": (
                "売上予測は単純な growth rate ではなく、営業効率やパートナー施策を overlay として比較しています。"
                " 現時点では営業効率の上積みが最も再現性の高い改善レバーです。"
            ),
            "metrics_to_check": "candidate-revenue-staged-sales の pl Δ / partner 候補との差分",
            "evidence_to_check": f"想定候補比較 / {external_label}",
            "frequency": "高",
            "accuracy": "中",
            "tags": ["sales", "partner"],
        },
        {
            "scope": "一般",
            "category": "コスト",
            "question": "費用計画はどのような前提で管理しているか？",
            "answer": (
                f"費用は人件費比率 {_fmt_pct(assumptions['personnel_ratio'][0])}、マーケ費比率 {_fmt_pct(assumptions['marketing_ratio'][0])}、"
                f"開発費比率 {_fmt_pct(assumptions['development_ratio'][0])} を起点に置いています。"
                " まず比率で全体像を押さえ、費用計画シートで明細へ落としています。"
            ),
            "metrics_to_check": "人件費比率 / マーケ費比率 / 開発費比率",
            "evidence_to_check": "費用計画 / （全Ver）前提条件",
            "frequency": "高",
            "accuracy": "高",
            "tags": [],
        },
        {
            "scope": "一般",
            "category": "コスト",
            "question": "固定費と変動費のバランスは妥当か？",
            "answer": (
                "現状は人件費が主な固定費で、マーケ費と一部の開発投資が調整レバーです。"
                " 固定費を先行させすぎず、検証後に投資を厚くする設計にしています。"
            ),
            "metrics_to_check": "人件費 / マーケ費 / OPEX合計",
            "evidence_to_check": "費用計画 / PDCA全体推移",
            "frequency": "中",
            "accuracy": "中",
            "tags": ["staged"],
        },
        {
            "scope": "一般",
            "category": "収益性",
            "question": "この計画で黒字化の筋は見えているか？",
            "answer": (
                "黒字化の筋は、売上を落とさずに PL 側の改善を積み上げられるかにかかっています。"
                f" 今回の採用候補では PL 再現度が {score_layers.get('pl', {}).get('delta', 0.0):+.4f} 改善しており、"
                "収益性の筋は前進しています。"
            ),
            "metrics_to_check": f"pl Δ={score_layers.get('pl', {}).get('delta', 0.0):+.4f}, 営業利益, 営業利益率",
            "evidence_to_check": "評価スコア / PL設計",
            "frequency": "高",
            "accuracy": "中",
            "tags": ["sales"],
        },
        {
            "scope": "一般",
            "category": "収益性",
            "question": "粗利とOPEXのどちらが収益性改善の主因か？",
            "answer": (
                "現状の改善は、粗利率よりも OPEX と売上のつながりを整える効果が大きいです。"
                " そのため、営業効率の上積みと費用配分の妥当性をセットで見ています。"
            ),
            "metrics_to_check": "粗利率 / OPEX合計 / pl Δ",
            "evidence_to_check": "PL設計 / 費用計画 / PDCA全体推移",
            "frequency": "中",
            "accuracy": "中",
            "tags": ["sales"],
        },
        {
            "scope": "一般",
            "category": "成長性",
            "question": "なぜ最初の数年を検証期間としているのか？",
            "answer": (
                "前半でユニットエコノミクスを確認し、後半で投資アクセルを踏む段階設計を採っています。"
                " これにより、検証前に固定費を先行させすぎるリスクを抑えます。"
            ),
            "metrics_to_check": "iterationごとの pl Δ / model_sheets Δ",
            "evidence_to_check": diagnosis.get("hypothesis", {}).get("detail", "-"),
            "frequency": "高",
            "accuracy": "高",
            "tags": ["staged"],
        },
        {
            "scope": "一般",
            "category": "成長性",
            "question": "成長投資のアクセル条件は何か？",
            "answer": (
                "前提は、初期の検証で unit economics の筋が見えたら営業投資を強めることです。"
                " つまり成長投資は一律ではなく、検証結果を踏まえた段階投入です。"
            ),
            "metrics_to_check": "PDCA全体推移 / 次の施策 / 売上成長差分",
            "evidence_to_check": "診断レポートの仮説とロジック",
            "frequency": "高",
            "accuracy": "中",
            "tags": ["staged"],
        },
        {
            "scope": "一般",
            "category": "リスク",
            "question": "現時点での主要リスクは何か？",
            "answer": (
                f"現時点の主要リスクは {next_actions[0]}。"
                " つまり、モデルの構造理解は進んだ一方で、consulting から PL への橋渡しはまだ改善余地があります。"
            ),
            "metrics_to_check": "悪化した点 / 次の改善施策",
            "evidence_to_check": current_summary.get("worsened_points", "-"),
            "frequency": "高",
            "accuracy": "高",
            "tags": ["sales"],
        },
        {
            "scope": "一般",
            "category": "市場",
            "question": "市場性や外部環境の裏づけはあるか？",
            "answer": (
                "外部根拠は PDF だけでなく、営業効率や市場浸透に関する external ソースも併用しています。"
                " ただし、完全な live ETL ではなく curated source cache ベースの運用です。"
            ),
            "metrics_to_check": "外部根拠タイプ / source cache の件数",
            "evidence_to_check": external_label,
            "frequency": "中",
            "accuracy": "中",
            "tags": ["partner", "branding"],
        },
        {
            "scope": "一般",
            "category": "オペレーション",
            "question": "事業モデルはどの単位で運用設計しているか？",
            "answer": (
                f"モデルは {' / '.join(segments) or 'ミール / アカデミー / コンサル'} の3系統で分けています。"
                " 各モデルシートの driver を PL 設計に接続し、運用単位ごとに改善できるようにしています。"
            ),
            "metrics_to_check": "モデルシート構成 / PL設計の参照式",
            "evidence_to_check": "ミールモデル / アカデミーモデル / コンサルモデル",
            "frequency": "高",
            "accuracy": "高",
            "tags": [],
        },
        {
            "scope": "一般",
            "category": "資金",
            "question": "開発投資は PL 上でどのように扱っているか？",
            "answer": (
                f"開発投資はキャッシュ支出と PL 計上を分け、定額法 {int(assumptions['development_amortization_years'][0])} 年で償却しています。"
                " そのため PL の開発費は当期投資額そのものではなく償却額です。"
            ),
            "metrics_to_check": "開発投資額 / 当期償却額 / 未償却残高",
            "evidence_to_check": "費用計画の開発償却ブロック",
            "frequency": "中",
            "accuracy": "高",
            "tags": [],
        },
        {
            "scope": "企画書固有",
            "category": "収益",
            "question": "なぜ営業効率を重ねた案が最有力なのか？",
            "answer": (
                "比較した候補の中で、営業効率を重ねた案が PL 改善と説明責任の両方で最も良い結果でした。"
                " partner や branding も候補ですが、現時点では sales overlay が最も筋の良い改善です。"
            ),
            "metrics_to_check": "candidate-revenue-staged-sales の total / pl Δ / explainability Δ",
            "evidence_to_check": "PDCA全体推移 / diagnosis verdict",
            "frequency": "高",
            "accuracy": "高",
            "tags": ["sales"],
        },
        {
            "scope": "企画書固有",
            "category": "収益",
            "question": "partner や branding を今回は採用しなかった理由は何か？",
            "answer": (
                "partner と branding は比較対象として有効でしたが、今回のベスト候補ほど PL 改善を押し上げませんでした。"
                " 補助レバーとしては残しつつ、一次採用は見送っています。"
            ),
            "metrics_to_check": "partner/branding 候補との差分",
            "evidence_to_check": "PDCA全体推移の改善点・悪化点",
            "frequency": "高",
            "accuracy": "高",
            "tags": ["partner", "branding"],
        },
        {
            "scope": "企画書固有",
            "category": "収益性",
            "question": "今回の改善で何がどこまで良くなったか？",
            "answer": (
                f"PL 再現度は {score_layers.get('pl', {}).get('delta', 0.0):+.4f}、"
                f"説明責任は {score_layers.get('explainability', {}).get('delta', 0.0):+.4f} 改善しました。"
                " 総合スコアだけでなく、どのレイヤーに効いたかを見える化しています。"
            ),
            "metrics_to_check": "structure Δ / model_sheets Δ / pl Δ / explainability Δ",
            "evidence_to_check": current_summary.get("improved_points", "-"),
            "frequency": "高",
            "accuracy": "高",
            "tags": ["sales"],
        },
        {
            "scope": "企画書固有",
            "category": "リスク",
            "question": "まだ弱い部分はどこか？",
            "answer": (
                f"現時点でまだ弱いのは {next_actions[0]}"
                " workbook の見やすさは改善しましたが、consulting driver と PL の接続は引き続き主要論点です。"
            ),
            "metrics_to_check": "次の改善施策 / worsened_points",
            "evidence_to_check": current_summary.get("worsened_points", "-"),
            "frequency": "高",
            "accuracy": "高",
            "tags": ["sales"],
        },
        {
            "scope": "企画書固有",
            "category": "市場",
            "question": "営業以外に市場浸透で比較した施策は何か？",
            "answer": (
                "パートナー戦略やブランド波及も比較候補として検証しました。"
                " 現時点では営業効率を重ねた案が最も有力ですが、市場浸透の補助レバーとして継続監視します。"
            ),
            "metrics_to_check": "candidate-revenue-staged-partner / staged-branding の差分",
            "evidence_to_check": external_label,
            "frequency": "中",
            "accuracy": "中",
            "tags": ["partner", "branding"],
        },
        {
            "scope": "企画書固有",
            "category": "オペレーション",
            "question": "ミール・アカデミー・コンサルの役割分担は何か？",
            "answer": (
                "ミールはユニットエコノミクス、アカデミーは育成と認証、コンサルは高単価売上の柱として置いています。"
                " 3事業を別シートで管理し、PL設計に統合しています。"
            ),
            "metrics_to_check": "各モデルシートの売上 / PL設計の事業別行",
            "evidence_to_check": "ミールモデル / アカデミーモデル / コンサルモデル",
            "frequency": "高",
            "accuracy": "高",
            "tags": [],
        },
        {
            "scope": "企画書固有",
            "category": "オペレーション",
            "question": "アカデミーの階層分解はどう置いているか？",
            "answer": (
                "アカデミーは C/B/A/S 級に分け、構成比・進級率・認証率で系列化しています。"
                " これにより単価と人数の内訳を透明に追えるようにしています。"
            ),
            "metrics_to_check": "C級新規構成比 / C→B進級率 / 各級売上",
            "evidence_to_check": "（全Ver）前提条件のアカデミー前提",
            "frequency": "中",
            "accuracy": "中",
            "tags": [],
        },
        {
            "scope": "企画書固有",
            "category": "オペレーション",
            "question": "コンサルモデルの数量はどう作っているか？",
            "answer": (
                "コンサルは P1〜P12 の SKU ごとに売上構成比を置き、総売上から数量を逆算しています。"
                " まだ proxy を含むため、今後は実データ寄りに精緻化する余地があります。"
            ),
            "metrics_to_check": "P1〜P12売上構成比 / 推定件数",
            "evidence_to_check": "コンサルモデル / （全Ver）前提条件",
            "frequency": "低",
            "accuracy": "中",
            "tags": [],
        },
        {
            "scope": "企画書固有",
            "category": "資金",
            "question": "なぜ開発費を償却で見せているのか？",
            "answer": (
                "この計画では開発投資を一時費用ではなく将来収益に効く投資として扱うため、PL上は償却費で見せています。"
                " そのほうが cash と PL の説明責任を分けて整理できます。"
            ),
            "metrics_to_check": "開発投資額 / 当期償却額 / 未償却残高",
            "evidence_to_check": "費用計画の開発償却ブロック / 前提条件",
            "frequency": "中",
            "accuracy": "高",
            "tags": [],
        },
        {
            "scope": "企画書固有",
            "category": "成長性",
            "question": "3年間検証してから投資するのは保守的すぎないか？",
            "answer": (
                "保守的に見えますが、今回の比較では段階設計全体のほうが単純加速より再現度が高い結果でした。"
                " 先に検証条件を明確にしておくことで、投資判断の説明力も上がります。"
            ),
            "metrics_to_check": "staged_acceleration と acceleration_period の差分",
            "evidence_to_check": "PDCA全体推移 / diagnosis verdict",
            "frequency": "中",
            "accuracy": "中",
            "tags": ["staged"],
        },
            {
                "scope": "企画書固有",
                "category": "資金",
                "question": "資金の使い道は何を優先する計画か？",
            "answer": (
                "優先順位は、検証を支える運営投資、その後の営業投資、そして必要な開発投資です。"
                " 先に sales を厚くしすぎず、勝ち筋確認後に投下する構えです。"
            ),
            "metrics_to_check": "人件費 / マーケ費 / 開発費（償却）",
            "evidence_to_check": "費用計画 / 仮説の詳細",
                "frequency": "中",
                "accuracy": "中",
                "tags": ["staged", "sales"],
            },
        ]

    items: list[dict[str, Any]] = []
    for spec in qa_specs:
        first_added, last_updated = _iteration_range_for_tags(iteration_summaries, spec["tags"])
        status = _qa_status(first_added, last_updated, current_iteration)
        adoption = _qa_adoption(spec["tags"], diagnosis)
        support_buckets = _build_qa_support_buckets(
            spec=spec,
            pdf_facts=pdf_facts,
            external_label=external_label,
            current_summary=current_summary,
            next_actions=next_actions,
            assumptions=assumptions,
            gross_margin_start=gross_margin_start,
            gross_margin_end=gross_margin_end,
            meal_retention_start=meal_retention_start,
            consult_retention_start=consult_retention_start,
            personnel_ratio_start=personnel_ratio_start,
            marketing_ratio_start=marketing_ratio_start,
            development_ratio_start=development_ratio_start,
            development_years=development_years,
            pl_delta=pl_delta,
            total_delta=total_delta,
        )
        support_categories = _select_qa_support_categories(spec)
        filtered_buckets = [bucket for bucket in support_buckets if bucket["category"] in support_categories]
        answer = spec["answer"]
        items.append(
            {
                "scope": spec["scope"],
                "category": spec["category"],
                "question": spec["question"],
                "answer": answer,
                "metrics_to_check": spec["metrics_to_check"],
                "evidence_to_check": spec["evidence_to_check"],
                "frequency": spec["frequency"],
                "accuracy": spec["accuracy"],
                "first_added_iteration": first_added,
                "last_updated_iteration": last_updated,
                "status": status,
                "adoption": adoption,
                "support_categories": " / ".join(bucket["category"] for bucket in filtered_buckets),
                "support_details": "\n".join(f"{bucket['category']}: {bucket['detail']}" for bucket in filtered_buckets),
                "support_sources": "\n".join(f"{bucket['category']}: {bucket['source']}" for bucket in filtered_buckets),
            }
        )
    return items


def _build_qa_support_buckets(
    *,
    spec: dict[str, Any],
    pdf_facts: list[str],
    external_label: str,
    current_summary: dict[str, Any],
    next_actions: list[str],
    assumptions: dict[str, list[float]],
    gross_margin_start: float,
    gross_margin_end: float,
    meal_retention_start: float,
    consult_retention_start: float,
    personnel_ratio_start: float,
    marketing_ratio_start: float,
    development_ratio_start: float,
    development_years: float,
    pl_delta: float,
    total_delta: float,
) -> list[dict[str, str]]:
    category = spec["category"]
    question = spec["question"]
    tags = spec["tags"]
    external_refs = _external_refs_for_tags(tags)
    external_detail = (
        " / ".join(f"{ref['publisher']}: {ref['quote']}" for ref in external_refs[:2])
        if external_refs
        else "該当する外部ベンチマークは未登録です。"
    )
    external_source = (
        "\n".join(f"{ref['publisher']} | {ref['title']} | {ref['url']}" for ref in external_refs[:2])
        if external_refs
        else "source cache 未登録"
    )

    fact_detail = "内部実績・PoC・営業実測など、計画の前提を直接裏づける一次ファクトは現行 workbook では未格納です。"
    fact_source = "内部実績 / PoC結果 / CRM実測 / 運用ログ（未連携）"
    if "先行指標" in question:
        fact_detail = "現状のモデルで直接追える先行指標は、受講人数・食数・コンサル構成比までです。パイプラインや受注残は未連携です。"
        fact_source = "アカデミーモデル / ミールモデル / コンサルモデル"
    elif "営業体制と採用計画" in question:
        fact_detail = "営業容量と採用計画を直接裏づける headcount・生産性・立ち上がり実測は未格納です。"
        fact_source = "営業計画原票 / 採用計画 / CRM生産性データ（未連携）"
    elif "下振れした場合" in question:
        fact_detail = "下振れ時の壊れ方を裏づける実際の感応度ログは未取得で、現状はモデル比較から推定しています。"
        fact_source = "シナリオ比較 / 実績感応度ログ（未整備）"
    elif "役割分担" in question:
        fact_detail = "ミール・アカデミー・コンサルの3本柱に分ける事業構造自体は、現行計画で明示的に整理済みです。"
        fact_source = "事業計画PDF / モデル構造整理"

    if category == "収益":
        data_detail = (
            f"売上目標は FY1 {_fmt_int(assumptions['revenue_target'][0])} -> FY5 {_fmt_int(assumptions['revenue_target'][-1])}、"
            f"継続率は meal {_fmt_pct(meal_retention_start)} / consult {_fmt_pct(consult_retention_start)} です。"
        )
    elif category == "収益性":
        data_detail = (
            f"粗利率は FY1 {_fmt_pct(gross_margin_start)} -> FY5 {_fmt_pct(gross_margin_end)}、"
            f"PL差分は {pl_delta:+.4f} です。"
        )
    elif category == "コスト":
        data_detail = (
            f"費用比率は人件費 {_fmt_pct(personnel_ratio_start)} / マーケ費 {_fmt_pct(marketing_ratio_start)} / "
            f"開発費 {_fmt_pct(development_ratio_start)} です。"
        )
    elif category == "資金":
        data_detail = (
            f"開発費は {development_years:.0f} 年償却、総合差分は {total_delta:+.4f}、PL差分は {pl_delta:+.4f} です。"
        )
    elif category == "オペレーション":
        data_detail = (
            f"アカデミー受講人数 FY1 {_fmt_int(assumptions['academy_students'][0])}、"
            f"meal 食数/年 FY1 {_fmt_int(assumptions['meal_meals_per_year'][0])} を起点にしています。"
        )
    else:
        data_detail = (
            f"売上目標 FY1 {_fmt_int(assumptions['revenue_target'][0])}、PL差分 {pl_delta:+.4f}、"
            f"次アクション {next_actions[0]}"
        )

    plan_detail = _plan_assumption_for_question(spec, current_summary, next_actions, pl_delta, total_delta)
    plan_source = spec["evidence_to_check"]
    other_detail = external_detail if external_refs else (external_label if external_label != "pdf" else current_summary.get("improved_points", "-"))
    if not other_detail:
        other_detail = next_actions[0]
    other_source = external_source if external_refs else f"外部比較 / PDCA全体推移 / 次の改善施策 / tags={', '.join(tags) or '-'}"

    return [
        {
            "category": "ファクト",
            "detail": fact_detail,
            "source": fact_source,
        },
        {
            "category": "データ",
            "detail": data_detail,
            "source": spec["metrics_to_check"],
        },
        {
            "category": "事業計画",
            "detail": plan_detail,
            "source": plan_source,
        },
        {
            "category": "その他",
            "detail": other_detail,
            "source": other_source,
        },
    ]


def _select_qa_support_categories(spec: dict[str, Any]) -> list[str]:
    question = spec["question"]
    category = spec["category"]
    tags = spec["tags"]

    selected: list[str] = ["事業計画"]

    if any(
        token in question
        for token in [
            "価格",
            "数量",
            "継続率",
            "粗利率",
            "黒字化",
            "費用",
            "下振れ",
            "追加資金",
            "投資回収",
            "受講人数",
            "数量ロジック",
            "KPI",
        ]
    ) or category in {"収益性", "コスト", "資金"}:
        selected.append("データ")

    if any(
        token in question
        for token in [
            "検証期間",
            "役割分担",
            "償却",
            "主軸",
            "市場成長ではなく",
            "営業投資",
            "顧客獲得仮説",
            "ミール",
            "アカデミー",
            "コンサル",
            "先行指標",
            "営業体制",
        ]
    ):
        selected.append("ファクト")

    if any(tag in {"partner", "branding"} for tag in tags) or any(
        token in question for token in ["市場成長ではなく", "主軸にしなかった", "sales efficiency", "外部"]
    ):
        selected.append("その他")

    ordered = []
    for label in ["ファクト", "データ", "事業計画", "その他"]:
        if label in selected and label not in ordered:
            ordered.append(label)
    return ordered


def _external_refs_for_tags(tags: list[str]) -> list[dict[str, str]]:
    tag_to_source_type = {
        "sales": "sales_efficiency_analysis",
        "partner": "partner_strategy_analysis",
        "staged": "staged_acceleration_analysis",
        "branding": "branding_lift_analysis",
    }
    refs: list[dict[str, str]] = []
    seen_urls: set[str] = set()
    for tag in tags:
        source_type = tag_to_source_type.get(tag)
        if not source_type:
            continue
        for ref in analysis_source_refs(source_type):
            url = ref.get("url", "")
            if url and url in seen_urls:
                continue
            if url:
                seen_urls.add(url)
            refs.append(ref)
    return refs


def _plan_assumption_for_question(
    spec: dict[str, Any],
    current_summary: dict[str, Any],
    next_actions: list[str],
    pl_delta: float,
    total_delta: float,
) -> str:
    question = spec["question"]
    tags = spec["tags"]
    if "価格・数量・継続率" in question:
        return "計画上は、価格改定より数量増と継続率維持で売上を伸ばす前提を採用しています。"
    if "市場成長ではなく" in question:
        return "計画上は、市場追い風ではなく営業効率の高い顧客獲得チャネルへ投資を寄せる前提です。"
    if "先行指標" in question:
        return "計画上は、受講人数・食数・コンサル構成比を先行管理指標として扱っています。"
    if "営業体制と採用計画" in question:
        return "計画上は sales overlay が効く前提ですが、営業容量の詳細前提はまだ薄い状態です。"
    if "営業投資を強める前提条件" in question or "検証期間" in question or "解放条件" in question:
        return "計画上は、検証期間で筋を確認した後に sales 投資を加速する段階設計を採用しています。"
    if "sales efficiency" in question:
        return f"計画上は、sales efficiency を重ねた候補が総合 {total_delta:+.4f} / PL {pl_delta:+.4f} と最良だったため主軸採用です。"
    if "partner" in question or "branding" in question:
        return "計画上は、partner / branding は補助レバーとして位置づけ、主軸採用はしていません。"
    if "償却" in question:
        return "計画上は、開発投資を cash と P/L で分けて説明するため、償却表示を採用しています。"
    if "役割分担" in question:
        return "計画上は、ミール・アカデミー・コンサルを別の収益エンジンとして管理する設計です。"
    if "最も説明が弱い" in question or "追加で確認すべきファクト" in question:
        return f"計画上の未解決論点は {next_actions[0]} で、ここを次の改善対象としています。"
    if "感応度" in question:
        return "計画上は、sales efficiency・partner・継続率・consulting 数量を主要感応度レバーとして扱います。"
    if tags:
        return f"計画上は、{', '.join(tags)} に関する前提を採用し、比較候補の中で最良のものを選んでいます。"
    return current_summary.get("summary", "計画上の採用前提は current run の summary に従います。")


def _current_iteration(iteration_summaries: list[dict[str, Any]], candidate_id: str) -> int:
    for item in iteration_summaries:
        if item.get("candidate_id") == candidate_id:
            return int(item.get("iteration", 0))
    return max((int(item.get("iteration", 0)) for item in iteration_summaries), default=0)


def _iteration_range_for_tags(iteration_summaries: list[dict[str, Any]], tags: list[str]) -> tuple[int, int]:
    if not tags:
        return (0, 0)
    matched_iterations: list[int] = []
    for item in iteration_summaries:
        changed_levers = str(item.get("changed_levers", ""))
        on_levers = changed_levers
        if " / OFF:" in changed_levers:
            on_levers = changed_levers.split(" / OFF:", 1)[0]
        haystack = " ".join(
            [
                str(item.get("candidate_id", "")).lower(),
                str(item.get("hypothesis", "")).lower(),
                on_levers.lower(),
            ]
        )
        if any(tag.lower() in haystack for tag in tags):
            matched_iterations.append(int(item.get("iteration", 0)))
    if not matched_iterations:
        return (0, 0)
    return (min(matched_iterations), max(matched_iterations))


def _qa_status(first_added: int, last_updated: int, current_iteration: int) -> str:
    if first_added == 0 and last_updated == 0:
        return "継続"
    if last_updated == current_iteration:
        return "新規" if first_added == current_iteration else "更新"
    return "継続"


def _qa_adoption(tags: list[str], diagnosis: dict[str, Any]) -> str:
    if not tags:
        return "今回採用"
    toggles_on = set(diagnosis.get("logic", {}).get("toggles_on", []) or [])
    if any(tag in toggles_on for tag in tags):
        return "今回採用"
    return "比較のみ"


def _fmt_int(value: float) -> str:
    return f"{int(round(value)):,}"


def _fmt_pct(value: float) -> str:
    return f"{value * 100:.1f}%"


def _write_score_table(sheet, row_index: int, score: dict[str, Any], baseline_total: float) -> int:
    headers = ["項目", "スコア", "差分", "補足"]
    for column_index, header in enumerate(headers, start=1):
        sheet.cell(row=row_index, column=column_index, value=header)
        sheet.cell(row=row_index, column=column_index).font = BOLD_FONT
    row_index += 1
    total = score.get("total", 0.0)
    total_delta = score.get("delta_vs_baseline", round(total - baseline_total, 4))
    score_rows = [("total", total, total_delta, "総合評価")]
    for layer_name, layer_payload in score.get("layers", {}).items():
        score_rows.append((layer_name, layer_payload.get("value", 0.0), layer_payload.get("delta", 0.0), "個別評価"))
    for label, value, delta, note in score_rows:
        sheet.cell(row=row_index, column=1, value=label)
        sheet.cell(row=row_index, column=2, value=value)
        sheet.cell(row=row_index, column=3, value=delta)
        sheet.cell(row=row_index, column=4, value=note)
        sheet.cell(row=row_index, column=2).number_format = "0.0000"
        sheet.cell(row=row_index, column=3).number_format = "+0.0000;-0.0000;0.0000"
        row_index += 1
    return row_index


def _write_assumption_table(sheet, row_index: int, assumptions: list[dict[str, Any]]) -> int:
    headers = ["source_type", "review_status", "evidence_count", "evidence_ids"]
    for column_index, header in enumerate(headers, start=1):
        sheet.cell(row=row_index, column=column_index, value=header)
        sheet.cell(row=row_index, column=column_index).font = BOLD_FONT
    row_index += 1
    if not assumptions:
        sheet.cell(row=row_index, column=1, value="前提データなし")
        return row_index + 1
    for assumption in assumptions:
        evidence_refs = assumption.get("evidence_refs", [])
        evidence_ids = ", ".join(
            ref.get("source_id") or ref.get("title") or ""
            for ref in evidence_refs
            if ref.get("source_id") or ref.get("title")
        )
        sheet.cell(row=row_index, column=1, value=assumption.get("source_type", ""))
        sheet.cell(row=row_index, column=2, value=assumption.get("review_status", ""))
        sheet.cell(row=row_index, column=3, value=len(evidence_refs))
        sheet.cell(row=row_index, column=4, value=evidence_ids)
        row_index += 1
    return row_index


def _write_bullet_list(sheet, row_index: int, lines: list[str]) -> int:
    for line in lines:
        sheet.cell(row=row_index, column=1, value="・")
        sheet.cell(row=row_index, column=4, value=line)
        row_index += 1
    return row_index


def _write_iteration_trend_table(sheet, row_index: int, iteration_summaries: list[dict[str, Any]]) -> int:
    headers = [
        "回",
        "候補",
        "仮説",
        "変更レバー",
        "良くなった点",
        "悪化した点",
        "structure Δ",
        "model_sheets Δ",
        "pl Δ",
        "explainability Δ",
        "成果物観点",
        "判定",
        "次の施策",
    ]
    for column_index, header in enumerate(headers, start=1):
        sheet.cell(row=row_index, column=column_index, value=header)
        sheet.cell(row=row_index, column=column_index).font = BOLD_FONT
    row_index += 1
    for item in iteration_summaries:
        sheet.cell(row=row_index, column=1, value=item.get("iteration"))
        sheet.cell(row=row_index, column=2, value=item.get("candidate_id", ""))
        sheet.cell(row=row_index, column=3, value=item.get("hypothesis", ""))
        sheet.cell(row=row_index, column=4, value=item.get("changed_levers", ""))
        sheet.cell(row=row_index, column=5, value=item.get("improved_points", ""))
        sheet.cell(row=row_index, column=6, value=item.get("worsened_points", ""))
        sheet.cell(row=row_index, column=7, value=item.get("structure_delta", 0.0))
        sheet.cell(row=row_index, column=8, value=item.get("model_sheets_delta", 0.0))
        sheet.cell(row=row_index, column=9, value=item.get("pl_delta", 0.0))
        sheet.cell(row=row_index, column=10, value=item.get("explainability_delta", 0.0))
        sheet.cell(row=row_index, column=11, value=item.get("artifact_impact", ""))
        sheet.cell(row=row_index, column=12, value=item.get("verdict", ""))
        sheet.cell(row=row_index, column=13, value=item.get("next_action", ""))
        for numeric_col in (7, 8, 9, 10):
            sheet.cell(row=row_index, column=numeric_col).number_format = "+0.0000;-0.0000;0.0000"
        row_index += 1
    return row_index


def _build_workbook_assumptions(candidate_payload: dict[str, Any]) -> dict[str, list[float]]:
    model_sheets = candidate_payload.get("model_sheets", {})
    academy = model_sheets.get("アカデミー", {})
    meal = model_sheets.get("ミール", {})
    consult = model_sheets.get("コンサル", {})
    pl_lines = candidate_payload.get("pl_lines", {})

    revenue_target = _normalize_series(pl_lines.get("売上", []), default=0.0)
    gross_profit_target = _normalize_series(pl_lines.get("粗利", []), default=0.0)
    opex_target = _normalize_series(pl_lines.get("事業運営費（OPEX）", []), default=0.0)

    academy_price = _normalize_series(academy.get("academy_price", []), default=0.0)
    academy_students = _normalize_series(academy.get("academy_students", []), default=0.0)
    academy_certified = _normalize_series(academy.get("academy_certified", []), default=0.0)
    academy_revenue = _normalize_series(
        academy.get("academy_revenue", []),
        default=0.0,
        fallback=_series_product(academy_price, academy_students),
    )

    meal_price = _normalize_series(meal.get("price_per_item", []), default=0.0)
    meal_items = _normalize_series(meal.get("items_per_meal", []), default=1.0)
    meal_meals = _normalize_series(meal.get("meals_per_year", []), default=0.0)
    meal_retention = _normalize_series(meal.get("retention_rate", []), default=0.0)

    consult_price = _normalize_series(consult.get("sku_unit_price", []), default=0.0)
    consult_retention = _normalize_series(consult.get("sku_retention", []), default=0.0)
    consult_standard_hours = _normalize_series(consult.get("sku_standard_hours", []), default=0.0)
    assumptions = {
        "revenue_target": revenue_target,
        "gross_profit_target": gross_profit_target,
        "opex_target": opex_target,
        "academy_public_price": academy_price,
        "academy_students": academy_students,
        "academy_certified": academy_certified,
        "academy_revenue": academy_revenue,
        "meal_price_per_item": meal_price,
        "meal_items_per_meal": meal_items,
        "meal_meals_per_year": meal_meals,
        "meal_retention_rate": meal_retention,
        "meal_share_non_academy": [0.15] * len(YEAR_HEADERS),
        "consult_unit_price": consult_price,
        "consult_retention": consult_retention,
        "consult_standard_hours": consult_standard_hours,
        "personnel_ratio": [0.45] * len(YEAR_HEADERS),
        "marketing_ratio": [0.25] * len(YEAR_HEADERS),
        "development_ratio": [0.20] * len(YEAR_HEADERS),
        "other_ratio": [0.10] * len(YEAR_HEADERS),
        "development_amortization_years": [5] * len(YEAR_HEADERS),
    }
    for level in ACADEMY_LEVELS:
        code = level["code"]
        assumptions[f"academy_{code}_share"] = list(level["share_default"])
        assumptions[f"academy_{code}_price_multiplier"] = [level["price_multiplier"]] * len(YEAR_HEADERS)
        assumptions[f"academy_{code}_completion"] = [level["completion"]] * len(YEAR_HEADERS)
        assumptions[f"academy_{code}_certification"] = [level["certification"]] * len(YEAR_HEADERS)
        if code != "s":
            assumptions[f"academy_{code}_to_{ACADEMY_LEVELS[ACADEMY_LEVELS.index(level)+1]['code']}"] = [level["progression_to_next"]] * len(YEAR_HEADERS)
    assumptions["blended_hourly_rate"] = [3150.0] * len(YEAR_HEADERS)
    for sku in CONSULT_SKUS:
        assumptions[f"consult_{sku['sku'].lower()}_share"] = list(sku["share_default"])
    return assumptions


def _normalize_series(
    values: list[float] | tuple[float, ...],
    *,
    default: float,
    fallback: list[float] | None = None,
    length: int = 5,
) -> list[float]:
    if values:
        series = [float(value) for value in values]
    elif fallback:
        series = [float(value) for value in fallback]
    else:
        series = []

    if not series:
        return [default] * length
    if len(series) >= length:
        return series[:length]
    series = series + [series[-1]] * (length - len(series))
    return series


def _series_product(left: list[float], right: list[float]) -> list[float]:
    left_series = _normalize_series(left, default=0.0)
    right_series = _normalize_series(right, default=0.0)
    return [left_value * right_value for left_value, right_value in zip(left_series, right_series)]


def _write_header_row(sheet) -> None:
    sheet.cell(row=1, column=1, value="項目")
    sheet["A1"].font = BOLD_FONT
    for column_index, header in enumerate(YEAR_HEADERS, start=2):
        sheet.cell(row=1, column=column_index, value=header)
        sheet.cell(row=1, column=column_index).font = BOLD_FONT


def _apply_row_style(sheet, row_index: int, start_col: int, end_col: int, *, fill: PatternFill, font: Font | None = None) -> None:
    for column_index in range(start_col, end_col + 1):
        cell = sheet.cell(row=row_index, column=column_index)
        cell.fill = fill
        if font is not None:
            cell.font = font


def _style_input_rows(sheet, row_indexes: list[int], *, start_col: int = 2, end_col: int = 6) -> None:
    for row_index in row_indexes:
        _apply_row_style(sheet, row_index, start_col, end_col, fill=INPUT_FILL)


def _style_formula_rows(sheet, row_indexes: list[int], *, start_col: int = 2, end_col: int = 6) -> None:
    for row_index in row_indexes:
        _apply_row_style(sheet, row_index, start_col, end_col, fill=FORMULA_FILL)


def _style_subtotal_rows(sheet, row_indexes: list[int], *, start_col: int = 1, end_col: int = 6) -> None:
    for row_index in row_indexes:
        _apply_row_style(sheet, row_index, start_col, end_col, fill=SUBTOTAL_FILL, font=BOLD_FONT)


def _style_total_rows(sheet, row_indexes: list[int], *, start_col: int = 1, end_col: int = 6) -> None:
    for row_index in row_indexes:
        _apply_row_style(sheet, row_index, start_col, end_col, fill=TOTAL_FILL, font=TOTAL_FONT)


def _set_number_format_rows(sheet, row_indexes: list[int], *, number_format: str, start_col: int = 2, end_col: int = 6) -> None:
    for row_index in row_indexes:
        for column_index in range(start_col, end_col + 1):
            sheet.cell(row=row_index, column=column_index).number_format = number_format


def _set_column_widths(sheet, widths: dict[str, float]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def _set_row_heights(sheet, heights: dict[int, float]) -> None:
    for row_index, height in heights.items():
        sheet.row_dimensions[row_index].height = height


def _align_range(sheet, *, start_row: int, end_row: int, start_col: int, end_col: int, alignment: Alignment) -> None:
    for row_index in range(start_row, end_row + 1):
        for column_index in range(start_col, end_col + 1):
            sheet.cell(row=row_index, column=column_index).alignment = alignment


def _apply_standard_layout(sheet, *, freeze_panes: str, label_width: float = 20, year_width: float = 12) -> None:
    sheet.freeze_panes = freeze_panes
    _set_column_widths(
        sheet,
        {
            "A": label_width,
            "B": year_width,
            "C": year_width,
            "D": year_width,
            "E": year_width,
            "F": year_width,
        },
    )
    _set_row_heights(sheet, {1: 22})
    _align_range(sheet, start_row=1, end_row=sheet.max_row, start_col=1, end_col=1, alignment=LEFT_ALIGN)
    if sheet.max_column >= 2:
        _align_range(sheet, start_row=1, end_row=sheet.max_row, start_col=2, end_col=min(6, sheet.max_column), alignment=RIGHT_ALIGN)


def _write_key_value_rows(sheet, rows: list[list[Any]]) -> None:
    for row_index, row in enumerate(rows, start=1):
        for col_index, value in enumerate(row, start=1):
            sheet.cell(row=row_index, column=col_index, value=value)
        sheet.cell(row=row_index, column=1).font = Font(bold=(row_index == 1 or len(row) > 1))


def _sheet_ref(sheet_name: str, cell_ref: str) -> str:
    escaped = sheet_name.replace("'", "''")
    return f"'{escaped}'!{cell_ref}"


def _year_columns() -> list[str]:
    return [excel_col(column_index) for column_index in range(2, 7)]


def excel_col(column_index: int) -> str:
    result = ""
    index = column_index
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        result = chr(65 + remainder) + result
    return result
