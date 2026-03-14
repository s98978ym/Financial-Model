"""Extract normalized benchmark data from a reference workbook."""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, Iterable, List, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class ReferenceWorkbook:
    segment_names: List[str] = field(default_factory=list)
    model_sheets: Dict[str, Dict[str, List[float]]] = field(default_factory=dict)
    pl_lines: Dict[str, List[float]] = field(default_factory=dict)


def extract_reference_workbook(path: Path) -> ReferenceWorkbook:
    workbook = load_workbook(path, data_only=True)

    pl_sheet = workbook["PL設計"]
    meal_sheet = _find_sheet(workbook.sheetnames, "ミールモデル")
    academy_sheet = _find_sheet(workbook.sheetnames, "アカデミーモデル")
    consulting_sheet = _find_sheet(workbook.sheetnames, "コンサルモデル")

    return ReferenceWorkbook(
        segment_names=_extract_segment_names(pl_sheet),
        model_sheets={
            "ミール": _extract_meal_sheet(workbook[meal_sheet]),
            "アカデミー": _extract_academy_sheet(workbook[academy_sheet]),
            "コンサル": _extract_consulting_sheet(workbook[consulting_sheet]),
        },
        pl_lines=_extract_pl_lines(pl_sheet),
    )


def _find_sheet(sheet_names: Iterable[str], prefix: str) -> str:
    for sheet_name in sheet_names:
        if sheet_name.startswith(prefix):
            return sheet_name
    raise KeyError(f"Sheet starting with '{prefix}' was not found")


def _extract_segment_names(sheet: Worksheet) -> List[str]:
    names: List[str] = []
    for row in sheet.iter_rows(values_only=True):
        for value in row:
            if value in {"アカデミー", "コンサル", "ミール"} and value not in names:
                names.append(value)
    return names


def _extract_meal_sheet(sheet: Worksheet) -> Dict[str, List[float]]:
    return {
        "price_per_item": _series_after_label(sheet, "価格/アイテム"),
        "items_per_meal": _series_after_label(sheet, "アイテム/食事"),
        "meals_per_year": _series_after_label(sheet, "食事数/年"),
        "retention_rate": _series_after_label(sheet, "継続率"),
    }


def _extract_academy_sheet(sheet: Worksheet) -> Dict[str, List[float]]:
    return {
        "academy_revenue": _series_after_label(sheet, "アカデミー"),
        "academy_price": _series_after_label(sheet, "単価"),
        "academy_students": _series_after_label(sheet, "受講人数"),
        "academy_certified": _series_after_label(sheet, "認証人数(期末)"),
    }


def _extract_consulting_sheet(sheet: Worksheet) -> Dict[str, List[float]]:
    unit_price = _first_numeric_in_column_pair(sheet, key_header="SKU", value_header="単価（円）")
    retention = _first_numeric_in_column_pair(sheet, key_header="SKU", value_header="継続率")
    standard_hours = _sum_numeric_row(sheet, "P1", start_col=5, end_col=7)
    return {
        "sku_unit_price": [unit_price] if unit_price is not None else [],
        "sku_retention": [retention] if retention is not None else [],
        "sku_standard_hours": [standard_hours] if standard_hours is not None else [],
    }


def _extract_pl_lines(sheet: Worksheet) -> Dict[str, List[float]]:
    labels = ["売上", "粗利", "事業運営費（OPEX）"]
    return {label: _series_after_label(sheet, label) for label in labels}


def _series_after_label(sheet: Worksheet, label: str, max_values: int = 5) -> List[float]:
    for row in sheet.iter_rows(values_only=True):
        values = list(row)
        if label not in values:
            continue
        label_index = values.index(label)
        series = [float(v) for v in values[label_index + 1 :] if isinstance(v, (int, float))]
        if series:
            return series[:max_values]
    return []


def _header_index(sheet: Worksheet, header: str) -> Optional[int]:
    for row in sheet.iter_rows(values_only=True):
        values = list(row)
        if header in values:
            return values.index(header) + 1
    return None


def _first_numeric_in_column_pair(sheet: Worksheet, key_header: str, value_header: str) -> Optional[float]:
    key_col = _header_index(sheet, key_header)
    value_col = _header_index(sheet, value_header)
    if key_col is None or value_col is None:
        return None

    for row_idx in range(1, sheet.max_row + 1):
        key = sheet.cell(row_idx, key_col).value
        value = sheet.cell(row_idx, value_col).value
        if isinstance(key, str) and key.startswith("P") and isinstance(value, (int, float)):
            return float(value)
    return None


def _sum_numeric_row(sheet: Worksheet, row_key: str, start_col: int, end_col: int) -> Optional[float]:
    for row_idx in range(1, sheet.max_row + 1):
        value = sheet.cell(row_idx, 2).value
        if value == row_key:
            numeric_values = [
                float(sheet.cell(row_idx, col_idx).value)
                for col_idx in range(start_col, end_col + 1)
                if isinstance(sheet.cell(row_idx, col_idx).value, (int, float))
            ]
            if numeric_values:
                return sum(numeric_values)
    return None
