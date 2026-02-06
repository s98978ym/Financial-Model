"""Template Catalog Generator.

Scans an Excel template and produces an InputCatalog -- a complete list of all
writable input cells, enriched with contextual metadata (labels, units,
periods, blocks).

Usage::

    from src.catalog.scanner import scan_template, export_catalog_json

    catalog = scan_template("templates/pl_template.xlsx")
    export_catalog_json(catalog, "output/input_catalog.json")
"""

from __future__ import annotations

import json
import logging
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

import openpyxl
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..config.models import CatalogItem, InputCatalog

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Japanese unit patterns commonly found in financial models
# ---------------------------------------------------------------------------
_UNIT_PATTERNS: List[Tuple[str, re.Pattern[str]]] = [
    ("円", re.compile(r"円")),
    ("万円", re.compile(r"万円")),
    ("百万円", re.compile(r"百万円")),
    ("億円", re.compile(r"億円")),
    ("千円", re.compile(r"千円")),
    ("%", re.compile(r"[%％]")),
    ("人", re.compile(r"人")),
    ("月", re.compile(r"月")),
    ("年", re.compile(r"年")),
    ("万", re.compile(r"万(?!円)")),
    ("億", re.compile(r"億(?!円)")),
    ("件", re.compile(r"件")),
    ("台", re.compile(r"台")),
    ("回", re.compile(r"回")),
    ("時間", re.compile(r"時間")),
    ("日", re.compile(r"日(?!付|本)")),
]

# Period detection patterns
_PERIOD_PATTERNS: List[Tuple[str, re.Pattern[str]]] = [
    # FY patterns: FY2024, FY24, FY2024E, etc.
    ("FY", re.compile(r"FY\s*(\d{2,4})\s*[EP]?", re.IGNORECASE)),
    # Quarter patterns: Q1, 1Q, Q1 FY24, etc.
    ("Q", re.compile(r"(\d)[QqＱ]|[QqＱ](\d)", re.IGNORECASE)),
    # Japanese fiscal year: 2024年度, 第N期, N期
    ("年度", re.compile(r"(\d{2,4})\s*年度")),
    ("期", re.compile(r"第?\s*(\d+)\s*期")),
    # Month patterns: 1月, Jan, January, 2024/01, etc.
    ("月", re.compile(r"(\d{1,2})\s*月")),
    ("Month", re.compile(
        r"(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*",
        re.IGNORECASE,
    )),
    # Yearly columns: 2024, 2025, etc.
    ("Year", re.compile(r"\b(20\d{2})\b")),
]

# KPI label patterns (used for block detection heuristic)
_SECTION_KEYWORDS: List[str] = [
    "売上",
    "原価",
    "粗利",
    "販管費",
    "営業利益",
    "経常利益",
    "EBITDA",
    "人件費",
    "減価償却",
    "設備投資",
    "運転資本",
    "キャッシュフロー",
    "CF",
    "BS",
    "PL",
    "前提",
    "KPI",
    "収益",
    "費用",
    "コスト",
    "投資",
]


# ===================================================================
# Public API
# ===================================================================

def scan_template(
    template_path: str,
    input_color: str = "FFFFF2CC",
    *,
    fuzzy_color: bool = True,
) -> InputCatalog:
    """Scan an Excel template and return a catalog of input cells.

    Parameters
    ----------
    template_path:
        Path to the ``.xlsx`` template file.
    input_color:
        Target fill colour in hex (without leading ``#``).  The default
        ``FFFFF2CC`` corresponds to the standard *light-yellow* input
        highlight.
    fuzzy_color:
        When *True*, colour matching strips common prefixes (``00``,
        ``FF``) and compares the core 6-hex-digit colour, making the
        match more tolerant to openpyxl serialisation quirks.

    Returns
    -------
    InputCatalog
        Catalog containing every matched cell with enriched metadata.
    """
    path = Path(template_path)
    if not path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    wb = openpyxl.load_workbook(str(path), data_only=False)
    catalog = InputCatalog()
    blocks_dict: Dict[str, List[CatalogItem]] = {}

    for ws in wb.worksheets:
        logger.info("Scanning sheet: %s", ws.title)
        _scan_sheet(ws, input_color, fuzzy_color, catalog, blocks_dict)

    catalog.blocks = blocks_dict

    writable_count = sum(1 for i in catalog.items if not i.has_formula)
    formula_count = sum(1 for i in catalog.items if i.has_formula)

    wb.close()
    logger.info(
        "Scan complete – %d items (%d writable, %d formula)",
        len(catalog.items),
        writable_count,
        formula_count,
    )
    return catalog


def export_catalog_json(catalog: InputCatalog, output_path: str) -> None:
    """Export *catalog* as a JSON file to *output_path*."""
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    with open(out, "w", encoding="utf-8") as fh:
        json.dump(catalog.model_dump(), fh, ensure_ascii=False, indent=2, default=str)
    logger.info("Catalog exported to %s", out)


# ===================================================================
# Internal helpers
# ===================================================================

def _scan_sheet(
    ws: Worksheet,
    input_color: str,
    fuzzy_color: bool,
    catalog: InputCatalog,
    blocks_dict: Dict[str, List[CatalogItem]],
) -> None:
    """Iterate every cell in *ws* and collect matching input cells."""
    merged_ranges = _get_merged_cell_values(ws)

    for row in ws.iter_rows():
        for cell in row:
            # Skip MergedCell placeholders – they have no independent fill
            if isinstance(cell, MergedCell):
                continue

            if not _match_color(cell.fill, input_color, fuzzy=fuzzy_color):
                continue

            item = _build_catalog_item(ws, cell, merged_ranges)
            catalog.items.append(item)

            # Group into blocks keyed by "SheetName::BlockName"
            block_key = f"{ws.title}::{item.block or '_unblocked'}"
            blocks_dict.setdefault(block_key, []).append(item)


def _build_catalog_item(
    ws: Worksheet,
    cell: Cell,
    merged_values: Dict[str, Any],
) -> CatalogItem:
    """Create a fully-enriched CatalogItem from *cell*."""
    row, col = cell.row, cell.column
    coord = cell.coordinate  # e.g. "C12"

    # Determine formula status
    has_formula = _cell_has_formula(cell)

    # Colour hex
    fill_color = _extract_fill_color(cell)

    # Nearby labels
    labels = _find_labels(ws, row, col, merged_values)

    # Units
    unit_candidates: List[str] = []
    for lbl in labels:
        unit_candidates.extend(_detect_units(lbl))
    # Also check the cell's own row header area for units
    for c in range(1, min(col, 4)):
        hdr_cell = ws.cell(row=row, column=c)
        if hdr_cell.value and isinstance(hdr_cell.value, str):
            unit_candidates.extend(_detect_units(hdr_cell.value))
    unit_candidates = list(dict.fromkeys(unit_candidates))  # dedupe, keep order

    # Period
    period = _detect_period(ws, row, col)

    # Block / section
    block = _detect_block(ws, row, col, merged_values)

    return CatalogItem(
        sheet=ws.title,
        cell=coord,
        current_value=cell.value,
        fill_color=fill_color,
        has_formula=has_formula,
        label_candidates=labels,
        unit_candidates=unit_candidates,
        year_or_period=period,
        block=block,
    )


# -------------------------------------------------------------------
# Colour matching
# -------------------------------------------------------------------

def _match_color(cell_fill, target_color: str, *, fuzzy: bool = True) -> bool:
    """Return *True* when the cell's fill colour matches *target_color*.

    Handles the common openpyxl quirks:
    * ``fgColor.rgb`` may carry a leading ``"00"`` or ``"FF"`` prefix.
    * Theme-based fills expose ``fgColor.theme`` instead of ``.rgb``.
    """
    if cell_fill is None or cell_fill.patternType is None:
        return False

    fg = cell_fill.fgColor
    if fg is None:
        return False

    rgb_value = fg.rgb
    if rgb_value is None or rgb_value == "00000000":
        # No explicit RGB – possibly theme-based; skip for now
        return False

    if isinstance(rgb_value, str):
        return _colors_equal(rgb_value, target_color, fuzzy=fuzzy)

    return False


def _colors_equal(a: str, b: str, *, fuzzy: bool = True) -> bool:
    """Compare two hex colour strings, optionally using fuzzy matching."""
    if not fuzzy:
        return a.upper() == b.upper()

    core_a = _strip_color_prefix(a).upper()
    core_b = _strip_color_prefix(b).upper()
    return core_a == core_b


def _strip_color_prefix(color: str) -> str:
    """Strip common leading prefixes from an openpyxl colour hex string.

    ``"00FFFFF2CC"`` -> ``"FFF2CC"``
    ``"FFFFF2CC"``   -> ``"FFF2CC"``
    ``"FFF2CC"``     -> ``"FFF2CC"``
    """
    c = color.upper().lstrip("#")
    # 8-char ARGB -> take last 6
    if len(c) == 8:
        return c[2:]
    # 9-char (rare openpyxl artefact with extra nibble)
    if len(c) == 9:
        return c[3:]
    # 10-char double prefix
    if len(c) == 10:
        return c[4:]
    return c


def _extract_fill_color(cell: Cell) -> str:
    """Return the hex fill colour string of *cell*, or ``""``."""
    try:
        fg = cell.fill.fgColor
        if fg and fg.rgb and isinstance(fg.rgb, str):
            return fg.rgb
    except Exception:  # noqa: BLE001
        pass
    return ""


# -------------------------------------------------------------------
# Formula detection
# -------------------------------------------------------------------

def _cell_has_formula(cell: Cell) -> bool:
    """Return True if the cell contains a formula."""
    if cell.data_type == "f":
        return True
    if isinstance(cell.value, str) and cell.value.startswith("="):
        return True
    return False


# -------------------------------------------------------------------
# Label detection
# -------------------------------------------------------------------

def _find_labels(
    ws: Worksheet,
    row: int,
    col: int,
    merged_values: Dict[str, Any],
) -> List[str]:
    """Gather label candidates from nearby cells.

    Strategy (in priority order):
    1. Same row, scan columns to the left until a text value is found.
    2. Row directly above the cell.
    3. Two rows above (sometimes headers are two rows up).
    4. Check merged cell registry for labels spanning the area.
    """
    candidates: List[str] = []
    seen: Set[str] = set()

    def _add(text: Any) -> None:
        if text is None:
            return
        s = str(text).strip()
        if s and s not in seen and not s.startswith("="):
            seen.add(s)
            candidates.append(s)

    # 1) Same row – scan leftward (up to 6 columns)
    for c in range(col - 1, max(0, col - 7), -1):
        val = _safe_cell_value(ws, row, c, merged_values)
        if val is not None:
            _add(val)
            break  # take closest label only

    # 2) One row above, same column
    if row > 1:
        _add(_safe_cell_value(ws, row - 1, col, merged_values))

    # 3) Two rows above, same column
    if row > 2:
        _add(_safe_cell_value(ws, row - 2, col, merged_values))

    # 4) Same column header (row 1 and row 2)
    for r in (1, 2):
        if r < row - 2:  # avoid duplicates from steps 2-3
            _add(_safe_cell_value(ws, r, col, merged_values))

    # 5) Same row, column A and B (common label positions)
    for c in (1, 2):
        if c < col - 6:  # avoid duplicates from step 1
            _add(_safe_cell_value(ws, row, c, merged_values))

    return candidates


def _safe_cell_value(
    ws: Worksheet,
    row: int,
    col: int,
    merged_values: Dict[str, Any],
) -> Any:
    """Get cell value, falling back to the merged-cell registry."""
    if row < 1 or col < 1:
        return None
    coord = f"{get_column_letter(col)}{row}"
    # Check merged values first (MergedCell instances have value=None)
    if coord in merged_values:
        return merged_values[coord]
    try:
        cell = ws.cell(row=row, column=col)
        if isinstance(cell, MergedCell):
            return merged_values.get(coord)
        val = cell.value
        # Only return string / numeric labels, not formulas
        if isinstance(val, str) and val.startswith("="):
            return None
        return val
    except Exception:  # noqa: BLE001
        return None


# -------------------------------------------------------------------
# Unit detection
# -------------------------------------------------------------------

def _detect_units(text: str) -> List[str]:
    """Return unit tokens found in *text*."""
    if not text or not isinstance(text, str):
        return []
    found: List[str] = []
    for unit_name, pattern in _UNIT_PATTERNS:
        if pattern.search(text):
            found.append(unit_name)
    return found


# -------------------------------------------------------------------
# Period / year detection
# -------------------------------------------------------------------

def _detect_period(ws: Worksheet, row: int, col: int) -> Optional[str]:
    """Try to detect a fiscal period label from header rows above *cell*.

    Checks the cell's own column in rows 1-5 and the cell directly above.
    """
    candidates: List[str] = []

    # Collect header text from the top rows and the row above
    search_rows = list(range(1, min(row, 6)))
    if row - 1 not in search_rows and row > 1:
        search_rows.append(row - 1)

    for r in search_rows:
        try:
            hdr = ws.cell(row=r, column=col).value
        except Exception:  # noqa: BLE001
            continue
        if hdr is None:
            continue
        hdr_str = str(hdr).strip()
        if not hdr_str:
            continue

        for period_name, pattern in _PERIOD_PATTERNS:
            m = pattern.search(hdr_str)
            if m:
                candidates.append(hdr_str)
                break  # one match per header cell is enough

    return candidates[0] if candidates else None


# -------------------------------------------------------------------
# Block / section detection
# -------------------------------------------------------------------

def _detect_block(
    ws: Worksheet,
    row: int,
    col: int,
    merged_values: Dict[str, Any],
) -> Optional[str]:
    """Detect the section / block name that a cell belongs to.

    Heuristics (top-down, first match wins):
    1. Walk upward in column A looking for a bold / merged cell whose
       text matches known section keywords.
    2. Walk upward in the same column looking for merged cells.
    3. Check column A of the current row for a section label.
    """
    # Strategy 1 & 2: walk upward from the cell's row
    for r in range(row - 1, 0, -1):
        # Check column A first (most common label column)
        for check_col in (1, 2):
            coord = f"{get_column_letter(check_col)}{r}"
            val = merged_values.get(coord)
            if val is None:
                try:
                    c = ws.cell(row=r, column=check_col)
                    if isinstance(c, MergedCell):
                        val = merged_values.get(coord)
                    else:
                        val = c.value
                        # Prefer bold cells as section headers
                        if val and c.font and c.font.bold:
                            return str(val).strip()
                except Exception:  # noqa: BLE001
                    continue

            if val is not None:
                text = str(val).strip()
                # Check if it looks like a section header
                if any(kw in text for kw in _SECTION_KEYWORDS):
                    return text
                # Large merged cells spanning several columns are often headers
                if coord in merged_values and len(text) > 1:
                    return text

        # Stop scanning if we've gone more than 30 rows up (performance)
        if row - r > 30:
            break

    return None


# -------------------------------------------------------------------
# Merged cell helpers
# -------------------------------------------------------------------

def _get_merged_cell_values(ws: Worksheet) -> Dict[str, Any]:
    """Build a mapping ``coord -> value`` for every cell inside a merged range.

    openpyxl stores the value only on the top-left cell of a merged range;
    all other cells in the range are ``MergedCell`` instances with
    ``value = None``.  This helper propagates the top-left value to every
    coordinate in the range so we can look it up easily.
    """
    mapping: Dict[str, Any] = {}
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        # The top-left cell holds the actual value
        top_left = ws.cell(row=min_row, column=min_col)
        val = top_left.value
        if val is None:
            continue
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                coord = f"{get_column_letter(c)}{r}"
                mapping[coord] = val
    return mapping
