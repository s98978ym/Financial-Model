"""Model Map / Formula Graph Analyzer.

Analyses the formula structure of an Excel template to produce:

1.  A human-readable model description.
2.  KPI definitions with formula translations.
3.  A dependency tree tracing KPIs back to input parameters.

Usage::

    from src.catalog.scanner import scan_template
    from src.modelmap.analyzer import analyze_model, generate_model_report_md

    catalog = scan_template("templates/pl_template.xlsx")
    report  = analyze_model("templates/pl_template.xlsx", catalog)
    md_text = generate_model_report_md(report)
"""

from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

import openpyxl
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..config.models import (
    AnalysisReport,
    CatalogItem,
    DependencyNode,
    FormulaInfo,
    InputCatalog,
    KPIDefinition,
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# KPI detection patterns (Japanese financial model terms)
# ---------------------------------------------------------------------------
_KPI_LABEL_PATTERNS: List[re.Pattern[str]] = [
    re.compile(r"売上(?:高|合計|総額)?"),
    re.compile(r"粗利(?:益)?(?:額)?"),
    re.compile(r"営業利益"),
    re.compile(r"経常利益"),
    re.compile(r"純利益"),
    re.compile(r"当期(?:純)?利益"),
    re.compile(r"EBITDA", re.IGNORECASE),
    re.compile(r"EBIT", re.IGNORECASE),
    re.compile(r"販管費(?:合計)?"),
    re.compile(r"原価(?:合計)?"),
    re.compile(r"売上総利益"),
    re.compile(r"営業(?:外)?(?:収益|費用)"),
    re.compile(r"減価償却"),
    re.compile(r"人件費(?:合計)?"),
    re.compile(r"(?:フリー)?キャッシュ\s*フロー"),
    re.compile(r"FCF", re.IGNORECASE),
    re.compile(r"ARR", re.IGNORECASE),
    re.compile(r"MRR", re.IGNORECASE),
    re.compile(r"ARPU", re.IGNORECASE),
    re.compile(r"LTV", re.IGNORECASE),
    re.compile(r"CAC", re.IGNORECASE),
    re.compile(r"チャーン(?:レート)?"),
    re.compile(r"Churn\s*Rate", re.IGNORECASE),
    re.compile(r"顧客(?:数|単価)"),
    re.compile(r"会員(?:数|単価)"),
    re.compile(r"従業員数"),
    re.compile(r"利益率"),
    re.compile(r"粗利率"),
    re.compile(r"営業利益率"),
]

# Sheet names that are likely to contain PL / KPI definitions
_KPI_SHEET_PATTERNS: List[re.Pattern[str]] = [
    re.compile(r"PL", re.IGNORECASE),
    re.compile(r"損益"),
    re.compile(r"P\s*&?\s*L", re.IGNORECASE),
    re.compile(r"収支"),
    re.compile(r"KPI", re.IGNORECASE),
    re.compile(r"サマリ"),
    re.compile(r"Summary", re.IGNORECASE),
    re.compile(r"設計"),
]

# ---------------------------------------------------------------------------
# Formula parsing regex
# ---------------------------------------------------------------------------

# Matches cell references like B12, $B$12, B$12, $B12
_CELL_REF_RE = re.compile(
    r"""
    (?:                           # optional sheet prefix
        (?:'([^']+)'|(\w+))       # 'Sheet Name' or SheetName
        !                         # sheet separator
    )?
    (\$?[A-Z]{1,3}\$?\d+)        # cell reference (with optional $ anchors)
    (?:                           # optional range end
        :
        (\$?[A-Z]{1,3}\$?\d+)    # end of range
    )?
    """,
    re.VERBOSE | re.IGNORECASE,
)

# Matches a bare cell reference (no sheet prefix) -- used for simple cases
_BARE_CELL_RE = re.compile(r"\$?([A-Z]{1,3})\$?(\d+)", re.IGNORECASE)

# Named ranges: just capture alphanumeric identifiers that are NOT
# Excel built-in function names
_EXCEL_FUNCTIONS: Set[str] = {
    "SUM", "SUMIF", "SUMIFS", "SUMPRODUCT",
    "AVERAGE", "AVERAGEIF", "AVERAGEIFS",
    "COUNT", "COUNTA", "COUNTIF", "COUNTIFS", "COUNTBLANK",
    "IF", "IFS", "IFERROR", "IFNA",
    "VLOOKUP", "HLOOKUP", "XLOOKUP", "INDEX", "MATCH",
    "MIN", "MAX", "LARGE", "SMALL",
    "ROUND", "ROUNDUP", "ROUNDDOWN", "INT", "MOD",
    "ABS", "SIGN", "SQRT", "POWER", "LOG", "LN", "EXP",
    "LEFT", "RIGHT", "MID", "LEN", "TRIM", "SUBSTITUTE",
    "CONCATENATE", "TEXTJOIN", "TEXT", "VALUE",
    "DATE", "YEAR", "MONTH", "DAY", "TODAY", "NOW",
    "AND", "OR", "NOT", "TRUE", "FALSE",
    "OFFSET", "INDIRECT", "ROW", "COLUMN", "ROWS", "COLUMNS",
    "TRANSPOSE", "SORT", "FILTER", "UNIQUE",
    "NPV", "IRR", "PMT", "PV", "FV", "RATE",
    "CHOOSE", "SWITCH",
    "EOMONTH", "EDATE", "DATEDIF", "NETWORKDAYS",
    "CEILING", "FLOOR",
    "RANK", "PERCENTILE",
    "LOOKUP",
}


# ===================================================================
# Public API
# ===================================================================

def analyze_model(
    template_path: str,
    catalog: InputCatalog,
) -> AnalysisReport:
    """Analyse the Excel template and return a structured report.

    Parameters
    ----------
    template_path:
        Path to the ``.xlsx`` template.
    catalog:
        The ``InputCatalog`` produced by :func:`src.catalog.scanner.scan_template`.

    Returns
    -------
    AnalysisReport
    """
    path = Path(template_path)
    if not path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    wb = openpyxl.load_workbook(str(path), data_only=False)

    # Step 1 -- Build a label map  (address -> best label)
    label_map = _build_label_map(catalog, wb)

    # Step 2 -- Build a set of known input addresses for quick lookup
    input_addresses: Set[str] = set()
    for item in catalog.items:
        input_addresses.add(f"'{item.sheet}'!{item.cell}")

    # Step 3 -- Collect all formula cells & parse them
    formulas: List[FormulaInfo] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                if not _cell_is_formula(cell):
                    continue
                raw = str(cell.value)
                refs = _parse_formula_refs(raw)
                # Qualify bare refs with current sheet name
                qualified_refs = _qualify_refs(refs, ws.title)
                human = _humanize_formula(raw, label_map)
                addr = f"'{ws.title}'!{cell.coordinate}"
                label = label_map.get(addr, "")
                formulas.append(FormulaInfo(
                    sheet=ws.title,
                    cell=cell.coordinate,
                    raw_formula=raw,
                    referenced_cells=qualified_refs,
                    human_readable=human,
                    label=label,
                ))

    # Step 4 -- Identify KPI cells
    kpi_cells: List[Tuple[str, str, str]] = []  # (sheet, coord, label)
    for ws in wb.worksheets:
        found = _identify_kpi_cells(ws)
        kpi_cells.extend(found)

    # Step 5 -- Build KPI definitions
    kpis: List[KPIDefinition] = []
    for sheet, coord, label in kpi_cells:
        addr = f"'{sheet}'!{coord}"
        # Find the matching FormulaInfo (if any)
        fi = next(
            (f for f in formulas if f.sheet == sheet and f.cell == coord),
            None,
        )
        raw = fi.raw_formula if fi else ""
        human = fi.human_readable if fi else ""
        deps = _collect_dependencies(
            addr, formulas, input_addresses, label_map,
        )
        kpis.append(KPIDefinition(
            name=label,
            sheet=sheet,
            cell=coord,
            raw_formula=raw,
            human_readable_formula=human,
            dependencies=deps,
        ))

    # Step 6 -- Build dependency trees for all KPIs
    dep_tree: Dict[str, DependencyNode] = {}
    formula_lookup = {
        f"'{f.sheet}'!{f.cell}": f for f in formulas
    }
    for kpi in kpis:
        addr = f"'{kpi.sheet}'!{kpi.cell}"
        node = _build_dependency_tree_node(
            addr, formula_lookup, input_addresses, label_map, visited=set(),
        )
        dep_tree[addr] = node

    # Step 7 -- Summary
    summary_parts = [
        f"Template: {path.name}",
        f"Sheets analysed: {len(wb.sheetnames)}",
        f"Total formulas found: {len(formulas)}",
        f"KPIs identified: {len(kpis)}",
        f"Input cells in catalog: {len(catalog.items)} "
        f"({sum(1 for i in catalog.items if not i.has_formula)} writable)",
    ]
    summary = "\n".join(summary_parts)

    wb.close()

    return AnalysisReport(
        template_path=str(path.resolve()),
        formulas=formulas,
        kpis=kpis,
        dependency_tree=dep_tree,
        label_map=label_map,
        summary=summary,
    )


def generate_model_report_md(report: AnalysisReport) -> str:
    """Generate a human-readable Markdown report from *report*."""
    lines: List[str] = []

    lines.append("# Financial Model Analysis Report")
    lines.append("")
    lines.append("## Overview")
    lines.append("")
    for line in report.summary.split("\n"):
        lines.append(f"- {line}")
    lines.append("")

    # KPI section
    if report.kpis:
        lines.append("## KPI Definitions")
        lines.append("")
        for kpi in report.kpis:
            lines.append(f"### {kpi.name}")
            lines.append("")
            lines.append(f"- **Location**: `'{kpi.sheet}'!{kpi.cell}`")
            if kpi.raw_formula:
                lines.append(f"- **Formula**: `{kpi.raw_formula}`")
            if kpi.human_readable_formula:
                lines.append(
                    f"- **Readable**: {kpi.human_readable_formula}"
                )
            if kpi.dependencies:
                lines.append("- **Depends on**:")
                for dep in kpi.dependencies:
                    lines.append(f"  - {dep}")
            lines.append("")

    # Dependency trees
    if report.dependency_tree:
        lines.append("## Dependency Trees")
        lines.append("")
        for addr, node in report.dependency_tree.items():
            label = node.label or addr
            lines.append(f"### {label}")
            lines.append("")
            lines.append("```")
            _render_tree(node, lines, indent=0)
            lines.append("```")
            lines.append("")

    # Formula summary
    if report.formulas:
        lines.append("## Formula Summary")
        lines.append("")
        lines.append("| Sheet | Cell | Label | Human-readable |")
        lines.append("|-------|------|-------|----------------|")
        for fi in report.formulas[:100]:  # cap at 100 for readability
            label = fi.label or ""
            human = fi.human_readable or fi.raw_formula
            # Escape pipe characters in table cells
            label = label.replace("|", "\\|")
            human = human.replace("|", "\\|")
            lines.append(f"| {fi.sheet} | {fi.cell} | {label} | {human} |")
        if len(report.formulas) > 100:
            lines.append(
                f"| ... | ... | ... | *({len(report.formulas) - 100} more)* |"
            )
        lines.append("")

    # Label map
    if report.label_map:
        lines.append("## Label Map (sample)")
        lines.append("")
        lines.append("| Address | Label |")
        lines.append("|---------|-------|")
        count = 0
        for addr, label in sorted(report.label_map.items()):
            label_escaped = label.replace("|", "\\|")
            lines.append(f"| `{addr}` | {label_escaped} |")
            count += 1
            if count >= 50:
                lines.append(
                    f"| ... | *({len(report.label_map) - 50} more)* |"
                )
                break
        lines.append("")

    return "\n".join(lines)


# ===================================================================
# Formula parsing
# ===================================================================

def _parse_formula_refs(formula: str) -> List[str]:
    """Extract cell references from an Excel formula string.

    Returns a list of reference strings.  Each entry is one of:
    - ``"B12"``             (bare cell)
    - ``"$B$12"``           (absolute cell)
    - ``"B12:B24"``         (range)
    - ``"'Sheet Name'!B12"`` (cross-sheet cell)
    - ``"SheetName!B12:C24"`` (cross-sheet range)

    Named ranges that are *not* Excel built-in functions are also
    captured when they look like identifiers followed by no ``(``.
    """
    if not formula or not isinstance(formula, str):
        return []

    # Strip leading "=" if present
    text = formula.lstrip("=")

    refs: List[str] = []
    seen: Set[str] = set()

    for m in _CELL_REF_RE.finditer(text):
        sheet_quoted = m.group(1)   # from 'Sheet Name'
        sheet_bare = m.group(2)     # from SheetName
        cell_start = m.group(3)     # e.g. B12 or $B$12
        cell_end = m.group(4)       # e.g. B24 (range end), may be None

        # Skip if the "sheet" part is actually a function name
        sheet = sheet_quoted or sheet_bare or ""
        if sheet.upper() in _EXCEL_FUNCTIONS:
            # This isn't really a sheet reference -- it's SUM(B12) etc.
            # The cell ref itself is still valid though.
            sheet = ""

        # Build the reference string
        if sheet:
            if sheet_quoted:
                ref = f"'{sheet}'!{cell_start}"
            else:
                ref = f"{sheet}!{cell_start}"
            if cell_end:
                ref += f":{cell_end}"
        else:
            ref = cell_start
            if cell_end:
                ref += f":{cell_end}"

        if ref not in seen:
            seen.add(ref)
            refs.append(ref)

    return refs


def _qualify_refs(
    refs: List[str],
    current_sheet: str,
) -> List[str]:
    """Prefix bare cell references with the current sheet name."""
    result: List[str] = []
    for ref in refs:
        if "!" in ref:
            result.append(ref)
        else:
            # Strip range part for qualification, keep it on the result
            result.append(f"'{current_sheet}'!{ref}")
    return result


# ===================================================================
# Human-readable formula translation
# ===================================================================

def _humanize_formula(formula: str, label_map: Dict[str, str]) -> str:
    """Replace cell references in *formula* with human-readable labels.

    Example::

        =B12*C12  -->  顧客数 * ARPU

    The replacement preserves operators and function calls.
    """
    if not formula or not isinstance(formula, str):
        return formula or ""

    text = formula

    # Replace cross-sheet references first (they contain '!' which
    # could interfere with bare-ref matching)
    def _replace_cross_sheet(m: re.Match) -> str:
        full = m.group(0)
        # Try exact lookup
        label = label_map.get(full)
        if label:
            return label
        # Try normalised form
        normalised = _normalise_address(full)
        label = label_map.get(normalised)
        return label if label else full

    # Pattern for 'Sheet'!REF or Sheet!REF
    cross_sheet_re = re.compile(
        r"(?:'[^']+'|[A-Za-z_]\w*)!\$?[A-Z]{1,3}\$?\d+"
        r"(?::\$?[A-Z]{1,3}\$?\d+)?",
        re.IGNORECASE,
    )
    text = cross_sheet_re.sub(_replace_cross_sheet, text)

    # Replace bare cell refs (B12, $B$12, etc.)
    def _replace_bare(m: re.Match) -> str:
        ref = m.group(0)
        # Try with current normalisation
        label = label_map.get(ref)
        if label:
            return label
        # Try stripping $ signs
        clean = ref.replace("$", "")
        label = label_map.get(clean)
        return label if label else ref

    bare_ref_re = re.compile(r"\$?[A-Z]{1,3}\$?\d+", re.IGNORECASE)
    text = bare_ref_re.sub(_replace_bare, text)

    # Prettify operators
    text = text.replace("*", " \u00d7 ")   # multiplication sign
    text = text.replace("/", " \u00f7 ")   # division sign
    text = text.replace("+", " + ")
    text = text.replace("-", " - ")

    # Clean up extra whitespace
    text = re.sub(r"\s{2,}", " ", text).strip()

    return text


# ===================================================================
# Label map construction
# ===================================================================

def _build_label_map(
    catalog: InputCatalog,
    wb: openpyxl.Workbook,
) -> Dict[str, str]:
    """Build a mapping from cell addresses to their best label.

    Sources:
    1. ``label_candidates`` from the catalog items.
    2. Scanning each worksheet for text cells adjacent to formula cells.
    """
    label_map: Dict[str, str] = {}

    # From catalog items
    for item in catalog.items:
        addr = f"'{item.sheet}'!{item.cell}"
        if item.label_candidates:
            label_map[addr] = item.label_candidates[0]
            # Also store bare coordinate for intra-sheet lookup
            label_map[item.cell] = item.label_candidates[0]

    # Scan worksheets for additional labels near formula cells
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                addr = f"'{ws.title}'!{cell.coordinate}"
                if addr in label_map:
                    continue
                # Try to find a label for this cell from its left neighbour
                if cell.column > 1:
                    left = ws.cell(row=cell.row, column=cell.column - 1)
                    if (
                        not isinstance(left, MergedCell)
                        and left.value
                        and isinstance(left.value, str)
                        and not left.value.startswith("=")
                    ):
                        label_map[addr] = left.value.strip()
                        label_map[cell.coordinate] = left.value.strip()
                        continue
                # Try row above
                if cell.row > 1:
                    above = ws.cell(row=cell.row - 1, column=cell.column)
                    if (
                        not isinstance(above, MergedCell)
                        and above.value
                        and isinstance(above.value, str)
                        and not above.value.startswith("=")
                    ):
                        label_map[addr] = above.value.strip()
                        label_map[cell.coordinate] = above.value.strip()

    return label_map


# ===================================================================
# KPI identification
# ===================================================================

def _identify_kpi_cells(
    ws: Worksheet,
) -> List[Tuple[str, str, str]]:
    """Return ``(sheet_name, coordinate, label)`` for KPI cells in *ws*.

    A cell is considered a KPI when:
    1. It is in a sheet whose name matches ``_KPI_SHEET_PATTERNS``, OR
    2. A nearby label matches ``_KPI_LABEL_PATTERNS``.

    Only formula cells are considered as KPIs (value cells are inputs).
    """
    is_kpi_sheet = any(p.search(ws.title) for p in _KPI_SHEET_PATTERNS)
    results: List[Tuple[str, str, str]] = []

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if not _cell_is_formula(cell):
                continue

            # Determine the label for this cell
            label = _get_cell_label(ws, cell)
            if not label:
                continue

            # Check if label matches a KPI pattern
            is_kpi_label = any(p.search(label) for p in _KPI_LABEL_PATTERNS)

            if is_kpi_sheet or is_kpi_label:
                results.append((ws.title, cell.coordinate, label))

    return results


def _get_cell_label(ws: Worksheet, cell: Cell) -> str:
    """Return the best label for *cell* by checking neighbours."""
    # Check left (same row)
    for c in range(cell.column - 1, max(0, cell.column - 5), -1):
        neighbour = ws.cell(row=cell.row, column=c)
        if isinstance(neighbour, MergedCell):
            continue
        if (
            neighbour.value
            and isinstance(neighbour.value, str)
            and not neighbour.value.startswith("=")
        ):
            return neighbour.value.strip()

    # Check above
    if cell.row > 1:
        above = ws.cell(row=cell.row - 1, column=cell.column)
        if (
            not isinstance(above, MergedCell)
            and above.value
            and isinstance(above.value, str)
            and not above.value.startswith("=")
        ):
            return above.value.strip()

    return ""


# ===================================================================
# Dependency collection & tree building
# ===================================================================

def _collect_dependencies(
    address: str,
    formulas: List[FormulaInfo],
    input_addresses: Set[str],
    label_map: Dict[str, str],
    *,
    max_depth: int = 10,
) -> List[str]:
    """Recursively collect input dependencies for *address*.

    Returns a list of human-readable dependency descriptions:
    ``"label (address)"`` for each input cell the KPI ultimately depends on.
    """
    formula_lookup = {
        f"'{f.sheet}'!{f.cell}": f for f in formulas
    }
    deps: List[str] = []
    seen: Set[str] = set()
    _collect_deps_recursive(
        address, formula_lookup, input_addresses, label_map,
        deps, seen, depth=0, max_depth=max_depth,
    )
    return deps


def _collect_deps_recursive(
    address: str,
    formula_lookup: Dict[str, FormulaInfo],
    input_addresses: Set[str],
    label_map: Dict[str, str],
    deps: List[str],
    seen: Set[str],
    depth: int,
    max_depth: int,
) -> None:
    if depth > max_depth or address in seen:
        return
    seen.add(address)

    fi = formula_lookup.get(address)
    if fi is None:
        # It's a leaf -- possibly an input cell
        if address in input_addresses:
            label = label_map.get(address, "")
            desc = f"{label} ({address})" if label else address
            if desc not in deps:
                deps.append(desc)
        return

    for ref in fi.referenced_cells:
        normalised = _normalise_address(ref)
        _collect_deps_recursive(
            normalised, formula_lookup, input_addresses, label_map,
            deps, seen, depth + 1, max_depth,
        )


def _build_dependency_tree_node(
    address: str,
    formula_lookup: Dict[str, FormulaInfo],
    input_addresses: Set[str],
    label_map: Dict[str, str],
    visited: Set[str],
    *,
    max_depth: int = 10,
    _depth: int = 0,
) -> DependencyNode:
    """Recursively build a dependency tree rooted at *address*."""
    label = label_map.get(address, "")
    is_input = address in input_addresses
    is_kpi = address in formula_lookup  # has a formula -> computed cell

    node = DependencyNode(
        address=address,
        label=label,
        is_input=is_input,
        is_kpi=is_kpi,
    )

    if _depth >= max_depth or address in visited:
        return node

    visited.add(address)

    fi = formula_lookup.get(address)
    if fi is not None:
        for ref in fi.referenced_cells:
            normalised = _normalise_address(ref)
            child = _build_dependency_tree_node(
                normalised,
                formula_lookup,
                input_addresses,
                label_map,
                visited,
                max_depth=max_depth,
                _depth=_depth + 1,
            )
            node.children.append(child)

    return node


def _build_dependency_tree(
    ws: Worksheet,
    kpi_cells: List[str],
    label_map: Dict[str, str],
) -> Dict[str, List[str]]:
    """Build a flat dependency mapping from KPI addresses to input labels.

    This is a simpler alternative to the full tree -- returns
    ``{kpi_address: [dep_label, ...]}`` for quick lookups.
    """
    result: Dict[str, List[str]] = {}
    for coord in kpi_cells:
        cell = ws[coord]
        if not _cell_is_formula(cell):
            continue
        refs = _parse_formula_refs(str(cell.value))
        dep_labels: List[str] = []
        for ref in refs:
            label = label_map.get(ref, ref)
            dep_labels.append(label)
        addr = f"'{ws.title}'!{coord}"
        result[addr] = dep_labels
    return result


# ===================================================================
# Helpers
# ===================================================================

def _cell_is_formula(cell: Cell) -> bool:
    """Check whether a cell contains a formula."""
    if cell.data_type == "f":
        return True
    if isinstance(cell.value, str) and cell.value.startswith("="):
        return True
    return False


def _normalise_address(ref: str) -> str:
    """Normalise a cell reference by stripping ``$`` anchors.

    ``'Sheet'!$B$12`` -> ``'Sheet'!B12``
    ``$C$5``          -> ``C5``
    """
    if "!" in ref:
        sheet_part, cell_part = ref.split("!", 1)
        cell_part = cell_part.replace("$", "")
        # Strip range notation for tree-building (take first cell)
        if ":" in cell_part:
            cell_part = cell_part.split(":")[0]
        return f"{sheet_part}!{cell_part}"
    clean = ref.replace("$", "")
    if ":" in clean:
        clean = clean.split(":")[0]
    return clean


def _render_tree(
    node: DependencyNode,
    lines: List[str],
    indent: int,
    *,
    prefix: str = "",
    is_last: bool = True,
) -> None:
    """Render a dependency tree node as indented text lines."""
    connector = "\u2514\u2500 " if is_last else "\u251c\u2500 "
    if indent == 0:
        display = node.label or node.address
        tag = ""
        if node.is_input:
            tag = " [INPUT]"
        elif node.is_kpi:
            tag = " [KPI]"
        lines.append(f"{display}{tag}")
    else:
        display = node.label or node.address
        tag = ""
        if node.is_input:
            tag = " [INPUT]"
        elif node.is_kpi:
            tag = " [KPI]"
        lines.append(f"{prefix}{connector}{display}{tag}")

    # Prepare prefix for children
    if indent == 0:
        child_prefix = ""
    else:
        child_prefix = prefix + ("\u2502  " if not is_last else "   ")

    for i, child in enumerate(node.children):
        _render_tree(
            child,
            lines,
            indent + 1,
            prefix=child_prefix if indent > 0 else "",
            is_last=(i == len(node.children) - 1),
        )
