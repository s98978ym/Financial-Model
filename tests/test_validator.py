"""Tests for src.excel.validator -- post-generation validation.

Validates that the PLValidator correctly detects formula changes,
new errors, fullCalcOnLoad, and distinguishes input vs non-input
cell changes.
"""

import shutil
from pathlib import Path

import pytest
import openpyxl
from openpyxl.styles import PatternFill

from src.config.models import ValidationResult
from src.excel.validator import PLValidator, EXCEL_ERRORS


# ===================================================================
# Helpers
# ===================================================================

INPUT_FILL = PatternFill(
    start_color="FFFFF2CC",
    end_color="FFFFF2CC",
    fill_type="solid",
)


def _make_valid_output(template_path: str, output_path: str, changes: dict = None):
    """Create a valid output file from a template.

    *changes* maps ``(sheet, cell)`` -> new_value and only applies to
    input cells (non-formula).  fullCalcOnLoad is set to True.
    """
    shutil.copy2(template_path, output_path)
    wb = openpyxl.load_workbook(output_path)

    if changes:
        for (sheet, cell_ref), value in changes.items():
            ws = wb[sheet]
            cell = ws[cell_ref]
            # Only change if it is NOT a formula
            if not (isinstance(cell.value, str) and str(cell.value).startswith("=")):
                cell.value = value

    # Set fullCalcOnLoad
    wb.calculation = openpyxl.workbook.properties.CalcProperties(fullCalcOnLoad=True)
    wb.save(output_path)
    wb.close()


def _make_broken_output_formula_changed(template_path: str, output_path: str):
    """Create an output where a formula cell has been overwritten."""
    shutil.copy2(template_path, output_path)
    wb = openpyxl.load_workbook(output_path)
    ws = wb["PL\u8a2d\u8a08"]
    # Overwrite formula in B5 with a plain number
    ws["B5"] = 99999
    wb.calculation = openpyxl.workbook.properties.CalcProperties(fullCalcOnLoad=True)
    wb.save(output_path)
    wb.close()


def _make_broken_output_no_calc(template_path: str, output_path: str):
    """Create an output without fullCalcOnLoad set.

    Explicitly clears any calculation properties so the validator
    sees ``wb.calculation is None`` and flags the missing setting.
    """
    shutil.copy2(template_path, output_path)
    wb = openpyxl.load_workbook(output_path)
    # Explicitly clear calculation properties so fullCalcOnLoad is absent
    wb.calculation = None
    wb.save(output_path)
    wb.close()


def _make_broken_output_with_error(template_path: str, output_path: str):
    """Create an output with a new Excel error value."""
    shutil.copy2(template_path, output_path)
    wb = openpyxl.load_workbook(output_path)
    ws = wb["PL\u8a2d\u8a08"]
    # Insert a new error that was not in the template
    ws["D3"] = "#REF!"
    wb.calculation = openpyxl.workbook.properties.CalcProperties(fullCalcOnLoad=True)
    wb.save(output_path)
    wb.close()


# ===================================================================
# Validation passes
# ===================================================================

class TestValidationPasses:
    """Test that validation passes for correctly generated output."""

    def test_valid_output_passes(self, sample_template_path, tmp_path):
        output = str(tmp_path / "valid_output.xlsx")
        _make_valid_output(
            sample_template_path,
            output,
            changes={("PL\u8a2d\u8a08", "B3"): 5000, ("PL\u8a2d\u8a08", "B4"): 2000},
        )
        validator = PLValidator(sample_template_path, output)
        result = validator.validate()
        assert isinstance(result, ValidationResult)
        assert result.passed is True
        assert result.formula_preserved is True
        assert result.no_new_errors is True
        assert result.full_calc_on_load is True

    def test_valid_output_changed_cells_tracked(self, sample_template_path, tmp_path):
        output = str(tmp_path / "valid_output.xlsx")
        _make_valid_output(
            sample_template_path,
            output,
            changes={("PL\u8a2d\u8a08", "B3"): 5000},
        )
        validator = PLValidator(sample_template_path, output)
        result = validator.validate()
        # B3 was changed -- should appear in changed_cells
        changed_coords = [(c.sheet, c.cell) for c in result.changed_cells]
        assert ("PL\u8a2d\u8a08", "B3") in changed_coords

    def test_unmodified_output_passes(self, sample_template_path, tmp_path):
        """An output identical to template (plus fullCalcOnLoad) should pass."""
        output = str(tmp_path / "same_output.xlsx")
        _make_valid_output(sample_template_path, output)
        validator = PLValidator(sample_template_path, output)
        result = validator.validate()
        assert result.passed is True
        assert result.formula_preserved is True


# ===================================================================
# Validation failures: formula changed
# ===================================================================

class TestFormulaChanged:
    """Test that validation detects formula cell modifications."""

    def test_formula_change_detected(self, sample_template_path, tmp_path):
        output = str(tmp_path / "broken_formula.xlsx")
        _make_broken_output_formula_changed(sample_template_path, output)
        validator = PLValidator(sample_template_path, output)
        result = validator.validate()
        assert result.passed is False
        assert result.formula_preserved is False

    def test_formula_change_in_errors(self, sample_template_path, tmp_path):
        output = str(tmp_path / "broken_formula.xlsx")
        _make_broken_output_formula_changed(sample_template_path, output)
        validator = PLValidator(sample_template_path, output)
        result = validator.validate()
        assert len(result.errors_found) > 0
        error_text = " ".join(result.errors_found)
        assert "FORMULA" in error_text.upper()


# ===================================================================
# Validation failures: fullCalcOnLoad
# ===================================================================

class TestFullCalcOnLoad:
    """Test that missing fullCalcOnLoad is detected."""

    def test_missing_calc_detected(self, sample_template_path, tmp_path):
        output = str(tmp_path / "no_calc.xlsx")
        _make_broken_output_no_calc(sample_template_path, output)
        validator = PLValidator(sample_template_path, output)
        result = validator.validate()
        assert result.passed is False
        assert result.full_calc_on_load is False


# ===================================================================
# Validation failures: new errors
# ===================================================================

class TestNewErrors:
    """Test that new Excel error values are detected."""

    def test_new_error_detected(self, sample_template_path, tmp_path):
        output = str(tmp_path / "with_error.xlsx")
        _make_broken_output_with_error(sample_template_path, output)
        validator = PLValidator(sample_template_path, output)
        result = validator.validate()
        assert result.no_new_errors is False

    def test_new_error_in_errors_found(self, sample_template_path, tmp_path):
        output = str(tmp_path / "with_error.xlsx")
        _make_broken_output_with_error(sample_template_path, output)
        validator = PLValidator(sample_template_path, output)
        result = validator.validate()
        assert len(result.errors_found) > 0
        error_text = " ".join(result.errors_found)
        assert "#REF!" in error_text


# ===================================================================
# EXCEL_ERRORS constant
# ===================================================================

class TestExcelErrors:
    """Test that the EXCEL_ERRORS constant is correct."""

    def test_contains_common_errors(self):
        assert "#REF!" in EXCEL_ERRORS
        assert "#DIV/0!" in EXCEL_ERRORS
        assert "#NAME?" in EXCEL_ERRORS
        assert "#N/A" in EXCEL_ERRORS
        assert "#VALUE!" in EXCEL_ERRORS
        assert "#NUM!" in EXCEL_ERRORS
        assert "#NULL!" in EXCEL_ERRORS


# ===================================================================
# Validation result attributes
# ===================================================================

class TestValidationResultAttributes:
    """Test that ValidationResult has all expected fields."""

    def test_result_has_all_fields(self, sample_template_path, tmp_path):
        output = str(tmp_path / "output.xlsx")
        _make_valid_output(sample_template_path, output)
        validator = PLValidator(sample_template_path, output)
        result = validator.validate()

        assert hasattr(result, "passed")
        assert hasattr(result, "formula_preserved")
        assert hasattr(result, "no_new_errors")
        assert hasattr(result, "full_calc_on_load")
        assert hasattr(result, "changed_cells")
        assert hasattr(result, "errors_found")
        assert hasattr(result, "warnings")

    def test_errors_found_is_list(self, sample_template_path, tmp_path):
        output = str(tmp_path / "output.xlsx")
        _make_valid_output(sample_template_path, output)
        validator = PLValidator(sample_template_path, output)
        result = validator.validate()
        assert isinstance(result.errors_found, list)

    def test_warnings_is_list(self, sample_template_path, tmp_path):
        output = str(tmp_path / "output.xlsx")
        _make_valid_output(sample_template_path, output)
        validator = PLValidator(sample_template_path, output)
        result = validator.validate()
        assert isinstance(result.warnings, list)
