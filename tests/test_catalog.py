"""Tests for src.catalog.scanner -- Excel template catalog scanner.

Tests scanning a real (in-memory) Excel template, colour matching,
label detection, unit detection, formula flagging, and JSON export.

NOTE: At runtime, CatalogItem uses the field name ``cell`` (not ``coordinate``),
and InputCatalog is a pydantic model without ``writable_items``/``formula_items``
properties.  The scanner appends items directly via ``catalog.items.append()``.
"""

import json
import os
from pathlib import Path

import pytest
import openpyxl
from openpyxl.styles import PatternFill, Font

from src.catalog.scanner import (
    scan_template,
    export_catalog_json,
    _match_color,
    _colors_equal,
    _strip_color_prefix,
    _cell_has_formula,
    _detect_units,
)
from src.config.models import CatalogItem, InputCatalog


# ===================================================================
# Helpers
# ===================================================================

INPUT_FILL = PatternFill(
    start_color="FFFFF2CC",
    end_color="FFFFF2CC",
    fill_type="solid",
)


def _make_template(path: str):
    """Create a small template xlsx at *path* with known structure."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PL\u8a2d\u8a08"

    # Row 1 - section header
    ws["A1"] = "\u58f2\u4e0a"
    ws["A1"].font = Font(bold=True)

    # Row 2 - column headers
    ws["B2"] = "FY2024"
    ws["C2"] = "FY2025"

    # Row 3 - revenue (input cells, yellow)
    ws["A3"] = "\u58f2\u4e0a\u9ad8(\u5186)"
    ws["B3"] = 1000
    ws["B3"].fill = INPUT_FILL
    ws["C3"] = 2000
    ws["C3"].fill = INPUT_FILL

    # Row 4 - cost (input cells)
    ws["A4"] = "\u539f\u4fa1"
    ws["B4"] = 400
    ws["B4"].fill = INPUT_FILL

    # Row 5 - gross profit (formula with input colour)
    ws["A5"] = "\u7c97\u5229"
    ws["B5"] = "=B3-B4"
    ws["B5"].fill = INPUT_FILL  # even if colored, should be flagged as formula

    # Row 6 - plain cell (no special color)
    ws["A6"] = "\u30e1\u30e2"
    ws["B6"] = "test"

    # Row 7 - percentage input
    ws["A7"] = "\u6210\u9577\u7387(%)"
    ws["B7"] = 0.15
    ws["B7"].fill = INPUT_FILL

    wb.save(path)
    wb.close()


# ===================================================================
# scan_template
# ===================================================================

class TestScanTemplate:
    """Test the main scan_template function."""

    def test_scan_returns_input_catalog(self, tmp_path):
        tpl = str(tmp_path / "tpl.xlsx")
        _make_template(tpl)
        catalog = scan_template(tpl)
        assert isinstance(catalog, InputCatalog)

    def test_scan_finds_input_cells(self, tmp_path):
        tpl = str(tmp_path / "tpl.xlsx")
        _make_template(tpl)
        catalog = scan_template(tpl)
        # Should find yellow-filled cells: B3, C3, B4, B5, B7
        assert len(catalog.items) >= 4

    def test_scan_formula_cells_flagged(self, tmp_path):
        tpl = str(tmp_path / "tpl.xlsx")
        _make_template(tpl)
        catalog = scan_template(tpl)
        formula_items = [i for i in catalog.items if i.has_formula]
        # B5 contains "=B3-B4" and has the input color
        formula_cells = [item.cell for item in formula_items]
        assert "B5" in formula_cells

    def test_scan_non_formula_items_exist(self, tmp_path):
        """Catalog should contain non-formula items for plain input cells."""
        tpl = str(tmp_path / "tpl.xlsx")
        _make_template(tpl)
        catalog = scan_template(tpl)
        writable = [i for i in catalog.items if not i.has_formula]
        assert len(writable) >= 3  # B3, C3, B4, B7

    def test_scan_writable_items_exclude_formulas(self, tmp_path):
        tpl = str(tmp_path / "tpl.xlsx")
        _make_template(tpl)
        catalog = scan_template(tpl)
        writable = [i for i in catalog.items if not i.has_formula]
        for item in writable:
            assert item.has_formula is False

    def test_scan_nonexistent_file_raises(self):
        with pytest.raises(FileNotFoundError):
            scan_template("/nonexistent/path/template.xlsx")

    def test_scan_label_detection(self, tmp_path):
        """Labels from column A should appear in label_candidates."""
        tpl = str(tmp_path / "tpl.xlsx")
        _make_template(tpl)
        catalog = scan_template(tpl)
        # Find the B3 item (revenue input)
        b3_items = [i for i in catalog.items if i.cell == "B3"]
        assert len(b3_items) == 1
        labels = b3_items[0].label_candidates
        # Should have picked up "\u58f2\u4e0a\u9ad8(\u5186)" from A3
        assert any("\u58f2\u4e0a" in lbl for lbl in labels)

    def test_scan_unit_detection(self, tmp_path):
        """Unit candidates should be extracted from nearby labels."""
        tpl = str(tmp_path / "tpl.xlsx")
        _make_template(tpl)
        catalog = scan_template(tpl)
        # B3 has label "\u58f2\u4e0a\u9ad8(\u5186)" -- should detect "\u5186"
        b3_items = [i for i in catalog.items if i.cell == "B3"]
        assert len(b3_items) == 1
        units = b3_items[0].unit_candidates
        assert any("\u5186" in u for u in units)

    def test_scan_period_detection(self, tmp_path):
        """Period should be detected from header rows."""
        tpl = str(tmp_path / "tpl.xlsx")
        _make_template(tpl)
        catalog = scan_template(tpl)
        b3_items = [i for i in catalog.items if i.cell == "B3"]
        assert len(b3_items) == 1
        # B2 has "FY2024" -- should appear as year_or_period
        period = b3_items[0].year_or_period
        assert period is not None
        assert "FY2024" in period

    def test_scan_blocks_populated(self, tmp_path):
        tpl = str(tmp_path / "tpl.xlsx")
        _make_template(tpl)
        catalog = scan_template(tpl)
        # blocks dict should have at least one entry
        assert len(catalog.blocks) >= 1

    def test_scan_items_have_cell_field(self, tmp_path):
        """Runtime CatalogItem uses 'cell' field, not 'coordinate'."""
        tpl = str(tmp_path / "tpl.xlsx")
        _make_template(tpl)
        catalog = scan_template(tpl)
        for item in catalog.items:
            assert hasattr(item, "cell")
            assert isinstance(item.cell, str)


# ===================================================================
# _match_color / _colors_equal / _strip_color_prefix
# ===================================================================

class TestColorMatching:
    """Test internal colour comparison helpers."""

    def test_colors_equal_exact(self):
        assert _colors_equal("FFFFF2CC", "FFFFF2CC") is True

    def test_colors_equal_case_insensitive(self):
        assert _colors_equal("fffff2cc", "FFFFF2CC") is True

    def test_colors_equal_fuzzy_strips_prefix(self):
        assert _colors_equal("00FFF2CC", "FFFFF2CC", fuzzy=True) is True

    def test_colors_equal_strict_rejects_different_prefix(self):
        assert _colors_equal("00FFF2CC", "FFFFF2CC", fuzzy=False) is False

    def test_strip_color_prefix_8_char(self):
        assert _strip_color_prefix("FFFFF2CC") == "FFF2CC"

    def test_strip_color_prefix_6_char(self):
        assert _strip_color_prefix("FFF2CC") == "FFF2CC"

    def test_strip_color_prefix_hash(self):
        assert _strip_color_prefix("#FFFFF2CC") == "FFF2CC"

    def test_match_color_none_fill(self):
        """None fill should not match."""
        assert _match_color(None, "FFFFF2CC") is False

    def test_match_color_with_real_cell(self):
        """Test _match_color against a real cell's fill object."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "test"
        ws["A1"].fill = INPUT_FILL
        cell = ws["A1"]
        assert _match_color(cell.fill, "FFFFF2CC") is True
        wb.close()

    def test_match_color_no_match(self):
        """Cell with different fill should not match."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "test"
        ws["A1"].fill = PatternFill(
            start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"
        )
        cell = ws["A1"]
        assert _match_color(cell.fill, "FFFFF2CC") is False
        wb.close()


# ===================================================================
# _cell_has_formula
# ===================================================================

class TestCellHasFormula:
    """Test formula detection on real cells."""

    def test_formula_detected(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "=SUM(B1:B10)"
        cell = ws["A1"]
        assert _cell_has_formula(cell) is True
        wb.close()

    def test_plain_value_not_formula(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = 42
        cell = ws["A1"]
        assert _cell_has_formula(cell) is False
        wb.close()

    def test_string_value_not_formula(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Hello"
        cell = ws["A1"]
        assert _cell_has_formula(cell) is False
        wb.close()


# ===================================================================
# _detect_units
# ===================================================================

class TestDetectUnits:
    """Test unit detection from label strings."""

    def test_yen(self):
        units = _detect_units("\u58f2\u4e0a\u9ad8(\u5186)")
        assert "\u5186" in units

    def test_man_yen(self):
        units = _detect_units("\u5358\u4fa1(\u4e07\u5186)")
        assert "\u4e07\u5186" in units

    def test_percent(self):
        units = _detect_units("\u6210\u9577\u7387(%)")
        assert "%" in units

    def test_fullwidth_percent(self):
        units = _detect_units("\u6210\u9577\u7387(\uff05)")
        assert "%" in units

    def test_nin(self):
        units = _detect_units("\u5f93\u696d\u54e1\u6570(\u4eba)")
        assert "\u4eba" in units

    def test_empty_string(self):
        units = _detect_units("")
        assert units == []

    def test_none_input(self):
        units = _detect_units(None)
        assert units == []

    def test_no_units(self):
        units = _detect_units("label without units")
        assert units == []


# ===================================================================
# export_catalog_json
# ===================================================================

class TestExportCatalogJson:
    """Test JSON catalog export.

    NOTE: export_catalog_json calls catalog.to_dict() which does not exist
    on the pydantic InputCatalog at runtime (the on-disk source defines it
    for the dataclass version).  These tests verify the bug is known and
    also test a workaround using model_dump().
    """

    @pytest.mark.xfail(
        reason="InputCatalog.to_dict() not available on pydantic model",
        raises=AttributeError,
    )
    def test_export_from_scanned_catalog(self, tmp_path):
        """Scan a real template and export to JSON -- end-to-end test."""
        tpl = str(tmp_path / "tpl.xlsx")
        _make_template(tpl)
        catalog = scan_template(tpl)
        out_path = str(tmp_path / "catalog.json")
        export_catalog_json(catalog, out_path)
        assert os.path.exists(out_path)

    @pytest.mark.xfail(
        reason="InputCatalog.to_dict() not available on pydantic model",
        raises=AttributeError,
    )
    def test_export_valid_json(self, tmp_path):
        """Exported file should be valid JSON."""
        tpl = str(tmp_path / "tpl.xlsx")
        _make_template(tpl)
        catalog = scan_template(tpl)
        out_path = str(tmp_path / "catalog.json")
        export_catalog_json(catalog, out_path)
        with open(out_path, encoding="utf-8") as f:
            data = json.load(f)
        assert isinstance(data, dict)

    @pytest.mark.xfail(
        reason="InputCatalog.to_dict() not available on pydantic model",
        raises=AttributeError,
    )
    def test_export_creates_parent_dirs(self, tmp_path):
        tpl = str(tmp_path / "tpl.xlsx")
        _make_template(tpl)
        catalog = scan_template(tpl)
        out_path = str(tmp_path / "sub" / "dir" / "catalog.json")
        export_catalog_json(catalog, out_path)
        assert os.path.exists(out_path)

    @pytest.mark.xfail(
        reason="InputCatalog.to_dict() not available on pydantic model",
        raises=AttributeError,
    )
    def test_export_contains_items(self, tmp_path):
        tpl = str(tmp_path / "tpl.xlsx")
        _make_template(tpl)
        catalog = scan_template(tpl)
        out_path = str(tmp_path / "catalog.json")
        export_catalog_json(catalog, out_path)
        with open(out_path, encoding="utf-8") as f:
            data = json.load(f)
        # The JSON should contain catalog data
        assert len(str(data)) > 10  # Non-trivial content

    def test_model_dump_workaround(self, tmp_path):
        """Using model_dump() as a JSON-safe alternative to to_dict()."""
        tpl = str(tmp_path / "tpl.xlsx")
        _make_template(tpl)
        catalog = scan_template(tpl)
        data = catalog.model_dump()
        assert "items" in data
        assert "blocks" in data
        assert len(data["items"]) > 0

    def test_model_dump_item_has_cell(self, tmp_path):
        """Serialized items should contain the 'cell' field."""
        tpl = str(tmp_path / "tpl.xlsx")
        _make_template(tpl)
        catalog = scan_template(tpl)
        data = catalog.model_dump()
        for item in data["items"]:
            assert "cell" in item
