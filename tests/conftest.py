"""Shared fixtures for the PL Generator test suite.

Provides reusable Excel template fixtures, configuration objects,
and sample extracted parameters that mirror real project structures.
"""

import os
import tempfile
from pathlib import Path

import pytest
import openpyxl
from openpyxl.styles import PatternFill, Font

from src.config.models import (
    CatalogItem,
    CellTarget,
    ColorConfig,
    Evidence,
    ExtractedParameter,
    InputCatalog,
    PhaseAConfig,
    ValidationResult,
)


# ---------------------------------------------------------------------------
# Directory / path helpers
# ---------------------------------------------------------------------------

@pytest.fixture
def tmp_dir(tmp_path):
    """Return a temporary directory path (pathlib.Path)."""
    return tmp_path


# ---------------------------------------------------------------------------
# Excel template fixtures
# ---------------------------------------------------------------------------

INPUT_FILL = PatternFill(
    start_color="FFFFF2CC",
    end_color="FFFFF2CC",
    fill_type="solid",
)

HEADER_FONT = Font(bold=True, size=12)


def _build_template_workbook():
    """Build a small Excel workbook that mimics a real PL template.

    Sheet layout for "PL\u8a2d\u8a08":

        Row 1:  A1="PL\u8a2d\u8a08" (bold header)
        Row 2:  A2="\u58f2\u4e0a", B2="FY2024", C2="FY2025"
        Row 3:  A3="\u58f2\u4e0a\u9ad8", B3=1000 (input, yellow), C3=2000 (input, yellow)
        Row 4:  A4="\u539f\u4fa1",  B4=400  (input, yellow), C4=800  (input, yellow)
        Row 5:  A5="\u7c97\u5229",  B5="=B3-B4" (formula),    C5="=C3-C4" (formula)
        Row 6:  A6="\u8ca9\u7ba1\u8cbb", B6=200 (input, yellow), C6=300 (input, yellow)
        Row 7:  A7="\u55b6\u696d\u5229\u76ca", B7="=B5-B6" (formula), C7="=C5-C6" (formula)
        Row 8:  A8="\u5358\u4fa1(\u5186)", B8=500  (input, yellow)
        Row 9:  A9="" (empty)
        Row 10: A10="\u5408\u8a08",  B10="=SUM(B3:B4)" (formula)

    The workbook also contains a second sheet "KPI" with a few cells.
    """
    wb = openpyxl.Workbook()

    # --- Sheet 1: PL\u8a2d\u8a08 ---
    ws = wb.active
    ws.title = "PL\u8a2d\u8a08"

    # Row 1 - header
    ws["A1"] = "PL\u8a2d\u8a08"
    ws["A1"].font = HEADER_FONT

    # Row 2 - column headers
    ws["A2"] = "\u58f2\u4e0a"
    ws["A2"].font = Font(bold=True)
    ws["B2"] = "FY2024"
    ws["C2"] = "FY2025"

    # Row 3 - revenue (input cells)
    ws["A3"] = "\u58f2\u4e0a\u9ad8"
    ws["B3"] = 1000
    ws["B3"].fill = INPUT_FILL
    ws["C3"] = 2000
    ws["C3"].fill = INPUT_FILL

    # Row 4 - COGS (input cells)
    ws["A4"] = "\u539f\u4fa1"
    ws["B4"] = 400
    ws["B4"].fill = INPUT_FILL
    ws["C4"] = 800
    ws["C4"].fill = INPUT_FILL

    # Row 5 - gross profit (formula cells)
    ws["A5"] = "\u7c97\u5229"
    ws["B5"] = "=B3-B4"
    ws["C5"] = "=C3-C4"

    # Row 6 - SGA (input cells)
    ws["A6"] = "\u8ca9\u7ba1\u8cbb"
    ws["B6"] = 200
    ws["B6"].fill = INPUT_FILL
    ws["C6"] = 300
    ws["C6"].fill = INPUT_FILL

    # Row 7 - operating profit (formula cells)
    ws["A7"] = "\u55b6\u696d\u5229\u76ca"
    ws["B7"] = "=B5-B6"
    ws["C7"] = "=C5-C6"

    # Row 8 - unit price (input, with unit label)
    ws["A8"] = "\u5358\u4fa1(\u5186)"
    ws["B8"] = 500
    ws["B8"].fill = INPUT_FILL

    # Row 9 - empty row
    # (left blank intentionally)

    # Row 10 - total row (formula)
    ws["A10"] = "\u5408\u8a08"
    ws["B10"] = "=SUM(B3:B4)"

    # --- Sheet 2: KPI ---
    ws2 = wb.create_sheet(title="KPI")
    ws2["A1"] = "KPI\u4e00\u89a7"
    ws2["A1"].font = HEADER_FONT
    ws2["A2"] = "\u6708\u6b21\u89e3\u7d04\u7387"
    ws2["B2"] = 0.05
    ws2["B2"].fill = INPUT_FILL
    ws2["A3"] = "\u6210\u9577\u7387"
    ws2["B3"] = 0.15
    ws2["B3"].fill = INPUT_FILL

    return wb


@pytest.fixture
def sample_template_path(tmp_path):
    """Create a sample .xlsx template on disk and return its path as a string."""
    wb = _build_template_workbook()
    file_path = tmp_path / "test_template.xlsx"
    wb.save(str(file_path))
    wb.close()
    return str(file_path)


@pytest.fixture
def sample_template_wb():
    """Return an in-memory openpyxl Workbook (not saved to disk)."""
    return _build_template_workbook()


# ---------------------------------------------------------------------------
# Configuration fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def default_color_config():
    """Return a ColorConfig with default values."""
    return ColorConfig()


@pytest.fixture
def default_phase_a_config():
    """Return a PhaseAConfig with sensible defaults for testing."""
    return PhaseAConfig(
        industry="SaaS",
        business_model="B2B",
        strictness="normal",
        cases=["base", "worst"],
        simulation=False,
    )


@pytest.fixture
def phase_a_config_with_template(sample_template_path):
    """Return a PhaseAConfig for testing (template path stored externally)."""
    return PhaseAConfig(
        industry="SaaS",
        business_model="B2B",
        strictness="normal",
        cases=["base", "best", "worst"],
        simulation=False,
    )


# ---------------------------------------------------------------------------
# Extracted parameter fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def sample_evidence():
    """Return a sample Evidence instance."""
    return Evidence(
        quote="\u58f2\u4e0a\u9ad8\u306f\u7d041.5\u5104\u5186\u3092\u898b\u8fbc\u3080",
        page_or_slide="3",
        rationale="\u4e8b\u696d\u8a08\u753b\u66f8p3\u306e\u58f2\u4e0a\u4e88\u6e2c\u304b\u3089\u62bd\u51fa",
    )


@pytest.fixture
def sample_parameters(sample_evidence):
    """Return a list of sample ExtractedParameter instances."""
    return [
        ExtractedParameter(
            key="revenue_fy2024",
            label="\u58f2\u4e0a\u9ad8 FY2024",
            value=150000000,
            unit="\u5186",
            mapped_targets=[CellTarget(sheet="PL\u8a2d\u8a08", cell="B3")],
            evidence=sample_evidence,
            confidence=0.9,
            source="document",
            selected=True,
            adjusted_value=None,
        ),
        ExtractedParameter(
            key="cogs_fy2024",
            label="\u539f\u4fa1 FY2024",
            value=60000000,
            unit="\u5186",
            mapped_targets=[CellTarget(sheet="PL\u8a2d\u8a08", cell="B4")],
            confidence=0.7,
            source="document",
            selected=True,
            adjusted_value=None,
        ),
        ExtractedParameter(
            key="sga_fy2024",
            label="\u8ca9\u7ba1\u8cbb FY2024",
            value=30000000,
            unit="\u5186",
            mapped_targets=[CellTarget(sheet="PL\u8a2d\u8a08", cell="B6")],
            confidence=0.6,
            source="inferred",
            selected=True,
            adjusted_value=None,
        ),
        ExtractedParameter(
            key="churn_rate",
            label="\u6708\u6b21\u89e3\u7d04\u7387",
            value=0.05,
            unit="%",
            mapped_targets=[CellTarget(sheet="KPI", cell="B2")],
            confidence=0.8,
            source="document",
            selected=True,
            adjusted_value=None,
        ),
        ExtractedParameter(
            key="growth_rate",
            label="\u6210\u9577\u7387",
            value=0.15,
            unit="%",
            mapped_targets=[CellTarget(sheet="KPI", cell="B3")],
            confidence=0.75,
            source="document",
            selected=True,
            adjusted_value=None,
        ),
    ]


@pytest.fixture
def formula_target_parameter():
    """Return a parameter that targets a formula cell (should be blocked)."""
    return ExtractedParameter(
        key="gross_profit_fy2024",
        label="\u7c97\u5229 FY2024",
        value=90000000,
        unit="\u5186",
        mapped_targets=[CellTarget(sheet="PL\u8a2d\u8a08", cell="B5")],
        confidence=0.9,
        source="document",
        selected=True,
        adjusted_value=None,
    )


@pytest.fixture
def deselected_parameter():
    """Return a parameter with selected=False (should be skipped)."""
    return ExtractedParameter(
        key="misc_cost",
        label="\u305d\u306e\u4ed6\u8cbb\u7528",
        value=5000000,
        unit="\u5186",
        mapped_targets=[CellTarget(sheet="PL\u8a2d\u8a08", cell="B6")],
        confidence=0.4,
        source="inferred",
        selected=False,
        adjusted_value=None,
    )


@pytest.fixture
def adjusted_parameter():
    """Return a parameter with an adjusted_value set."""
    return ExtractedParameter(
        key="revenue_fy2024",
        label="\u58f2\u4e0a\u9ad8 FY2024",
        value=150000000,
        unit="\u5186",
        mapped_targets=[CellTarget(sheet="PL\u8a2d\u8a08", cell="B3")],
        confidence=0.9,
        source="document",
        selected=True,
        adjusted_value=160000000,
    )


# ---------------------------------------------------------------------------
# Catalog fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def sample_catalog_items():
    """Return a list of sample CatalogItem instances."""
    return [
        CatalogItem(
            sheet="PL\u8a2d\u8a08",
            cell="B3",
            current_value=1000,
            fill_color="FFFFF2CC",
            has_formula=False,
            label_candidates=["\u58f2\u4e0a\u9ad8"],
            unit_candidates=["\u5186"],
            year_or_period="FY2024",
            block="\u58f2\u4e0a",
        ),
        CatalogItem(
            sheet="PL\u8a2d\u8a08",
            cell="B5",
            current_value="=B3-B4",
            fill_color="FFFFFFFF",
            has_formula=True,
            label_candidates=["\u7c97\u5229"],
            unit_candidates=[],
            year_or_period="FY2024",
            block="\u58f2\u4e0a",
        ),
    ]


@pytest.fixture
def sample_input_catalog(sample_catalog_items):
    """Return a sample InputCatalog instance."""
    catalog = InputCatalog(
        items=sample_catalog_items,
        blocks={"PL\u8a2d\u8a08::\u58f2\u4e0a": sample_catalog_items},
    )
    return catalog
