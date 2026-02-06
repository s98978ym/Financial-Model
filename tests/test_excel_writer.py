"""Tests for src.excel.writer -- PLWriter Excel generation.

Verifies that the writer correctly writes parameters to input cells,
preserves formulas, respects the selected flag, uses adjusted_value,
and sets fullCalcOnLoad.
"""

import shutil
from pathlib import Path

import pytest
import openpyxl
from openpyxl.styles import PatternFill

from src.config.models import (
    CellTarget,
    ColorConfig,
    Evidence,
    ExtractedParameter,
    PhaseAConfig,
)
from src.excel.writer import PLWriter


# ===================================================================
# Helpers
# ===================================================================

INPUT_FILL = PatternFill(
    start_color="FFFFF2CC",
    end_color="FFFFF2CC",
    fill_type="solid",
)


@pytest.fixture
def writer_setup(sample_template_path, tmp_path, default_phase_a_config):
    """Set up a PLWriter with the sample template and a temp output path."""
    output_path = str(tmp_path / "output.xlsx")
    writer = PLWriter(
        template_path=sample_template_path,
        output_path=output_path,
        config=default_phase_a_config,
    )
    return writer, output_path


def _make_formula_target_param():
    """Create a parameter that targets a formula cell (B5).

    Built inline to avoid the conftest fixture which passes evidence=None
    to a non-Optional Evidence field.
    """
    return ExtractedParameter(
        key="gross_profit_fy2024",
        label="\u7c97\u5229 FY2024",
        value=90000000,
        unit="\u5186",
        mapped_targets=[CellTarget(sheet="PL\u8a2d\u8a08", cell="B5")],
        evidence=Evidence(
            quote="\u7c97\u5229\u306f\u7d049\u5343\u4e07\u5186",
            page_or_slide="5",
        ),
        confidence=0.9,
        source="document",
        selected=True,
    )


# ===================================================================
# Basic generation
# ===================================================================

class TestPLWriterGenerate:
    """Test the generate method end-to-end."""

    def test_generate_creates_output_file(self, writer_setup, sample_parameters):
        writer, output_path = writer_setup
        result = writer.generate(sample_parameters)
        assert Path(result).exists()

    def test_generate_returns_output_path(self, writer_setup, sample_parameters):
        writer, output_path = writer_setup
        result = writer.generate(sample_parameters)
        assert result == output_path

    def test_generate_writes_values_to_input_cells(self, writer_setup, sample_parameters):
        writer, output_path = writer_setup
        writer.generate(sample_parameters)

        wb = openpyxl.load_workbook(output_path)
        ws = wb["PL\u8a2d\u8a08"]
        # revenue_fy2024 -> B3 should be 150000000
        assert ws["B3"].value == 150_000_000
        # cogs_fy2024 -> B4 should be 60000000
        assert ws["B4"].value == 60_000_000
        wb.close()

    def test_generate_change_log_populated(self, writer_setup, sample_parameters):
        writer, output_path = writer_setup
        writer.generate(sample_parameters)
        summary = writer.get_change_summary()
        assert summary["total_changes"] > 0


# ===================================================================
# Formula preservation
# ===================================================================

class TestFormulaPreservation:
    """Verify that formula cells are never overwritten."""

    def test_formula_cells_preserved(self, writer_setup, sample_parameters):
        """Formulas in the template should remain after generation."""
        writer, output_path = writer_setup
        writer.generate(sample_parameters)

        wb = openpyxl.load_workbook(output_path)
        ws = wb["PL\u8a2d\u8a08"]
        # B5 should still be "=B3-B4"
        assert ws["B5"].value == "=B3-B4"
        # B7 should still be "=B5-B6"
        assert ws["B7"].value == "=B5-B6"
        wb.close()

    def test_formula_target_blocked(self, writer_setup):
        """A parameter targeting a formula cell should be blocked."""
        param = _make_formula_target_param()
        writer, output_path = writer_setup
        writer.generate([param])

        wb = openpyxl.load_workbook(output_path)
        ws = wb["PL\u8a2d\u8a08"]
        # B5 should NOT be overwritten with 90000000
        assert ws["B5"].value == "=B3-B4"
        wb.close()

    def test_formula_target_in_skipped_log(self, writer_setup):
        """A parameter targeting a formula cell should appear in skipped log."""
        param = _make_formula_target_param()
        writer, output_path = writer_setup
        writer.generate([param])
        summary = writer.get_change_summary()
        assert summary["total_skipped"] > 0
        skipped_reasons = [s["reason"] for s in summary["skipped"]]
        assert "formula_cell" in skipped_reasons


# ===================================================================
# fullCalcOnLoad
# ===================================================================

class TestFullCalcOnLoad:
    """Verify that fullCalcOnLoad is set in the generated workbook."""

    def test_full_calc_on_load_set(self, writer_setup, sample_parameters):
        writer, output_path = writer_setup
        writer.generate(sample_parameters)

        wb = openpyxl.load_workbook(output_path)
        calc = wb.calculation
        assert calc is not None
        assert calc.fullCalcOnLoad is True
        wb.close()


# ===================================================================
# selected / adjusted_value
# ===================================================================

class TestSelectedAndAdjustedValue:
    """Test that selected=False params are skipped, adjusted_value is used."""

    def test_deselected_parameter_skipped(self, writer_setup, deselected_parameter):
        """Parameters with selected=False should not be written."""
        writer, output_path = writer_setup
        writer.generate([deselected_parameter])

        wb = openpyxl.load_workbook(output_path)
        ws = wb["PL\u8a2d\u8a08"]
        # B6 should still have the original template value (200)
        assert ws["B6"].value == 200
        wb.close()

    def test_adjusted_value_written(self, writer_setup, adjusted_parameter):
        """When adjusted_value is set, it should be written instead of value."""
        writer, output_path = writer_setup
        writer.generate([adjusted_parameter])

        wb = openpyxl.load_workbook(output_path)
        ws = wb["PL\u8a2d\u8a08"]
        # B3 should have adjusted_value=160000000, not value=150000000
        assert ws["B3"].value == 160_000_000
        wb.close()


# ===================================================================
# Edge cases
# ===================================================================

class TestWriterEdgeCases:
    """Test edge cases in the writer."""

    def test_missing_sheet_logged(self, writer_setup):
        """Parameter targeting a non-existent sheet should be skipped."""
        param = ExtractedParameter(
            key="nonexistent",
            label="\u5b58\u5728\u3057\u306a\u3044\u30b7\u30fc\u30c8",
            value=999,
            mapped_targets=[CellTarget(sheet="DoesNotExist", cell="A1")],
            selected=True,
        )
        writer, output_path = writer_setup
        writer.generate([param])
        summary = writer.get_change_summary()
        assert summary["total_skipped"] >= 1
        skipped_reasons = [s["reason"] for s in summary["skipped"]]
        assert "sheet_not_found" in skipped_reasons

    def test_empty_parameters_list(self, writer_setup):
        """Generating with empty parameters should still create output."""
        writer, output_path = writer_setup
        result = writer.generate([])
        assert Path(result).exists()

    def test_multiple_targets_per_parameter(self, writer_setup):
        """A single parameter with multiple targets should write to all."""
        param = ExtractedParameter(
            key="revenue",
            label="\u58f2\u4e0a\u9ad8",
            value=5000,
            mapped_targets=[
                CellTarget(sheet="PL\u8a2d\u8a08", cell="B3"),
                CellTarget(sheet="PL\u8a2d\u8a08", cell="C3"),
            ],
            selected=True,
        )
        writer, output_path = writer_setup
        writer.generate([param])

        wb = openpyxl.load_workbook(output_path)
        ws = wb["PL\u8a2d\u8a08"]
        assert ws["B3"].value == 5000
        assert ws["C3"].value == 5000
        wb.close()

    def test_output_directory_created(self, sample_template_path, default_phase_a_config):
        """Output directory should be created if it does not exist."""
        import tempfile
        with tempfile.TemporaryDirectory() as td:
            output_path = str(Path(td) / "subdir" / "output.xlsx")
            writer = PLWriter(
                template_path=sample_template_path,
                output_path=output_path,
                config=default_phase_a_config,
            )
            result = writer.generate([])
            assert Path(result).exists()


# ===================================================================
# Change summary
# ===================================================================

class TestChangeSummary:
    """Test the change summary report."""

    def test_summary_structure(self, writer_setup, sample_parameters):
        writer, output_path = writer_setup
        writer.generate(sample_parameters)
        summary = writer.get_change_summary()
        assert "total_changes" in summary
        assert "total_skipped" in summary
        assert "changes" in summary
        assert "skipped" in summary

    def test_change_entry_has_required_fields(self, writer_setup, sample_parameters):
        writer, output_path = writer_setup
        writer.generate(sample_parameters)
        summary = writer.get_change_summary()
        if summary["total_changes"] > 0:
            change = summary["changes"][0]
            assert "sheet" in change
            assert "cell" in change
            assert "old_value" in change
            assert "new_value" in change
            assert "param_key" in change
            assert "confidence" in change
