import json
from pathlib import Path

import subprocess
import sys

from openpyxl import load_workbook

from src.evals.candidate_profiles import CandidateProfile
from src.evals.pdca_loop import run_reference_pdca


def test_run_reference_pdca_selects_highest_scoring_candidate(tmp_path) -> None:
    result = run_reference_pdca(
        plan_pdf=Path("/tmp/fake.pdf"),
        reference_workbook=Path("tests/fixtures/evals/reference_workbook_minimal.xlsx"),
        artifact_root=tmp_path,
        runner="fixture",
    )

    assert result.best_candidate_id == "candidate-better"
    assert (tmp_path / result.run_id / "summary.md").exists()


def test_run_reference_pdca_writes_scores_and_summary(tmp_path) -> None:
    result = run_reference_pdca(
        plan_pdf=Path("/tmp/fake.pdf"),
        reference_workbook=Path("tests/fixtures/evals/reference_workbook_minimal.xlsx"),
        artifact_root=tmp_path,
        runner="fixture",
    )

    assert (tmp_path / result.run_id / "scores.json").exists()
    assert result.baseline_score is not None
    summary_text = (tmp_path / result.run_id / "summary.md").read_text(encoding="utf-8")
    assert "## 評価項目の説明" in summary_text
    assert "## スコア推移グラフ" in summary_text
    assert "## 施策・効果・課題" in summary_text
    assert "`structure`" in summary_text
    assert "`model_sheets`" in summary_text
    assert "`pl`" in summary_text
    assert "`explainability`" in summary_text
    assert "## 総合評価" in summary_text
    assert "## 個別評価" in summary_text
    assert "## 検証した仮説" in summary_text
    assert "## 結果" in summary_text
    assert "## 課題" in summary_text
    assert "## 改善内容" in summary_text
    assert "## 次の方針" in summary_text


def test_run_reference_pdca_writes_diagnosis_and_score_deltas(tmp_path) -> None:
    result = run_reference_pdca(
        plan_pdf=Path("/tmp/fake.pdf"),
        reference_workbook=Path("tests/fixtures/evals/reference_workbook_minimal.xlsx"),
        artifact_root=tmp_path,
        runner="fixture",
    )

    run_root = tmp_path / result.run_id
    scores = json.loads((run_root / "scores.json").read_text(encoding="utf-8"))
    diagnosis = json.loads((run_root / "diagnosis.json").read_text(encoding="utf-8"))

    candidate = next(iter(scores["candidates"].values()))
    assert "layer_deltas" in candidate
    assert "rank" in candidate
    assert diagnosis["candidates"]


def test_run_reference_pdca_writes_workbook_exports(tmp_path) -> None:
    result = run_reference_pdca(
        plan_pdf=Path("/tmp/fake.pdf"),
        reference_workbook=Path("tests/fixtures/evals/reference_workbook_minimal.xlsx"),
        artifact_root=tmp_path,
        runner="fixture",
    )

    run_root = tmp_path / result.run_id
    assert (run_root / "exports" / "baseline.xlsx").exists()
    assert (run_root / "exports" / "best-practical.xlsx").exists()
    assert result.best_practical_candidate_id == "candidate-better"
    assert (run_root / "exports" / "best-practical-candidate-better.xlsx").exists()

    workbook = load_workbook(run_root / "exports" / "best-practical.xlsx", data_only=False)
    pdca_sheet = workbook["PDCAチェックシート"]
    pdca_labels = {
        pdca_sheet.cell(row=row_index, column=1).value
        for row_index in range(1, pdca_sheet.max_row + 1)
        if pdca_sheet.cell(row=row_index, column=1).value
    }
    assert "PDCA全体推移" in pdca_labels
    trend_section_row = next(
        row_index
        for row_index in range(1, pdca_sheet.max_row + 1)
        if pdca_sheet.cell(row=row_index, column=1).value == "PDCA全体推移"
    )
    trend_header_row = trend_section_row + 1
    first_data_row = trend_section_row + 2
    trend_header = [pdca_sheet.cell(row=trend_header_row, column=column_index).value for column_index in range(1, 14)]
    assert trend_header == [
        "回",
        "候補",
        "仮説",
        "変更レバー",
        "良くなった点",
        "悪化した点",
        "structure Δ",
        "model_sheets Δ",
        "pl Δ",
        "explainability Δ",
        "成果物観点",
        "判定",
        "次の施策",
    ]
    assert pdca_sheet.cell(row=first_data_row + 1, column=2).value == "candidate-better"
    assert "fixture 改善候補の検証" in str(pdca_sheet.cell(row=first_data_row + 1, column=3).value)
    assert "fixture_better" in str(pdca_sheet.cell(row=first_data_row + 1, column=4).value)
    assert pdca_sheet.cell(row=first_data_row + 1, column=12).value == "hit"


def test_summary_contains_hypothesis_logic_evidence_and_next_actions(tmp_path) -> None:
    result = run_reference_pdca(
        plan_pdf=Path("/tmp/fake.pdf"),
        reference_workbook=Path("tests/fixtures/evals/reference_workbook_minimal.xlsx"),
        artifact_root=tmp_path,
        runner="fixture",
    )

    summary = (tmp_path / result.run_id / "summary.md").read_text(encoding="utf-8")

    assert "## 仮説内容" in summary
    assert "## 仮説検証結果" in summary
    assert "## ロジック" in summary
    assert "## 根拠ファクトとデータ" in summary
    assert "## 次の改善施策" in summary


def test_summary_mentions_workbook_exports(tmp_path) -> None:
    result = run_reference_pdca(
        plan_pdf=Path("/tmp/fake.pdf"),
        reference_workbook=Path("tests/fixtures/evals/reference_workbook_minimal.xlsx"),
        artifact_root=tmp_path,
        runner="fixture",
    )

    summary = (tmp_path / result.run_id / "summary.md").read_text(encoding="utf-8")

    assert "## Workbook Artifacts" in summary
    assert "baseline.xlsx" in summary
    assert "best-practical.xlsx" in summary


def test_fam_reference_cli_writes_artifacts(tmp_path) -> None:
    result = subprocess.run(
        [
            sys.executable,
            "-m",
            "src.cli.main",
            "eval",
            "fam-reference",
            "--plan-pdf",
            "/tmp/fake.pdf",
            "--reference-workbook",
            "tests/fixtures/evals/reference_workbook_minimal.xlsx",
            "--artifact-root",
            str(tmp_path),
            "--runner",
            "fixture",
        ],
        cwd=Path(__file__).resolve().parents[2],
        capture_output=True,
        text=True,
    )

    assert result.returncode == 0
    assert "best_candidate_id" in result.stdout
    assert "best_practical_candidate_id" in result.stdout
    assert "baseline_workbook_path" in result.stdout
    assert "best_practical_workbook_path" in result.stdout
    assert "best_practical_labeled_workbook_path" in result.stdout


def test_run_reference_pdca_uses_explicit_profiles_when_provided(tmp_path) -> None:
    result = run_reference_pdca(
        plan_pdf=Path("/tmp/fake.pdf"),
        reference_workbook=Path("tests/fixtures/evals/reference_workbook_minimal.xlsx"),
        artifact_root=tmp_path,
        runner="fixture",
        profiles=[
            CandidateProfile(
                candidate_id="candidate-baseline-like",
                label="Fixture baseline-like candidate",
                runner="fixture",
                config={"fixture_name": "baseline_result.json"},
            )
        ],
    )

    assert result.best_candidate_id == "candidate-baseline-like"


def test_run_reference_pdca_uses_explicit_baseline_mode_when_provided(tmp_path) -> None:
    result = run_reference_pdca(
        plan_pdf=Path("/tmp/fake.pdf"),
        reference_workbook=Path("tests/fixtures/evals/reference_workbook_minimal.xlsx"),
        artifact_root=tmp_path,
        runner="fixture",
        baseline_mode="candidate_result.json",
        profiles=[
            CandidateProfile(
                candidate_id="candidate-baseline-like",
                label="Fixture baseline-like candidate",
                runner="fixture",
                config={"fixture_name": "baseline_result.json"},
            )
        ],
    )

    assert result.best_candidate_score < result.baseline_score
