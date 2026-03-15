import json
from pathlib import Path

from openpyxl import load_workbook

from src.evals.diagnosis import build_candidate_diagnosis
from src.evals.candidate_profiles import CandidateProfile, fixture_profiles
from src.evals.reference_workbook import extract_reference_workbook
from src.evals.scoring import score_candidate
from src.evals.workbook_export import export_candidate_workbook


def _rgb_suffix(cell) -> str:
    return (cell.fill.fgColor.rgb or "").upper()[-6:]


def test_export_candidate_workbook_writes_expected_sheets(tmp_path) -> None:
    fixture_dir = Path("tests/fixtures/evals")
    candidate = json.loads((fixture_dir / "candidate_result.json").read_text(encoding="utf-8"))
    baseline = json.loads((fixture_dir / "baseline_result.json").read_text(encoding="utf-8"))
    reference = extract_reference_workbook(fixture_dir / "reference_workbook_minimal.xlsx")
    candidate_score = score_candidate(reference, candidate)
    baseline_score = score_candidate(reference, baseline)
    profile = next(profile for profile in fixture_profiles() if profile.candidate_id == "candidate-better")
    diagnosis = build_candidate_diagnosis(
        profile,
        candidate_score,
        baseline_score,
        evidence_summary={"pdf_facts": [], "external_sources": [], "benchmark_fills": [], "seed_notes": []},
    )

    output_path = tmp_path / "candidate.xlsx"
    export_candidate_workbook(
        output_path=output_path,
        candidate_id=profile.candidate_id,
        candidate_payload=candidate,
        diagnosis=diagnosis,
        baseline_total=baseline_score.total_score,
        run_root=tmp_path,
    )

    assert output_path.exists()
    workbook = load_workbook(output_path, data_only=False)
    assert workbook.sheetnames == [
        "PDCAチェックシート",
        "想定Q&A",
        "PL設計",
        "ミールモデル",
        "アカデミーモデル",
        "コンサルモデル",
        "費用計画",
        "（全Ver）前提条件",
    ]

    review_sheet = workbook["PDCAチェックシート"]
    assert review_sheet.freeze_panes == "B2"
    assert review_sheet.column_dimensions["A"].width >= 22
    assert review_sheet.column_dimensions["B"].width >= 18
    pdca_labels = {
        review_sheet.cell(row=row_index, column=1).value
        for row_index in range(1, review_sheet.max_row + 1)
        if review_sheet.cell(row=row_index, column=1).value
    }
    assert "今回の結論" in pdca_labels
    assert "PDCA全体推移" in pdca_labels
    assert "評価スコア" in pdca_labels
    assert "根拠と前提" in pdca_labels
    assert "関連ファイル" in pdca_labels
    assert "次の改善施策" in pdca_labels
    all_labels = {
        review_sheet.cell(row=row_index, column=1).value
        for row_index in range(1, review_sheet.max_row + 1)
    }
    assert "仮説タイトル" not in all_labels
    assert "仮説の要点" not in all_labels
    assert "今回の仮説" in all_labels
    assert "仮説の詳細" in all_labels

    pl_sheet = workbook["PL設計"]
    assert pl_sheet.freeze_panes == "B2"
    assert pl_sheet.column_dimensions["A"].width >= 18
    assert 10 <= pl_sheet.column_dimensions["B"].width <= 14
    assert pl_sheet.row_dimensions[1].height >= 20
    pl_rows = {
        pl_sheet.cell(row=row_index, column=1).value: row_index
        for row_index in range(1, pl_sheet.max_row + 1)
        if pl_sheet.cell(row=row_index, column=1).value
    }
    assert "営業利益" in pl_rows
    assert "開発費（償却）" in pl_rows
    assert isinstance(pl_sheet.cell(row=pl_rows["営業利益"], column=2).value, str)
    assert pl_sheet.cell(row=pl_rows["営業利益"], column=2).value.startswith("=")

    assumptions_sheet = workbook["（全Ver）前提条件"]
    assert assumptions_sheet.freeze_panes == "B2"
    assert assumptions_sheet.column_dimensions["A"].width >= 18
    assert assumptions_sheet.column_dimensions["G"].width >= 28
    assumption_labels = {
        assumptions_sheet.cell(row=row_index, column=1).value
        for row_index in range(1, assumptions_sheet.max_row + 1)
        if assumptions_sheet.cell(row=row_index, column=1).value
    }
    assert "売上目標" in assumption_labels
    assert "人件費比率" in assumption_labels
    assert "開発投資額（キャッシュ）" in assumption_labels
    assert "開発償却期間（年）" in assumption_labels
    assert "PL計上開発費（償却）" in assumption_labels
    assert assumptions_sheet["B2"].fill.fill_type == "solid"
    assert _rgb_suffix(assumptions_sheet["B2"]) == "DDEBF7"
    assert assumptions_sheet["B2"].number_format == "#,##0"
    assert assumptions_sheet["B10"].fill.fill_type == "solid"
    assert _rgb_suffix(assumptions_sheet["B10"]) == "E2F0D9"
    assert assumptions_sheet["B10"].number_format == "#,##0"
    assert assumptions_sheet["B28"].number_format == "0.0%"

    cost_sheet = workbook["費用計画"]
    assert cost_sheet.freeze_panes == "B2"
    assert cost_sheet.column_dimensions["A"].width >= 20
    cost_labels = {
        cost_sheet.cell(row=row_index, column=1).value
        for row_index in range(1, cost_sheet.max_row + 1)
        if cost_sheet.cell(row=row_index, column=1).value
    }
    assert "費用サマリー" in cost_labels
    assert "開発償却ブロック" in cost_labels
    assert "費用明細" in cost_labels
    assert "OPEX合計" in cost_labels
    assert "開発投資（キャッシュ）" in cost_labels
    assert "償却方法" in cost_labels
    assert "償却期間（年）" in cost_labels
    assert "当期償却額（PL計上）" in cost_labels
    assert "当期投資の期末未償却残高" in cost_labels
    assert cost_sheet["B17"].value == "FY1"
    assert cost_sheet["F17"].value == "FY5"
    assert cost_sheet["G17"].value == "カテゴリ"
    assert cost_sheet["A7"].fill.fill_type == "solid"
    assert _rgb_suffix(cost_sheet["A7"]) == "A6A6A6"
    assert cost_sheet["B7"].fill.fill_type == "solid"
    assert _rgb_suffix(cost_sheet["B7"]) == "A6A6A6"
    assert cost_sheet["B7"].number_format == "#,##0"
    assert cost_sheet["B11"].value == "定額法"
    assert cost_sheet["B12"].value == 5
    assert isinstance(cost_sheet["B10"].value, str)
    assert cost_sheet["B10"].value.startswith("=")
    assert isinstance(cost_sheet["B13"].value, str)
    assert cost_sheet["B13"].value.startswith("=")
    assert isinstance(cost_sheet["B18"].value, str)
    assert cost_sheet["B18"].value.startswith("=")
    assert cost_sheet["G18"].value == "人件費"


def test_export_candidate_workbook_adds_qa_sheet_with_iteration_tracking(tmp_path) -> None:
    fixture_dir = Path("tests/fixtures/evals")
    candidate = json.loads((fixture_dir / "candidate_result.json").read_text(encoding="utf-8"))
    baseline = json.loads((fixture_dir / "baseline_result.json").read_text(encoding="utf-8"))
    reference = extract_reference_workbook(fixture_dir / "reference_workbook_minimal.xlsx")
    candidate_score = score_candidate(reference, candidate)
    baseline_score = score_candidate(reference, baseline)
    profile = CandidateProfile(
        candidate_id="candidate-revenue-staged-sales",
        label="Revenue staged sales candidate",
        hypothesis_title="検証後アクセルに営業効率を重ねる",
        hypothesis_detail="3年間の検証後に営業投資を加速し、sales efficiency overlay を重ねると PL 再現が改善するはず。",
        toggles_on=["staged", "sales"],
        toggles_off=["partner", "branding"],
        logic_steps=[
            "前半3年を検証期間として扱う。",
            "後半で営業投資を加速させる。",
            "sales efficiency overlay を consulting/revenue 系列に反映する。",
        ],
        expected_impacts={"pl": 0.03},
        evidence_source_types=["pdf", "external"],
        next_if_success=["consulting driver と revenue line の橋渡しをさらに強化する。"],
        next_if_fail=["sales overlay の効き方を見直す。"],
    )
    diagnosis = build_candidate_diagnosis(
        profile,
        candidate_score,
        baseline_score,
        evidence_summary={
            "pdf_facts": ["3年間ユニットエコノミクスを磨く方針"],
            "external_sources": [{"title": "Sales benchmark", "quote": "高効率営業は win rate を改善する。"}],
            "benchmark_fills": [],
            "seed_notes": [],
        },
    )
    iteration_summaries = [
        {
            "iteration": 0,
            "candidate_id": "baseline",
            "hypothesis": "baseline",
            "changed_levers": "-",
            "improved_points": "-",
            "worsened_points": "-",
            "structure_delta": 0.0,
            "model_sheets_delta": 0.0,
            "pl_delta": 0.0,
            "explainability_delta": 0.0,
            "artifact_impact": "比較基準",
            "verdict": "baseline",
            "next_action": "以降の候補を比較する",
        },
        {
            "iteration": 1,
            "candidate_id": "candidate-revenue-staged-sales",
            "hypothesis": "検証後アクセルに営業効率を重ねる",
            "changed_levers": "ON: staged, sales / OFF: partner, branding",
            "improved_points": "pl +0.0478 / explainability +0.0027",
            "worsened_points": "-",
            "structure_delta": 0.0,
            "model_sheets_delta": 0.0,
            "pl_delta": 0.0478,
            "explainability_delta": 0.0027,
            "artifact_impact": "PL改善 / 説明強化",
            "verdict": "hit",
            "next_action": "consulting driver と revenue line の橋渡しをさらに強化する。",
        },
        {
            "iteration": 2,
            "candidate_id": "candidate-revenue-staged-partner",
            "hypothesis": "検証後アクセルにパートナー戦略を重ねる",
            "changed_levers": "ON: staged, partner / OFF: sales, branding",
            "improved_points": "pl +0.0417 / explainability +0.0027",
            "worsened_points": "model_sheets -0.0045",
            "structure_delta": 0.0,
            "model_sheets_delta": -0.0045,
            "pl_delta": 0.0417,
            "explainability_delta": 0.0027,
            "artifact_impact": "PL改善 / 説明強化 / 一部悪化",
            "verdict": "partial_hit",
            "next_action": "partner 寄与を sales efficiency と複合して比較する。",
        },
    ]

    output_path = tmp_path / "candidate.xlsx"
    export_candidate_workbook(
        output_path=output_path,
        candidate_id=profile.candidate_id,
        candidate_payload=candidate,
        diagnosis=diagnosis,
        baseline_total=baseline_score.total_score,
        run_root=tmp_path,
        iteration_summaries=iteration_summaries,
    )

    workbook = load_workbook(output_path, data_only=False)
    qa_sheet = workbook["想定Q&A"]
    assert qa_sheet.freeze_panes == "A2"
    headers = [qa_sheet.cell(row=1, column=column_index).value for column_index in range(1, 9)]
    assert headers == [
        "カテゴリ",
        "想定質問",
        "回答",
        "根拠",
        "初回追加Iteration",
        "今回更新Iteration",
        "状態",
        "採用状況",
    ]
    categories = {
        qa_sheet.cell(row=row_index, column=1).value
        for row_index in range(2, qa_sheet.max_row + 1)
        if qa_sheet.cell(row=row_index, column=1).value
    }
    assert {"収益", "コスト", "収益性", "成長性", "リスク", "市場", "オペレーション", "資金"}.issubset(categories)
    statuses = {
        qa_sheet.cell(row=row_index, column=7).value
        for row_index in range(2, qa_sheet.max_row + 1)
        if qa_sheet.cell(row=row_index, column=7).value
    }
    assert "新規" in statuses
    assert "継続" in statuses
    adoption = {
        qa_sheet.cell(row=row_index, column=8).value
        for row_index in range(2, qa_sheet.max_row + 1)
        if qa_sheet.cell(row=row_index, column=8).value
    }
    assert "今回採用" in adoption
    assert "比較のみ" in adoption
    revenue_rows = [
        row_index
        for row_index in range(2, qa_sheet.max_row + 1)
        if qa_sheet.cell(row=row_index, column=1).value == "収益"
    ]
    assert revenue_rows
    first_revenue_row = revenue_rows[0]
    assert qa_sheet.cell(row=first_revenue_row, column=5).value == 1
    assert qa_sheet.cell(row=first_revenue_row, column=6).value == 1


def test_export_candidate_workbook_expands_academy_and_consulting_structure(tmp_path) -> None:
    fixture_dir = Path("tests/fixtures/evals")
    candidate = json.loads((fixture_dir / "candidate_result.json").read_text(encoding="utf-8"))
    baseline = json.loads((fixture_dir / "baseline_result.json").read_text(encoding="utf-8"))
    reference = extract_reference_workbook(fixture_dir / "reference_workbook_minimal.xlsx")
    candidate_score = score_candidate(reference, candidate)
    baseline_score = score_candidate(reference, baseline)
    profile = next(profile for profile in fixture_profiles() if profile.candidate_id == "candidate-better")
    diagnosis = build_candidate_diagnosis(
        profile,
        candidate_score,
        baseline_score,
        evidence_summary={"pdf_facts": [], "external_sources": [], "benchmark_fills": [], "seed_notes": []},
    )

    output_path = tmp_path / "candidate.xlsx"
    export_candidate_workbook(
        output_path=output_path,
        candidate_id=profile.candidate_id,
        candidate_payload=candidate,
        diagnosis=diagnosis,
        baseline_total=baseline_score.total_score,
        run_root=tmp_path,
    )

    workbook = load_workbook(output_path, data_only=False)

    academy_sheet = workbook["アカデミーモデル"]
    assert academy_sheet.freeze_panes == "B2"
    assert academy_sheet.column_dimensions["A"].width >= 18
    academy_labels = {
        academy_sheet.cell(row=row_index, column=1).value
        for row_index in range(1, academy_sheet.max_row + 1)
        if academy_sheet.cell(row=row_index, column=1).value
    }
    assert {"C級課程", "B級課程", "A級課程", "S級課程"}.issubset(academy_labels)

    academy_formula_cells = [
        academy_sheet["B5"].value,
        academy_sheet["B10"].value,
        academy_sheet["B15"].value,
        academy_sheet["B20"].value,
    ]
    assert all(isinstance(value, str) and value.startswith("=") for value in academy_formula_cells)

    consult_sheet = workbook["コンサルモデル"]
    assert consult_sheet.freeze_panes == "A3"
    assert consult_sheet.column_dimensions["B"].width >= 20
    assert consult_sheet.column_dimensions["H"].width >= 12
    consult_headers = [consult_sheet.cell(row=2, column=column_index).value for column_index in range(1, 8)]
    assert consult_headers == [
        "SKU",
        "サービス名",
        "単位",
        "単価（円）",
        "継続率",
        "デリバリー原価単価",
        "標準時間",
    ]

    sku_rows = {
        consult_sheet.cell(row=row_index, column=1).value: row_index
        for row_index in range(1, consult_sheet.max_row + 1)
        if consult_sheet.cell(row=row_index, column=1).value
    }
    assert {"P1", "P2", "P3", "P4", "P5", "P6", "P8", "P9", "P10", "P11", "P12"}.issubset(sku_rows)
    assert isinstance(consult_sheet["H3"].value, str) and consult_sheet["H3"].value.startswith("=")
    assert isinstance(consult_sheet["I3"].value, str) and consult_sheet["I3"].value.startswith("=")
    assert consult_sheet["M15"].fill.fill_type == "solid"
    assert _rgb_suffix(consult_sheet["M15"]) == "D9E2F3"
    assert consult_sheet["D3"].number_format == "#,##0"
    assert consult_sheet["E3"].number_format == "0.0%"
    assert consult_sheet["H3"].number_format == "#,##0.0"
    assert consult_sheet["B3"].alignment.wrap_text is True

    assumptions_sheet = workbook["（全Ver）前提条件"]
    assumption_labels = {
        assumptions_sheet.cell(row=row_index, column=1).value
        for row_index in range(1, assumptions_sheet.max_row + 1)
        if assumptions_sheet.cell(row=row_index, column=1).value
    }
    assert "C級新規構成比" in assumption_labels
    assert "C→B進級率" in assumption_labels
    assert "P1売上構成比" in assumption_labels
    assert "ブレンド時給" in assumption_labels
    assert _rgb_suffix(assumptions_sheet["B33"]) == "DDEBF7"
    assert _rgb_suffix(assumptions_sheet["B58"]) == "DDEBF7"

    pl_sheet = workbook["PL設計"]
    assert pl_sheet["A14"].fill.fill_type == "solid"
    assert _rgb_suffix(pl_sheet["A14"]) == "A6A6A6"
    assert pl_sheet["B14"].fill.fill_type == "solid"
    assert _rgb_suffix(pl_sheet["B14"]) == "A6A6A6"
    assert pl_sheet["A3"].fill.fill_type == "solid"
    assert _rgb_suffix(pl_sheet["A3"]) == "D9E2F3"
    assert pl_sheet["B2"].number_format == "#,##0"
    assert pl_sheet["B8"].number_format == "0.0%"
    assert pl_sheet["B15"].number_format == "0.0%"
    assert pl_sheet["A3"].alignment.horizontal == "left"
