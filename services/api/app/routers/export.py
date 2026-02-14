"""Excel export endpoints — generate and download."""

from __future__ import annotations

import io
import logging
import os
import tempfile
import threading

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse

from .. import db
from .jobs import create_job

logger = logging.getLogger(__name__)
router = APIRouter()

_CELERY_ENABLED = bool(os.environ.get("REDIS_URL"))

# In-memory file store for local dev (when no object storage)
_file_store: dict[str, bytes] = {}


@router.post("/export/excel", status_code=202)
async def export_excel(body: dict):
    """Generate Excel file(s) for download (async job).

    Creates Best/Base/Worst scenario files + needs_review.csv.
    """
    project_id = body.get("project_id")
    if not project_id:
        raise HTTPException(
            status_code=422,
            detail={"code": "VALIDATION_ERROR", "message": "project_id required"},
        )

    run = db.get_latest_run(project_id)
    if not run:
        run = db.create_run(project_id)

    job = create_job(phase=6, run_id=run["id"], payload={
        "project_id": project_id,
        "parameters": body.get("parameters", {}),
        "scenarios": body.get("scenarios", ["base", "best", "worst"]),
        "options": body.get("options", {}),
    })

    if _CELERY_ENABLED:
        try:
            from services.worker.celery_app import app as celery_app
            celery_app.send_task("tasks.export.generate_excel", args=[job["id"]])
            logger.info("Dispatched export task for job %s", job["id"])
        except Exception as e:
            logger.exception("Celery dispatch failed for export (job %s), falling back to sync", job["id"])
            _dispatch_export_sync(job["id"], run["id"], body)
    else:
        _dispatch_export_sync(job["id"], run["id"], body)

    return {
        "job_id": job["id"],
        "status": "queued",
        "phase": 6,
        "poll_url": f"/v1/jobs/{job['id']}",
        "download_url": f"/v1/export/download/{job['id']}",
    }


@router.get("/export/download/{job_id}")
async def download_excel(job_id: str):
    """Download generated Excel file.

    Returns the .xlsx file as a streaming response.
    """
    job = db.get_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail={"code": "JOB_NOT_FOUND"})

    if job["status"] != "completed":
        raise HTTPException(
            status_code=409,
            detail={
                "code": "NOT_READY",
                "message": f"Job status is '{job['status']}', wait for completion",
                "status": job["status"],
            },
        )

    # Try to load from file store
    file_bytes = None

    # Check in-memory store
    if job_id in _file_store:
        file_bytes = _file_store[job_id]

    if not file_bytes:
        raise HTTPException(
            status_code=404,
            detail={"code": "FILE_NOT_FOUND", "message": "Excel file no longer available"},
        )

    project_name = "pl_model"
    # Try to get project name for filename
    if job.get("run_id"):
        run = None
        # Get project info from run
        for p in db.list_projects():
            r = db.get_latest_run(p["id"])
            if r and r["id"] == job["run_id"]:
                project_name = p.get("name", "pl_model").replace(" ", "_")[:30]
                break

    filename = f"{project_name}_base.xlsx"

    # Use RFC 5987 encoding for non-ASCII filenames
    from urllib.parse import quote
    encoded_filename = quote(filename)
    disposition = f"attachment; filename*=UTF-8''{encoded_filename}"

    return StreamingResponse(
        io.BytesIO(file_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": disposition},
    )


def _dispatch_export_sync(job_id: str, run_id: str, body: dict):
    """Run Excel generation in a background thread (non-blocking)."""
    def _run():
        try:
            _generate_local_excel(job_id, run_id, body)
        except Exception as e:
            logger.exception("Sync export failed for job %s", job_id)
            db.update_job(job_id, status="failed", error_msg=str(e))

    thread = threading.Thread(target=_run, daemon=True)
    thread.start()


def _generate_local_excel(job_id: str, run_id: str, body: dict):
    """Generate Excel file locally (no Celery) for development."""
    try:
        db.update_job(job_id, status="running", progress=10, log_msg="Starting local generation")

        # Try real PLWriter first
        try:
            from src.excel.writer import PLWriter
            from src.excel.case_generator import CaseGenerator
            from src.config.models import ExtractedParameter, CellTarget

            # Load Phase 5 results
            phase5 = db.get_phase_result(run_id, 5)
            if phase5 and phase5.get("raw_json", {}).get("extractions"):
                extractions = phase5["raw_json"]["extractions"]
                params = []
                for ext in extractions:
                    targets = [CellTarget(sheet=t["sheet"], cell=t["cell"])
                               for t in ext.get("mapped_targets", [])]
                    params.append(ExtractedParameter(
                        key=ext.get("key", ""),
                        label=ext.get("label", ""),
                        value=ext.get("value"),
                        unit=ext.get("unit", ""),
                        mapped_targets=targets,
                        confidence=ext.get("confidence", 0.5),
                        source=ext.get("source", "document"),
                        selected=True,
                    ))

                # Find template
                from services.api.app.routers.phases import _resolve_template
                template_path = _resolve_template("v2_ib_grade")

                writer = PLWriter(template_path=template_path)
                output_path = writer.generate(params, scenario="base")

                with open(output_path, "rb") as f:
                    _file_store[job_id] = f.read()

                # Save phase_results for phase 6 and link via result_ref (UUID FK)
                pr = db.save_phase_result(run_id, phase=6, raw_json={
                    "type": "excel_export_plwriter",
                    "output_path": output_path,
                })
                db.update_job(
                    job_id, status="completed", progress=100,
                    log_msg="Excel generated", result_ref=pr["id"],
                )
                return
        except ImportError:
            logger.info("PLWriter not available, using openpyxl fallback")
        except Exception as e:
            logger.warning("PLWriter failed: %s — falling back", e)

        # Fallback: create full v2 workbook with computed PL data
        from .recalc import _compute_pl, _apply_scenario_multipliers

        db.update_job(job_id, status="running", progress=30, log_msg="Building v2 workbook")

        parameters = body.get("parameters", {})
        options = body.get("options", {})

        # Determine segment count from Phase 2/4 results using user's selected proposal
        num_segments = 3  # default
        segment_names = []
        try:
            phase2 = db.get_phase_result(run_id, 2)
            if phase2 and phase2.get("raw_json"):
                p2 = phase2["raw_json"]
                proposals = p2.get("proposals", [])
                if proposals:
                    # Respect user's Phase 2 selection
                    selected_idx = 0
                    try:
                        phase2_edits = db.get_edits(run_id, phase=2)
                        for ed in reversed(phase2_edits):
                            pj = ed.get("patch_json", {})
                            if "selected_proposal_index" in pj:
                                selected_idx = pj["selected_proposal_index"]
                                break
                    except Exception:
                        logger.debug("Could not load Phase 2 edits for proposal selection")
                    chosen = proposals[selected_idx] if selected_idx < len(proposals) else proposals[0]
                    segs = chosen.get("segments", [])
                    if segs:
                        num_segments = max(1, min(len(segs), 10))
                        segment_names = [
                            s.get("name", "") if isinstance(s, dict) else str(s)
                            for s in segs
                        ]
        except Exception as e:
            logger.debug("Could not load Phase 2 segments, using default %d: %s", num_segments, e)

        # Determine extra sheets from Phase 3 adopted proposals
        extra_sheets = []
        try:
            phase3_edits = db.get_edits(run_id, phase=3)
            for ed in reversed(phase3_edits):
                pj = ed.get("patch_json", {})
                adopted = pj.get("adopted", [])
                if adopted:
                    for dec in adopted:
                        text = (dec.get("proposalText", "") + " " + dec.get("selectedOption", "")).lower()
                        if any(k in text for k in ["人員", "headcount", "fte", "採用"]):
                            extra_sheets.append("headcount")
                        if any(k in text for k in ["kpi", "ダッシュボード", "指標", "モニタリング"]):
                            extra_sheets.append("kpi_dashboard")
                        if any(k in text for k in ["リスク", "感応度", "sensitivity", "シナリオ分析"]):
                            extra_sheets.append("sensitivity")
                    break
        except Exception as e:
            logger.debug("Could not load Phase 3 decisions for extra sheets: %s", e)

        # Load Phase 5 parameters if available
        try:
            phase5 = db.get_phase_result(run_id, 5)
            if phase5 and phase5.get("raw_json"):
                from .recalc import _extract_params_from_phase5
                p5_params = _extract_params_from_phase5(phase5["raw_json"])
                parameters = {**p5_params, **parameters}
        except Exception as e:
            logger.debug("Could not load Phase 5 parameters: %s", e)

        # Load user's scenario parameter edits from DB (phase 6 edits)
        try:
            scenario_edits = db.get_edits(run_id, phase=6)
            for ed in reversed(scenario_edits):
                pj = ed.get("patch_json", {})
                user_params = pj.get("parameters", {})
                if user_params:
                    parameters = {**parameters, **user_params}
                    break
        except Exception as e:
            logger.debug("Could not load scenario parameter edits: %s", e)

        # Apply scenario multipliers from export options
        best_mult = options.get("best_multipliers", {"revenue": 1.2, "cost": 0.9})
        worst_mult = options.get("worst_multipliers", {"revenue": 0.8, "cost": 1.15})
        scenario = body.get("scenario", "base")
        parameters = _apply_scenario_multipliers(parameters, scenario, best_mult, worst_mult)

        db.update_job(job_id, status="running", progress=50, log_msg="Computing PL data")

        # Compute PL from driver parameters
        result = _compute_pl(parameters)
        pl = result["pl_summary"]

        # SGA/R&D mode: "inline" (default) or "separate"
        sga_rd_mode = options.get("sga_rd_mode", "inline")
        if sga_rd_mode not in ("inline", "separate"):
            sga_rd_mode = "inline"

        # Create full v2 workbook
        from src.excel.template_v2 import (
            create_v2_workbook, PL_ROWS, FY_COLS, SEG_ROWS, SEG_STREAM_ROWS_PER,
            MAX_STREAMS, YELLOW_FILL, NUMBER_FONT, NUMBER_FMT, PERCENT_FMT,
            SGA_ROWS, RD_ROWS,
        )
        from openpyxl.styles import Alignment

        wb = create_v2_workbook(
            num_segments=num_segments,
            extra_sheets=extra_sheets,
            sga_rd_mode=sga_rd_mode,
        )

        db.update_job(job_id, status="running", progress=70, log_msg="Populating data")

        # --- Populate PL設計 sheet with driver-computed values ---
        ws_pl = wb["PL設計"]

        # Override revenue formula with computed values (driven by revenue_fy1 + growth_rate)
        for i, col in enumerate(FY_COLS):
            c = ws_pl.cell(row=PL_ROWS["revenue"], column=col)
            c.value = pl["revenue"][i]
            c.fill = YELLOW_FILL
            c.font = NUMBER_FONT
            c.number_format = NUMBER_FMT
            c.alignment = Alignment(horizontal="right")

        # Set COGS rate (driven by cogs_rate driver)
        cogs_rate = float(parameters.get("cogs_rate", 0.3))
        for i, col in enumerate(FY_COLS):
            ws_pl.cell(row=PL_ROWS["cogs_rate"], column=col).value = cogs_rate

        # Split OPEX into components (driven by opex_base + opex_growth)
        opex_list = pl["opex"]

        if sga_rd_mode == "separate":
            # Populate detail sheets: 販管費明細 and 開発費明細
            ws_sga = wb["販管費明細"]
            ws_rd = wb["開発費明細"]

            for i, col in enumerate(FY_COLS):
                ox = opex_list[i]
                # SGA breakdown (88% of OPEX): payroll 45%, marketing 20%, office 15%, other 8%
                ws_sga.cell(row=SGA_ROWS["payroll"], column=col).value = round(ox * 0.45)
                ws_sga.cell(row=SGA_ROWS["marketing"], column=col).value = round(ox * 0.20)
                ws_sga.cell(row=SGA_ROWS["office"], column=col).value = round(ox * 0.15)
                ws_sga.cell(row=SGA_ROWS["other"], column=col).value = round(ox * 0.08)

                # R&D breakdown (12% of OPEX)
                rd_total = round(ox * 0.12)
                ws_rd.cell(row=RD_ROWS["internal"], column=col).value = round(rd_total * 0.50)
                ws_rd.cell(row=RD_ROWS["outsource"], column=col).value = round(rd_total * 0.25)
                ws_rd.cell(row=RD_ROWS["cloud_infra"], column=col).value = round(rd_total * 0.20)
                ws_rd.cell(row=RD_ROWS["other"], column=col).value = round(rd_total * 0.05)
            # PL rows 12-13 are formulas referencing detail sheets (already set by template)
        else:
            # Inline mode: write OPEX directly to PL sheet
            for i, col in enumerate(FY_COLS):
                ox = opex_list[i]
                ws_pl.cell(row=PL_ROWS["payroll"], column=col).value = round(ox * 0.45)
                ws_pl.cell(row=PL_ROWS["marketing"], column=col).value = round(ox * 0.20)
                ws_pl.cell(row=PL_ROWS["office"], column=col).value = round(ox * 0.15)
                ws_pl.cell(row=PL_ROWS["system"], column=col).value = round(ox * 0.12)
                ws_pl.cell(row=PL_ROWS["other_opex"], column=col).value = round(ox * 0.08)

        # --- Populate segment sheets with revenue breakdown ---
        # Use largest-remainder method to avoid rounding errors (sum must equal total)
        def _distribute(total: int, n: int) -> list[int]:
            base = total // n
            remainder = total - base * n
            return [base + (1 if i < remainder else 0) for i in range(n)]

        for seg_idx in range(1, num_segments + 1):
            sheet_name = f"セグメント{seg_idx}モデル"
            if sheet_name not in wb.sheetnames:
                continue
            ws_seg = wb[sheet_name]

            # Set segment name if available
            if seg_idx <= len(segment_names) and segment_names[seg_idx - 1]:
                ws_seg.cell(row=2, column=1).value = f"セグメント: {segment_names[seg_idx - 1]}"

            # Put all revenue in stream 1 as a simple model
            stream1_base = SEG_ROWS["stream_start"]
            ws_seg.cell(row=stream1_base + 1, column=2).value = (
                segment_names[seg_idx - 1] if seg_idx <= len(segment_names) and segment_names[seg_idx - 1]
                else f"セグメント{seg_idx}"
            )
            for i, col in enumerate(FY_COLS):
                # Distribute revenue across segments without rounding loss
                seg_rev = _distribute(pl["revenue"][i], num_segments)[seg_idx - 1]
                # Formula: 単価 × 数量 × 頻度 × 12 = seg_rev
                # Set: 単価 = seg_rev, 数量 = 1, 頻度 = 1/12
                # → seg_rev * 1 * (1/12) * 12 = seg_rev (exact)
                ws_seg.cell(row=stream1_base + 2, column=col).value = seg_rev
                ws_seg.cell(row=stream1_base + 3, column=col).value = 1
                ws_seg.cell(row=stream1_base + 4, column=col).value = round(1 / 12, 10)

        db.update_job(job_id, status="running", progress=90, log_msg="Saving workbook")

        buf = io.BytesIO()
        wb.save(buf)
        _file_store[job_id] = buf.getvalue()

        # Save phase_results for phase 6 and link via result_ref (UUID FK)
        pr = db.save_phase_result(run_id, phase=6, raw_json={
            "type": "excel_export",
            "scenario": scenario,
            "num_segments": num_segments,
            "sga_rd_mode": sga_rd_mode,
        })
        db.update_job(
            job_id, status="completed", progress=100,
            log_msg="Excel generated (v2 workbook)",
            result_ref=pr["id"],
        )

    except Exception as e:
        logger.exception("Local Excel generation failed")
        db.update_job(job_id, status="failed", error_msg=str(e))
